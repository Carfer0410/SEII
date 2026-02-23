import io
import os
import sys

from openpyxl import load_workbook

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)

import app


def _normalize_family_code(value):
    text = str(value or "").strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def main():
    client = app.app.test_client()
    res = client.get(
        "/reports/accounting_monthly_excel?month=2&year=2026"
        "&report_title=CHECK_CONCILIACION&generated_by=VALIDADOR"
    )
    if res.status_code != 200:
        print(f"ERROR: endpoint devolvio {res.status_code}")
        try:
            print(res.get_json())
        except Exception:
            print(res.get_data(as_text=True)[:1000])
        return 1

    wb = load_workbook(io.BytesIO(res.data), data_only=False)
    ws_des = wb["DESGLOSE"]
    ws_inf = wb["INFORME"]

    excluded = {"1114001", "4619001"}
    required = {"6502001"}
    codes_in_des = set()
    detail_rows = 0

    for r in range(2, ws_des.max_row + 1):
        fam_code = _normalize_family_code(ws_des.cell(r, 8).value)
        if fam_code:
            codes_in_des.add(fam_code)
            detail_rows += 1

    missing_required = sorted(code for code in required if code not in codes_in_des)
    present_excluded = sorted(code for code in excluded if code in codes_in_des)
    subtotal_formula = str(ws_inf.cell(31, 5).value or "")

    problems = []
    if missing_required:
        problems.append(f"Familias requeridas ausentes en DESGLOSE: {', '.join(missing_required)}")
    if present_excluded:
        problems.append(f"Familias excluidas presentes en DESGLOSE: {', '.join(present_excluded)}")
    if "E30" in subtotal_formula.upper():
        problems.append("El SUBTOTAL de INFORME (E31) no debe incluir E30")
    if str(ws_inf.cell(30, 3).value or "").strip():
        problems.append("La fila 30 del INFORME debe quedar vacia")
    for row_num in [34, 35, 36, 38, 41]:
        if str(ws_inf.cell(row_num, 5).value or "").strip():
            problems.append(f"La celda E{row_num} debe quedar vacia")

    if problems:
        print("INTEGRITY CHECK FAILED")
        for p in problems:
            print("-", p)
        return 1

    print("INTEGRITY CHECK PASSED")
    print(f"- Filas de detalle DESGLOSE detectadas: {detail_rows}")
    print(f"- Formula SUBTOTAL E31: {subtotal_formula}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
