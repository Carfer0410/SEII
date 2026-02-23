# Helper para limpiar nombres de archivo
def clean_filename(text):
    import unicodedata
    text = str(text or '').lower().replace(' ', '_')
    text = unicodedata.normalize('NFD', text)
    text = ''.join(ch for ch in text if unicodedata.category(ch) != 'Mn')
    text = ''.join(c for c in text if c.isalnum() or c in ('_', '-', '.'))
    return text
import os
import zipfile
import unicodedata
import json
from threading import Lock
from functools import lru_cache
from copy import copy
from datetime import datetime, timezone
from io import BytesIO
from zoneinfo import ZoneInfo

# Intentamos resoluciones más robustas de zona horaria en entornos Windows
def _resolve_zoneinfo(name):
    try:
        return ZoneInfo(name)
    except Exception:
        try:
            # python-dateutil normalmente está disponible (dependencia de pandas)
            from dateutil import tz as dateutil_tz

            tzinfo = dateutil_tz.gettz(name)
            if tzinfo:
                return tzinfo
        except Exception:
            pass
        return None


import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, Protection
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from PIL import Image as PILImage
from flask import Flask, render_template, request, jsonify, send_file, has_app_context
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image as RLImage
from reportlab.lib.utils import ImageReader
from reportlab.graphics.shapes import Drawing, String
from reportlab.graphics.charts.barcharts import HorizontalBarChart
from reportlab.graphics.charts.piecharts import Pie
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from sqlalchemy import UniqueConstraint, text
from xml.sax.saxutils import escape

BASE_DIR = os.path.dirname(__file__)
APP_TIMEZONE = os.getenv('APP_TIMEZONE', 'America/Bogota')
LOCAL_TZ = _resolve_zoneinfo(APP_TIMEZONE) or timezone.utc
DB_PATH = os.path.join(BASE_DIR, 'assets.db')
REPORTS_DIR = os.path.join(BASE_DIR, 'generated_reports')
TEMPLATE_A22_PATH = os.path.join(BASE_DIR, 'formato a22.xlsx')
ACCOUNTING_TEMPLATE_CANDIDATES = [
    os.path.join(BASE_DIR, 'INFORME CONTABILIDAD REF.xlsx'),
    os.path.join(BASE_DIR, 'INFORME CONTABILIDAD REFERENCIA.xlsx'),
]
FAMILY_CATALOG_PATH = os.path.join(BASE_DIR, 'FAMILIA DE ACTIVOS FIJOS.xlsx')
A22_LOGO_CANDIDATES = [
    os.path.join(BASE_DIR, 'logo_a22.png'),
    os.path.join(BASE_DIR, 'logo_a22.jpg'),
    os.path.join(BASE_DIR, 'logo_a22.jpeg'),
    os.path.join(BASE_DIR, 'logo.png'),
    os.path.join(BASE_DIR, 'logo.jpg'),
]
CODIFICACION_CANDIDATES = [
    os.path.join(BASE_DIR, 'codificacion.png'),
    os.path.join(BASE_DIR, 'codificacion.jpg'),
    os.path.join(BASE_DIR, 'codificacion.jpeg'),
]
TEMPLATES_DIR = os.path.join(BASE_DIR, 'templates')
STATIC_DIR = os.path.join(BASE_DIR, 'static')
DISPOSAL_TYPE_KEYS = ['BIOMEDICO', 'MUEBLE Y ENSER', 'INDUSTRIAL', 'TECNOLOGICO', 'CONTROL']
ACCOUNTING_FAMILY_ORDER = """
5504
5504001
5504002
5504003
5506
5506001
5511
5511001
5523
5523001
6002
6002001
6002002
6002003
6002004
6002005
6003
6003001
6003002
6005
6005001
6006
6006001
6006002
6006003
6006004
6006005
6006006
6006007
6006008
6006009
6006010
6006011
6007
6007001
6007002
6007003
6007004
6007005
6007006
6007007
6007008
6007009
6007010
6007011
6007012
6007013
6007015
6008
6008001
6008002
6008003
6008004
6008005
6008006
6008007
6501
6501001
6501002
6501003
6501004
6501005
6501006
6501007
6501008
6501009
6501010
6501011
6501012
6501013
6501014
6501015
6501016
6501017
6501018
6501019
6501020
6501021
6501022
6501023
6501024
6501025
6501026
6501027
6501028
6501029
6501030
6501031
6501032
6501033
6501034
6501035
6501036
6501037
6501038
6502
6502002
6502003
6502004
7001
7001001
7001002
7001003
7001004
7001005
7002
7002001
7002002
7002003
7002004
7002005
7002006
7003
7003001
7003002
7003003
7003004
7003005
7003006
7003007
7003008
7004
7004001
7502
7502001
7506
7506001
8002
8002001
8002002
8002003
8002004
8004
8004001
8004002
""".split()

ACCOUNTING_CACHE_LOCK = Lock()
BASE_DATA_VERSION = 0
ACCOUNTING_REPORT_CACHE = {
    'version': None,
    'algo_version': None,
    'bytes': None,
    'filename': None,
}


def now_local_dt():
    return datetime.now(LOCAL_TZ)


def now_iso():
    return now_local_dt().isoformat()


def parse_dt(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        dt = value
    else:
        raw = str(value).strip()
        if not raw:
            return None
        try:
            dt = datetime.fromisoformat(raw.replace('Z', '+00:00'))
        except Exception:
            return None
    # Si el datetime no tiene tzinfo, asumir que fue ingresado en la zona local
    # (APP_TIMEZONE). Esto evita interpretar horas locales como UTC y desplazar
    # la hora al convertir.
    if dt.tzinfo is None:
        try:
            dt = dt.replace(tzinfo=LOCAL_TZ)
        except Exception:
            dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(LOCAL_TZ)


def format_dt_local(value, fmt='%Y-%m-%d %H:%M'):
    dt = parse_dt(value)
    if not dt:
        return ''
    return dt.strftime(fmt)
ACCOUNTING_REPORT_ALGO_VERSION = 12
ACCOUNTING_EXCLUDED_FAMILIES = {'1114001', '4619001'}
STRICT_ACCOUNTING_VALIDATION = True
MONTH_LABELS_ES = {
    1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril', 5: 'mayo', 6: 'junio',
    7: 'julio', 8: 'agosto', 9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
}
EXCLUDED_SERVICE_NAMES = {
    'HOSPITAL FRANCISCO DE PAULA',
}
ISSUE_STATUSES = ['Nuevo', 'En analisis', 'Escalado', 'Cerrado']
ISSUE_SEVERITIES = ['Alta', 'Media', 'Baja']
ISSUE_TYPE_LABELS = {
    'NOT_FOUND_CRITICAL': 'No encontrado critico',
    'NOT_FOUND_HIGH_VALUE': 'No encontrado de alto valor',
    'SCANNED_OTHER_SERVICE': 'Escaneado en servicio distinto',
    'RESPONSIBLE_REVIEW': 'Revision de responsable',
    'LOCATION_REVIEW': 'Revision de ubicacion',
    'DUPLICATE_CODE': 'Duplicidad probable de codigo',
    'MISSING_SERIAL_REF': 'Sin serial/referencia',
    'MISSING_MODEL_BRAND': 'Falta marca/modelo',
    'MISSING_CUSTODY_DATA': 'Falta responsable/ubicacion',
    'PENDING_UNSCANNED': 'Activo pendiente sin escaneo',
    'INVALID_FINANCIAL_VALUES': 'Valores financieros inconsistentes',
    'DEPRECIATION_INCONSISTENT': 'Depreciacion/vida util inconsistente',
    'CANDIDATE_DISPOSAL': 'Riesgo por baja pendiente',
}
ACCOUNTING_REPORT_STRUCTURE = [
    {
        'parent_code': '1655',
        'parent_name': 'MAQUINARIA Y EQUIPO',
        'children': [
            {'report_code': '165504', 'name': 'MAQUINARIA INDUSTRIAL', 'source_prefix': '5504'},
            {'report_code': '165506', 'name': 'EQ DE RECREACION Y DEPORTES', 'source_prefix': '5506'},
            {'report_code': '165511', 'name': 'HERRAMIENTAS Y ACCESORIOS', 'source_prefix': '5511'},
            {'report_code': '165523', 'name': 'EQUIPO DE ASEO', 'source_prefix': '5523'},
        ],
    },
    {
        'parent_code': '1660',
        'parent_name': 'EQUIPO MEDICO Y CIENTIFICO',
        'children': [
            {'report_code': '166002', 'name': 'EQUIPO DE LABORATORIO', 'source_prefix': '6002'},
            {'report_code': '166003', 'name': 'EQUIPO DE URGENCIAS', 'source_prefix': '6003'},
            {'report_code': '166005', 'name': 'EQUIPO DE HOSPITALIZACION', 'source_prefix': '6005'},
            {'report_code': '166006', 'name': 'EQUIPO DE CX Y SALA DE PARTOS', 'source_prefix': '6006'},
            {'report_code': '166007', 'name': 'EQUIPO DE APOYO DIAGNOSTICO', 'source_prefix': '6007'},
            {'report_code': '166008', 'name': 'EQUIPO DE APOYO TERAPEUTICO', 'source_prefix': '6008'},
        ],
    },
    {
        'parent_code': '1665',
        'parent_name': 'MUEBLES, ENSERES Y EQUIPOS DE OFICINA',
        'children': [
            {'report_code': '166501', 'name': 'MUEBLES Y ENSERES', 'source_prefix': '6501'},
            {'report_code': '166502', 'name': 'EQUIPOS Y MAQUINAS DE OFICINA', 'source_prefix': '6502'},
        ],
    },
    {
        'parent_code': '1670',
        'parent_name': 'EQUIPO DE COMUNICACION Y COMPUTACION',
        'children': [
            {'report_code': '167001', 'name': 'EQUIPO DE COMUNICACION', 'source_prefixes': ['7001', '7004']},
            {'report_code': '167002', 'name': 'EQUIPO DE COMPUTACION', 'source_prefixes': ['7002', '7003']},
        ],
    },
    {
        'parent_code': '1675',
        'parent_name': 'EQUIPO DE TRANSPORTE Y TRACCION',
        'children': [
            {'report_code': '1675002', 'name': 'EQUIPO TERRESTRE', 'source_prefix': '7502'},
            {'report_code': '167506', 'name': 'EQUIPO DE TRACCION', 'source_prefix': '7506'},
        ],
    },
    {
        'parent_code': '1680',
        'parent_name': 'EQUIPO COMEDOR, DESPENSA Y COCINA',
        'children': [
            {'report_code': '168002', 'name': 'EQUIPO DE RESTAURANTE Y CAFETERIA', 'source_prefix': '8002'},
            {'report_code': '168004', 'name': 'EQUIPO DE LAVANDERIA', 'source_prefix': '8004'},
        ],
    },
]

app = Flask(__name__, template_folder=TEMPLATES_DIR, static_folder=STATIC_DIR)
CORS(app)
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_PATH}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)


class Asset(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    c_act = db.Column(db.String, unique=True, nullable=False)
    nom = db.Column(db.String)
    modelo = db.Column(db.String)
    ref = db.Column(db.String)
    serie = db.Column(db.String)
    nom_marca = db.Column(db.String)
    c_fam = db.Column(db.String)
    nom_fam = db.Column(db.String)
    c_tiac = db.Column(db.String)
    desc_tiac = db.Column(db.String)
    desc_subtiac = db.Column(db.String)
    deprecia = db.Column(db.String)
    vida_util = db.Column(db.String)
    tipo_activo_cache = db.Column(db.String)
    des_ubi = db.Column(db.String)
    nom_ccos = db.Column(db.String)
    nom_resp = db.Column(db.String)
    est = db.Column(db.String)
    costo = db.Column(db.Float)
    saldo = db.Column(db.Float)
    fecha_compra = db.Column(db.String)

    # campos de inventario
    estado_inventario = db.Column(db.String, default='No verificado')
    fecha_verificacion = db.Column(db.String)
    usuario_verificador = db.Column(db.String)
    observacion_inventario = db.Column(db.String)
    raw_row_json = db.Column(db.Text)

    def to_dict(self):
        return {
            'id': self.id,
            'C_ACT': self.c_act,
            'NOM': self.nom,
            'MODELO': self.modelo,
            'REF': self.ref,
            'SERIE': self.serie,
            'NOM_MARCA': self.nom_marca,
            'C_FAM': self.c_fam,
            'NOM_FAM': self.nom_fam,
            'C_TIAC': self.c_tiac,
            'DESC_TIAC': self.desc_tiac,
            'DES_SUBTIAC': self.desc_subtiac,
            'DEPRECIA': self.deprecia,
            'VIDA_UTIL': self.vida_util,
            'TIPO_ACTIVO': self.tipo_activo_cache,
            'DES_UBI': self.des_ubi,
            'NOM_CCOS': self.nom_ccos,
            'NOM_RESP': self.nom_resp,
            'EST': self.est,
            'COSTO': self.costo,
            'SALDO': self.saldo,
            'FECHA_COMPRA': self.fecha_compra,
            'estado_inventario': self.estado_inventario,
            'fecha_verificacion': self.fecha_verificacion,
            'usuario_verificador': self.usuario_verificador,
            'observacion_inventario': self.observacion_inventario,
        }


class InventoryRun(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String, nullable=False)
    period_id = db.Column(db.Integer, db.ForeignKey('inventory_period.id'))
    service = db.Column(db.String)
    service_scope_json = db.Column(db.String)
    status = db.Column(db.String, default='active', nullable=False)
    started_at = db.Column(db.String, nullable=False)
    closed_at = db.Column(db.String)
    created_by = db.Column(db.String)
    cancelled_at = db.Column(db.String)
    cancelled_by = db.Column(db.String)
    cancel_reason = db.Column(db.String)

    def to_dict(self):
        scope = []
        if self.service_scope_json:
            try:
                parsed = json.loads(self.service_scope_json)
                if isinstance(parsed, list):
                    scope = [str(x).strip() for x in parsed if str(x or '').strip()]
            except Exception:
                scope = []
        if (not scope) and self.service:
            scope = [str(self.service).strip()]
        return {
            'id': self.id,
            'name': self.name,
            'period_id': self.period_id,
            'service': self.service,
            'service_scope': scope,
            'service_scope_count': len(scope),
            'service_scope_label': ', '.join(scope[:3]) + (' ...' if len(scope) > 3 else ''),
            'status': self.status,
            'started_at': self.started_at,
            'started_at_local': format_dt_local(self.started_at),
            'closed_at': self.closed_at,
            'closed_at_local': format_dt_local(self.closed_at),
            'created_by': self.created_by,
            'cancelled_at': self.cancelled_at,
            'cancelled_at_local': format_dt_local(self.cancelled_at),
            'cancelled_by': self.cancelled_by,
            'cancel_reason': self.cancel_reason,
        }


class InventoryPeriod(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String, unique=True, nullable=False)
    period_type = db.Column(db.String, nullable=False, default='semestral')
    start_date = db.Column(db.String)
    end_date = db.Column(db.String)
    status = db.Column(db.String, nullable=False, default='open')
    notes = db.Column(db.String)
    created_at = db.Column(db.String, nullable=False)
    cancelled_at = db.Column(db.String)
    cancelled_by = db.Column(db.String)
    cancel_reason = db.Column(db.String)

    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'period_type': self.period_type,
            'start_date': self.start_date,
            'end_date': self.end_date,
            'status': self.status,
            'notes': self.notes,
            'created_at': self.created_at,
            'created_at_local': format_dt_local(self.created_at),
            'cancelled_at': self.cancelled_at,
            'cancelled_at_local': format_dt_local(self.cancelled_at),
            'cancelled_by': self.cancelled_by,
            'cancel_reason': self.cancel_reason,
        }


class RunAssetStatus(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    run_id = db.Column(db.Integer, db.ForeignKey('inventory_run.id'), nullable=False)
    asset_id = db.Column(db.Integer, db.ForeignKey('asset.id'), nullable=False)
    status = db.Column(db.String, nullable=False)
    scanned_at = db.Column(db.String, nullable=False)
    scanned_by = db.Column(db.String)

    __table_args__ = (
        UniqueConstraint('run_id', 'asset_id', name='uq_run_asset'),
    )


class AssetDisposal(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    asset_id = db.Column(db.Integer, db.ForeignKey('asset.id'), nullable=False, unique=True)
    period_id = db.Column(db.Integer, db.ForeignKey('inventory_period.id'))
    status = db.Column(db.String, nullable=False, default='Pendiente baja')
    reason = db.Column(db.String)
    requested_by = db.Column(db.String)
    requested_at = db.Column(db.String, nullable=False)
    reviewed_by = db.Column(db.String)
    reviewed_at = db.Column(db.String)
    review_notes = db.Column(db.String)

    def to_dict(self, asset=None):
        base = {
            'id': self.id,
            'asset_id': self.asset_id,
            'period_id': self.period_id,
            'status': self.status,
            'reason': self.reason,
            'requested_by': self.requested_by,
            'requested_at': self.requested_at,
            'requested_at_local': format_dt_local(self.requested_at),
            'reviewed_by': self.reviewed_by,
            'reviewed_at': self.reviewed_at,
            'reviewed_at_local': format_dt_local(self.reviewed_at),
            'review_notes': self.review_notes,
        }
        if asset:
            base['asset'] = asset.to_dict()
        return base


class SystemMeta(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    meta_key = db.Column(db.String, unique=True, nullable=False)
    meta_value = db.Column(db.String)


class GeneratedReport(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    report_type = db.Column(db.String, nullable=False)
    title = db.Column(db.String, nullable=False)
    period_id = db.Column(db.Integer, db.ForeignKey('inventory_period.id'))
    period_label = db.Column(db.String)
    file_name = db.Column(db.String, nullable=False)
    file_path = db.Column(db.String, nullable=False)
    generated_at = db.Column(db.String, nullable=False)

    def to_dict(self):
        return {
            'id': self.id,
            'report_type': self.report_type,
            'title': self.title,
            'period_id': self.period_id,
            'period_label': self.period_label,
            'file_name': self.file_name,
            'generated_at': self.generated_at,
            'generated_at_local': format_dt_local(self.generated_at),
        }


class AssetIssue(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    issue_type = db.Column(db.String, nullable=False)
    title = db.Column(db.String, nullable=False)
    severity = db.Column(db.String, nullable=False, default='Media')
    status = db.Column(db.String, nullable=False, default='Nuevo')
    source = db.Column(db.String, nullable=False, default='auto')
    period_id = db.Column(db.Integer, db.ForeignKey('inventory_period.id'))
    run_id = db.Column(db.Integer, db.ForeignKey('inventory_run.id'))
    asset_id = db.Column(db.Integer, db.ForeignKey('asset.id'))
    service = db.Column(db.String)
    detected_value = db.Column(db.Float, default=0.0)
    description = db.Column(db.String)
    assigned_to = db.Column(db.String)
    due_date = db.Column(db.String)
    resolution_notes = db.Column(db.String)
    created_at = db.Column(db.String, nullable=False)
    updated_at = db.Column(db.String, nullable=False)

    def to_dict(self):
        asset = Asset.query.get(self.asset_id) if self.asset_id else None
        return {
            'id': self.id,
            'issue_type': self.issue_type,
            'issue_type_label': ISSUE_TYPE_LABELS.get(self.issue_type, self.issue_type),
            'title': self.title,
            'severity': self.severity,
            'status': self.status,
            'source': self.source,
            'period_id': self.period_id,
            'run_id': self.run_id,
            'asset_id': self.asset_id,
            'asset_code': asset.c_act if asset else '',
            'asset_name': asset.nom if asset else '',
            'service': self.service or (asset.nom_ccos if asset else ''),
            'detected_value': to_number(self.detected_value),
            'description': self.description or '',
            'assigned_to': self.assigned_to or '',
            'due_date': self.due_date or '',
            'resolution_notes': self.resolution_notes or '',
            'created_at': self.created_at,
            'created_at_local': format_dt_local(self.created_at),
            'updated_at': self.updated_at,
            'updated_at_local': format_dt_local(self.updated_at),
        }


def ensure_db():
    if has_app_context():
        db.create_all()
        ensure_schema_updates()
        assign_legacy_period_to_old_runs()
        assign_legacy_period_to_old_disposals()
    else:
        with app.app_context():
            db.create_all()
            ensure_schema_updates()
            assign_legacy_period_to_old_runs()
            assign_legacy_period_to_old_disposals()


def invalidate_accounting_report_cache():
    global BASE_DATA_VERSION
    with ACCOUNTING_CACHE_LOCK:
        BASE_DATA_VERSION += 1
        ACCOUNTING_REPORT_CACHE['version'] = None
        ACCOUNTING_REPORT_CACHE['algo_version'] = None
        ACCOUNTING_REPORT_CACHE['bytes'] = None
        ACCOUNTING_REPORT_CACHE['filename'] = None


def get_system_meta(meta_key, default=None):
    row = SystemMeta.query.filter_by(meta_key=meta_key).first()
    if not row:
        return default
    return row.meta_value if row.meta_value is not None else default


def set_system_meta(meta_key, meta_value):
    row = SystemMeta.query.filter_by(meta_key=meta_key).first()
    if not row:
        row = SystemMeta(meta_key=meta_key)
        db.session.add(row)
    row.meta_value = str(meta_value)


def bump_assets_revision():
    raw = get_system_meta('assets_revision', '0')
    try:
        current = int(str(raw).strip())
    except Exception:
        current = 0
    next_value = current + 1
    set_system_meta('assets_revision', str(next_value))
    db.session.commit()
    return next_value


def get_assets_revision():
    raw = get_system_meta('assets_revision', '0')
    try:
        return int(str(raw).strip())
    except Exception:
        return 0


def ensure_schema_updates():
    with db.engine.begin() as conn:
        columns = {row[1] for row in conn.execute(text('PRAGMA table_info(asset)')).fetchall()}
        if 'c_fam' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN c_fam VARCHAR'))
        if 'nom_fam' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN nom_fam VARCHAR'))
        if 'desc_subtiac' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN desc_subtiac VARCHAR'))
        if 'deprecia' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN deprecia VARCHAR'))
        if 'vida_util' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN vida_util VARCHAR'))
        if 'tipo_activo_cache' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN tipo_activo_cache VARCHAR'))
        if 'raw_row_json' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN raw_row_json TEXT'))

        run_columns = {row[1] for row in conn.execute(text('PRAGMA table_info(inventory_run)')).fetchall()}
        if 'period_id' not in run_columns:
            conn.execute(text('ALTER TABLE inventory_run ADD COLUMN period_id INTEGER'))
        if 'service_scope_json' not in run_columns:
            conn.execute(text('ALTER TABLE inventory_run ADD COLUMN service_scope_json VARCHAR'))
        if 'cancelled_at' not in run_columns:
            conn.execute(text('ALTER TABLE inventory_run ADD COLUMN cancelled_at VARCHAR'))
        if 'cancelled_by' not in run_columns:
            conn.execute(text('ALTER TABLE inventory_run ADD COLUMN cancelled_by VARCHAR'))
        if 'cancel_reason' not in run_columns:
            conn.execute(text('ALTER TABLE inventory_run ADD COLUMN cancel_reason VARCHAR'))

        period_columns = {row[1] for row in conn.execute(text('PRAGMA table_info(inventory_period)')).fetchall()}
        if 'cancelled_at' not in period_columns:
            conn.execute(text('ALTER TABLE inventory_period ADD COLUMN cancelled_at VARCHAR'))
        if 'cancelled_by' not in period_columns:
            conn.execute(text('ALTER TABLE inventory_period ADD COLUMN cancelled_by VARCHAR'))
        if 'cancel_reason' not in period_columns:
            conn.execute(text('ALTER TABLE inventory_period ADD COLUMN cancel_reason VARCHAR'))

        report_columns = {row[1] for row in conn.execute(text('PRAGMA table_info(generated_report)')).fetchall()}
        if 'period_id' not in report_columns:
            conn.execute(text('ALTER TABLE generated_report ADD COLUMN period_id INTEGER'))

        disposal_columns = {row[1] for row in conn.execute(text('PRAGMA table_info(asset_disposal)')).fetchall()}
        if 'period_id' not in disposal_columns:
            conn.execute(text('ALTER TABLE asset_disposal ADD COLUMN period_id INTEGER'))


def assign_legacy_period_to_old_runs():
    orphan_runs = InventoryRun.query.filter(InventoryRun.period_id.is_(None)).count()
    if orphan_runs <= 0:
        return
    legacy = get_or_create_default_period()
    InventoryRun.query.filter(InventoryRun.period_id.is_(None)).update({'period_id': legacy.id})
    db.session.commit()


def assign_legacy_period_to_old_disposals():
    orphan_disposals = AssetDisposal.query.filter(AssetDisposal.period_id.is_(None)).count()
    if orphan_disposals <= 0:
        return
    legacy = get_or_create_default_period()
    AssetDisposal.query.filter(AssetDisposal.period_id.is_(None)).update({'period_id': legacy.id})
    db.session.commit()


def normalize_columns(cols):
    return {c.strip().upper(): c for c in cols}


def is_excluded_service_name(value):
    txt = str(value or '').strip().upper()
    return txt in EXCLUDED_SERVICE_NAMES


def parse_bool(value, default=False):
    if isinstance(value, bool):
        return value
    if value is None:
        return default
    txt = str(value).strip().lower()
    if txt in {'1', 'true', 'si', 'sí', 'yes', 'on'}:
        return True
    if txt in {'0', 'false', 'no', 'off', ''}:
        return False
    return default


def parse_int(value, default=None):
    try:
        if value in (None, ''):
            return default
        return int(value)
    except Exception:
        return default


def normalize_service_name(value):
    txt = str(value or '').strip()
    if not txt:
        return ''
    return txt


def normalize_service_scope(values):
    if values is None:
        return []
    if not isinstance(values, list):
        values = [values]
    result = []
    seen = set()
    for item in values:
        svc = normalize_service_name(item)
        if not svc or is_excluded_service_name(svc):
            continue
        key = svc.casefold()
        if key in seen:
            continue
        seen.add(key)
        result.append(svc)
    return result


def run_scope_services(run):
    if not run:
        return []
    scope = []
    raw = getattr(run, 'service_scope_json', None)
    if raw:
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, list):
                scope = parsed
        except Exception:
            scope = []
    scope = normalize_service_scope(scope)
    if (not scope) and getattr(run, 'service', None):
        scope = normalize_service_scope([run.service])
    return scope


def apply_run_scope_filter(query, run):
    scope = run_scope_services(run)
    if scope:
        query = query.filter(Asset.nom_ccos.in_(scope))
    return query


def get_cell(row, cols_map, key):
    """Retorna el valor de la columna mapeada o None si no existe."""
    if key in cols_map:
        try:
            v = row[cols_map[key]]
            if pd.isna(v):
                return None
            return v
        except Exception:
            return None
    return None


def get_cell_first(row, cols_map, keys):
    for key in keys:
        value = get_cell(row, cols_map, key)
        if value is not None and str(value).strip().lower() != 'nan':
            return value
    return None


def is_non_depreciable(value):
    if value is None:
        return False

    raw = str(value).strip()
    if not raw:
        return False

    # Si viene numérico (ej: 0, 0.0), aplica regla directa.
    numeric = raw.replace(',', '.')
    try:
        return float(numeric) <= 0
    except Exception:
        pass

    txt = raw.upper()
    txt = unicodedata.normalize('NFD', txt)
    txt = ''.join(ch for ch in txt if unicodedata.category(ch) != 'Mn')
    true_tokens = {'SI', 'S', 'TRUE', '1', 'Y', 'YES'}
    false_tokens = {'NO', 'N', 'FALSE', '0', '0.0', '0.00'}
    if txt in false_tokens:
        return True
    if txt in true_tokens:
        return False
    if 'NO' in txt and 'DEPREC' in txt:
        return True
    if 'SIN DEPREC' in txt:
        return True
    return False


def is_zero_useful_life(value):
    if value is None:
        return False
    txt = str(value).strip()
    if not txt:
        return False
    txt = txt.replace(',', '.')
    try:
        return float(txt) <= 0
    except Exception:
        return txt.upper() in {'CERO', 'SIN VIDA UTIL', 'NO APLICA', 'N/A'}


@app.route('/')
def index():
    ensure_db()
    return render_template('home.html')


@app.route('/inventario')
def inventario_page():
    ensure_db()
    return render_template('inventario.html')


@app.route('/jornadas')
def jornadas_page():
    ensure_db()
    return render_template('jornadas.html')


@app.route('/bajas')
def bajas_page():
    ensure_db()
    return render_template('bajas.html')


@app.route('/dashboard')
def dashboard_page():
    ensure_db()
    return render_template('dashboard.html')


@app.route('/informes')
def informes_page():
    ensure_db()
    return render_template('informes.html')


@app.route('/cronograma')
def cronograma_page():
    ensure_db()
    return render_template('cronograma.html')


@app.route('/novedades')
def novedades_page():
    ensure_db()
    return render_template('novedades.html')


@app.route('/logo')
def logo_file():
    logo_path = os.path.join(BASE_DIR, 'logo.png')
    if not os.path.exists(logo_path):
        return jsonify({'error': 'Logo no encontrado'}), 404
    return send_file(logo_path, mimetype='image/png')


@app.route('/import', methods=['POST'])
def import_file():
    ensure_db()
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file uploaded'}), 400

    filename = f.filename.lower()
    try:
        if filename.endswith('.csv'):
            df = pd.read_csv(f)
        else:
            df = pd.read_excel(f)
    except Exception as e:
        return jsonify({'error': f'Error leyendo archivo: {e}'}), 400

    cols = normalize_columns(df.columns)
    if 'C_ACT' not in cols:
        return jsonify({'error': 'El archivo debe contener la columna C_ACT'}), 400

    imported = 0
    updated = 0
    ordered_cols = list(df.columns)

    def serialize_raw_value(v):
        if pd.isna(v):
            return None
        if isinstance(v, pd.Timestamp):
            return v.isoformat()
        if hasattr(v, 'item'):
            try:
                return v.item()
            except Exception:
                pass
        return v

    for _, row in df.iterrows():
        c_act_val = get_cell(row, cols, 'C_ACT')
        c_act = str(c_act_val).strip() if c_act_val is not None else ''
        if not c_act:
            continue
        raw_payload = {}
        for col_name in ordered_cols:
            key = str(col_name).strip().upper()
            raw_payload[key] = serialize_raw_value(row[col_name])

        asset = Asset.query.filter_by(c_act=c_act).first()
        data = {
            'nom': get_cell(row, cols, 'NOM'),
            'modelo': get_cell(row, cols, 'MODELO'),
            'ref': get_cell(row, cols, 'REF'),
            'serie': get_cell(row, cols, 'SERIE'),
            'nom_marca': get_cell(row, cols, 'NOM_MARCA'),
            'c_fam': get_cell(row, cols, 'C_FAM'),
            'nom_fam': get_cell(row, cols, 'NOM_FAM'),
            'c_tiac': get_cell(row, cols, 'C_TIAC'),
            'desc_tiac': get_cell(row, cols, 'DESC_TIAC'),
            'desc_subtiac': get_cell_first(row, cols, [
                'DES_SUBTIAC',
                'DESC_SUBTIAC',
                'DES_SUB_TIA',
                'DESC_SUB_TIA',
                'SUBTIAC',
                'DESC_SUBTIPO_ACTIVO',
                'DES_SUBTIPO_ACTIVO',
            ]),
            'deprecia': get_cell_first(row, cols, [
                'DEPRECIA',
                'DEP',
                'SE_DEPRECIA',
            ]),
            'vida_util': get_cell_first(row, cols, [
                'VIDA_UTIL',
                'VIDA UTIL',
                'V_UTIL',
            ]),
            'des_ubi': get_cell(row, cols, 'DES_UBI'),
            'nom_ccos': get_cell(row, cols, 'NOM_CCOS'),
            'nom_resp': get_cell(row, cols, 'NOM_RESP'),
            'est': get_cell(row, cols, 'EST'),
            'costo': try_float(get_cell(row, cols, 'COSTO')),
            'saldo': try_float(get_cell(row, cols, 'SALDO')),
            'fecha_compra': get_cell(row, cols, 'FECHA_COMPRA'),
            'raw_row_json': json.dumps(raw_payload, ensure_ascii=False, default=str),
        }

        if asset:
            for k, v in data.items():
                if v is not None and v != 'nan':
                    setattr(asset, k, v)
            refresh_asset_type_cache(asset)
            updated += 1
        else:
            asset = Asset(c_act=c_act, **{k: v for k, v in data.items() if v is not None})
            refresh_asset_type_cache(asset)
            db.session.add(asset)
            imported += 1
    # Persist metadata so UI can show the currently imported base at all times.
    set_system_meta('last_import_file_name', str(f.filename or '').strip())
    set_system_meta('last_import_at', now_iso())
    set_system_meta('last_import_imported', str(imported))
    set_system_meta('last_import_updated', str(updated))

    db.session.commit()
    bump_assets_revision()
    invalidate_accounting_report_cache()
    return jsonify({'imported': imported, 'updated': updated})


@app.route('/import/status')
def import_status():
    ensure_db()

    file_name = (get_system_meta('last_import_file_name', '') or '').strip()
    imported_at = (get_system_meta('last_import_at', '') or '').strip()
    imported_raw = get_system_meta('last_import_imported', '0')
    updated_raw = get_system_meta('last_import_updated', '0')
    try:
        imported = int(str(imported_raw).strip())
    except Exception:
        imported = 0
    try:
        updated = int(str(updated_raw).strip())
    except Exception:
        updated = 0

    has_assets = db.session.query(Asset.id).first() is not None
    has_import = bool(file_name or imported_at or has_assets)

    return jsonify({
        'has_import': has_import,
        'file_name': file_name,
        'imported_at': imported_at,
        'imported_at_local': format_dt_local(imported_at) if imported_at else '',
        'imported': imported,
        'updated': updated,
    })


@app.route('/export_pdf')
def export_pdf():
    ensure_db()
    service = request.args.get('service')
    q = Asset.query
    if service:
        q = q.filter(Asset.nom_ccos == service)
    assets = [a.to_dict() for a in q.all()]
    if not assets:
        return jsonify({'error': 'No assets for given filter'}), 400

    out = BytesIO()
    c = canvas.Canvas(out, pagesize=letter)
    width, height = letter
    x_margin = 40
    y = height - 40
    c.setFont('Helvetica-Bold', 14)
    c.drawString(x_margin, y, f'A22 - Inventario - {service or "Todos"}')
    y -= 24
    c.setFont('Helvetica', 10)

    headers = [
        ('C_ACT', 'COD ACTIVO'),
        ('NOM', 'DESCRIPCION ACTIVO'),
        ('MODELO', 'MODELO'),
        ('SERIE', 'SERIAL'),
        ('DES_UBI', 'UBICACION'),
        ('NOM_RESP', 'RESPONSABLE'),
        ('estado_inventario', 'ESTADO INVENTARIO'),
    ]
    col_widths = [90, 140, 80, 80, 110, 110, 80]

    # draw header
    x = x_margin
    for i, (_, label) in enumerate(headers):
        c.drawString(x + 2, y, label)
        x += col_widths[i]
    y -= 14
    c.line(x_margin, y + 8, width - x_margin, y + 8)

    for a in assets:
        x = x_margin
        if y < 80:
            c.showPage()
            y = height - 40
        for i, (key, _) in enumerate(headers):
            text = str(a.get(key, '') or '')
            c.drawString(x + 2, y, text[:int(col_widths[i] / 6)])
            x += col_widths[i]
        y -= 14

    c.showPage()
    c.save()
    out.seek(0)
    base = f"a22_inventario_{service or 'almacen'}"
    filename = f"{clean_filename(base)}.pdf"
    return send_file(out, download_name=filename, as_attachment=True, mimetype='application/pdf')


def try_float(x):
    try:
        return float(x)
    except Exception:
        return None


def get_run_or_404(run_id):
    run = InventoryRun.query.get(run_id)
    if not run:
        return None, (jsonify({'error': 'Jornada no encontrada'}), 404)
    return run, None


def normalize_scan_code(raw):
    text = str(raw or '').strip().replace('\r', '').replace('\n', '')
    if not text:
        return ''
    compact = text.replace(' ', '')
    if compact and all(ch.isdigit() or ch == '.' for ch in compact):
        try:
            num = float(compact)
            if num.is_integer():
                return str(int(num))
        except Exception:
            pass
    return text


def scan_code_equals(left, right):
    a = normalize_scan_code(left)
    b = normalize_scan_code(right)
    if not a or not b:
        return False
    if a.casefold() == b.casefold():
        return True
    if a.isdigit() and b.isdigit():
        return (a.lstrip('0') or '0') == (b.lstrip('0') or '0')
    return False


def get_asset_by_code(code):
    if code is None:
        return None, None
    scan_code = normalize_scan_code(code)
    if not scan_code:
        return None, None

    for candidate in [scan_code, str(code).strip()]:
        if not candidate:
            continue
        asset = Asset.query.filter_by(c_act=candidate).first()
        if asset:
            return asset, 'C_ACT'

    if scan_code.isdigit():
        int_code = str(int(scan_code))
        asset = Asset.query.filter(Asset.c_act.in_([f'{int_code}.0', f'{int_code}.00'])).first()
        if asset:
            return asset, 'C_ACT'
        variants = Asset.query.filter(Asset.c_act.like(f'{int_code}.%')).limit(20).all()
        for row in variants:
            if scan_code_equals(row.c_act, scan_code):
                return row, 'C_ACT'

    keys = [
        'CODINTELIGENTE',
        'COD_BARRAS',
        'CODBARRAS',
        'CODIGO_BARRAS',
        'CODIGO DE BARRAS',
        'BARCODE',
        'BARRAS',
    ]
    candidates = Asset.query.filter(
        Asset.raw_row_json.isnot(None),
        Asset.raw_row_json.contains(scan_code)
    ).limit(500).all()
    for row in candidates:
        payload = asset_raw_payload(row)
        for key in keys:
            if scan_code_equals(payload.get(key), scan_code):
                return row, key
    return None, None


def get_or_create_default_period():
    period = InventoryPeriod.query.filter_by(name='LEGADO').first()
    if period:
        return period
    period = InventoryPeriod(
        name='LEGADO',
        period_type='historico',
        status='closed',
        created_at=now_iso(),
        notes='Periodo historico creado automaticamente para jornadas antiguas.',
    )
    db.session.add(period)
    db.session.commit()
    return period


def classify_area(service_name):
    service = (service_name or '').upper()
    assistential_keywords = [
        'URGEN', 'UCI', 'HOSP', 'QUIR', 'CIRUG', 'LAB', 'IMAGEN', 'CONSULT',
        'ODONTO', 'NEON', 'PEDIAT', 'FARM', 'SANGRE', 'RAYOS'
    ]
    administrative_keywords = [
        'ADMIN', 'GEREN', 'TALENTO', 'CONTAB', 'FINAN', 'SISTEM', 'ARCHIV',
        'ALMACE', 'JURID', 'FACTUR', 'CARTERA', 'COMPR', 'MANTEN'
    ]
    if any(k in service for k in assistential_keywords):
        return 'Asistencial'
    if any(k in service for k in administrative_keywords):
        return 'Administrativa'
    logistic_keywords = [
        'ALMACEN', 'LOGIST', 'MANTEN', 'SERVICIOS GENERALES', 'ACTIVOS FIJOS'
    ]
    if any(k in service for k in logistic_keywords):
        return 'Logistico'
    return 'Sin clasificar'


def a22_type_order_value(asset):
    group = classify_asset_group(asset).upper().strip()
    order_map = {
        'MUEBLE Y ENSER': 1,
        'BIOMEDICO': 2,
        'INDUSTRIAL': 3,
        'TECNOLOGICO': 4,
        'CONTROL': 5,
    }
    return order_map.get(group, 99)


def sort_assets_for_a22(assets):
    return sorted(
        list(assets or []),
        key=lambda a: (
            a22_type_order_value(a),
            str(classify_asset_group(a) or ''),
            str(a.c_act or ''),
        )
    )


def summarize_status(records):
    total = len(records)
    found = sum(1 for r in records if r.get('status') == 'Encontrado')
    not_found = sum(1 for r in records if r.get('status') == 'No encontrado')
    pending = max(total - found - not_found, 0)
    found_pct = round((found / total) * 100, 2) if total else 0
    not_found_pct = round((not_found / total) * 100, 2) if total else 0
    return {
        'total': total,
        'found': found,
        'not_found': not_found,
        'pending': pending,
        'found_pct': found_pct,
        'not_found_pct': not_found_pct,
    }


def to_number(value):
    try:
        if value is None:
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def asset_book_value(asset):
    saldo = to_number(asset.saldo)
    costo = to_number(asset.costo)
    if saldo > 0:
        return saldo
    if costo > 0:
        return costo
    return 0.0


def money_text(value):
    return f"${value:,.0f}"


def classify_critical_asset(asset):
    text = ' '.join([
        str(asset.nom or ''),
        str(asset.desc_tiac or ''),
        str(asset.c_tiac or ''),
        str(asset.modelo or ''),
        str(asset.ref or ''),
    ]).upper()

    keyword_weights = {
        'VENTIL': 10,
        'VENTILADOR': 10,
        'MONITOR': 9,
        'SIGNOS': 9,
        'DESFIBR': 10,
        'INFUSION': 8,
        'ANESTES': 9,
        'RAYOS': 8,
        'RX': 8,
        'ECOGRAF': 8,
        'CAMILLA': 7,
        'UCI': 8,
        'RESPIR': 9,
        'TOMOGRAF': 10,
    }
    matched = [(k, w) for k, w in keyword_weights.items() if k in text]
    value = asset_book_value(asset)

    value_score = 0
    if value >= 80_000_000:
        value_score = 10
    elif value >= 40_000_000:
        value_score = 8
    elif value >= 20_000_000:
        value_score = 6
    elif value >= 10_000_000:
        value_score = 4
    elif value >= 5_000_000:
        value_score = 2

    key_score = max([w for _, w in matched], default=0)
    score = key_score + value_score
    is_critical = score >= 8

    reasons = []
    if matched:
        reasons.append('Tipo critico')
    if value >= 10_000_000:
        reasons.append('Alto valor')
    if not reasons and is_critical:
        reasons.append('Prioridad tecnica')

    return {
        'is_critical': is_critical,
        'score': score,
        'reasons': ', '.join(reasons) if reasons else 'Sin marca critica',
    }


def build_management_insights(payload):
    insights = []
    k = payload.get('kpis', {})
    f = payload.get('financial', {})
    by_service = payload.get('by_service', [])
    by_type = payload.get('by_type', [])
    by_service_value = payload.get('top_not_found_by_service_value', [])
    critical = payload.get('critical_not_found', [])

    if by_service:
        top_service = max(by_service, key=lambda x: x.get('not_found', 0))
        insights.append(
            f"Servicio con mas no encontrados: {top_service.get('name', 'N/D')} "
            f"({top_service.get('not_found', 0)} activos)."
        )
    if by_type:
        top_type = max(by_type, key=lambda x: x.get('not_found', 0))
        insights.append(
            f"Tipo con mayor faltante: {top_type.get('name', 'N/D')} "
            f"({top_type.get('not_found', 0)} no encontrados)."
        )
    if by_service_value:
        s = by_service_value[0]
        insights.append(
            f"Mayor impacto economico por servicio: {s.get('name', 'N/D')} "
            f"({money_text(to_number(s.get('not_found_value', 0)))})."
        )
    insights.append(
        f"Valor no encontrado total: {money_text(to_number(f.get('not_found_value', 0)))} "
        f"({f.get('not_found_value_pct', 0)}% del valor inventariado)."
    )
    if critical:
        insights.append(
            f"Activos criticos no encontrados: {len(critical)} "
            f"por {money_text(to_number(f.get('critical_not_found_value', 0)))}."
        )
    if k.get('found_pct', 0) < 95:
        insights.append("Cumplimiento inferior al 95%; se recomienda plan de choque por servicio.")
    return insights


def build_executive_narrative(payload):
    k = payload.get('kpis', {})
    f = payload.get('financial', {})
    c = payload.get('coverage', {})
    meta = payload.get('meta', {})
    period_name = (meta.get('period') or {}).get('name') or 'Periodo seleccionado'
    run_name = (meta.get('run') or {}).get('name') or 'todas las jornadas'
    service_filter = meta.get('service_filter') or 'todos los servicios del alcance'

    objetivo_general = (
        f"Evaluar el estado de los activos fijos inventariados en {period_name}, "
        f"considerando el alcance operativo definido para {run_name} y {service_filter}, "
        "con el fin de soportar decisiones de control, custodia y mejora continua."
    )
    objetivos_especificos = [
        "Cuantificar activos encontrados, no encontrados y pendientes dentro del alcance evaluado.",
        "Estimar el impacto economico de los no encontrados y priorizar riesgos criticos.",
        "Contrastar cobertura del inventario frente a la base total institucional para contexto gerencial.",
        "Identificar servicios y tipos de activo con mayor brecha para acciones correctivas.",
    ]

    total = k.get('total', 0)
    found = k.get('found', 0)
    not_found = k.get('not_found', 0)
    pending = k.get('pending', 0)
    found_pct = k.get('found_pct', 0)
    not_found_pct = k.get('not_found_pct', 0)
    not_found_value = money_text(to_number(f.get('not_found_value', 0)))
    total_value = money_text(to_number(f.get('total_value', 0)))
    scope_assets = c.get('scope_assets', total)
    base_assets = c.get('base_total_assets', total)
    scope_assets_pct = c.get('scope_assets_pct', 0)

    resumen = (
        f"En el corte analizado se evaluaron {total} activos dentro del alcance del periodo. "
        f"Se registraron {found} encontrados ({found_pct}%), {not_found} no encontrados "
        f"({not_found_pct}%) y {pending} pendientes. En terminos economicos, el valor no "
        f"encontrado asciende a {not_found_value} sobre un valor total evaluado de {total_value}. "
        f"La cobertura del alcance corresponde a {scope_assets} activos sobre una base de "
        f"{base_assets} ({scope_assets_pct}%)."
    )

    interpretacion = []
    if found_pct >= 98:
        interpretacion.append("El nivel de cumplimiento de encontrados es sobresaliente para el corte evaluado.")
    elif found_pct >= 95:
        interpretacion.append("El cumplimiento de encontrados es aceptable, con oportunidades puntuales de mejora.")
    else:
        interpretacion.append("El cumplimiento de encontrados es bajo para el estandar institucional esperado.")

    if not_found_pct >= 5:
        interpretacion.append("El porcentaje de no encontrados requiere intervencion prioritaria por riesgo operativo.")
    else:
        interpretacion.append("El porcentaje de no encontrados se mantiene en una franja controlable.")

    if c.get('base_not_in_scope_assets', 0) > 0:
        interpretacion.append(
            f"Existe una brecha de cobertura de {c.get('base_not_in_scope_assets', 0)} activos frente a la base total."
        )
    else:
        interpretacion.append("La cobertura del periodo frente a la base institucional es completa.")

    return {
        'objetivo_general': objetivo_general,
        'objetivos_especificos': objetivos_especificos,
        'resumen': resumen,
        'interpretacion': interpretacion,
    }


def build_executive_action_plan(payload):
    k = payload.get('kpis', {})
    f = payload.get('financial', {})
    by_service = payload.get('by_service', [])
    by_type = payload.get('by_type', [])

    found_pct = float(k.get('found_pct', 0) or 0)
    not_found_pct = float(k.get('not_found_pct', 0) or 0)
    not_found_value = to_number(f.get('not_found_value', 0))

    risk_level = 'BAJO'
    risk_reason = 'Cumplimiento estable y brecha controlada.'
    if found_pct < 95 or not_found_pct >= 5 or not_found_value >= 50_000_000:
        risk_level = 'ALTO'
        risk_reason = 'Riesgo operativo y economico alto por brecha de inventario.'
    elif found_pct < 98 or not_found_pct >= 2 or not_found_value >= 20_000_000:
        risk_level = 'MEDIO'
        risk_reason = 'Riesgo moderado; requiere seguimiento dirigido.'

    top_service = max(by_service, key=lambda x: x.get('not_found', 0)) if by_service else None
    top_type = max(by_type, key=lambda x: x.get('not_found', 0)) if by_type else None

    actions = [
        {
            'priority': '1 - Inmediata',
            'action': 'Plan de choque de localizacion y saneamiento de no encontrados.',
            'focus': (top_service.get('name') if top_service else 'Servicios con mayor brecha'),
            'owner': 'Lideres de servicio + Activos fijos',
            'term': '15 dias',
        },
        {
            'priority': '2 - Corto plazo',
            'action': 'Auditoria dirigida a activos de mayor valor no encontrados.',
            'focus': (top_type.get('name') if top_type else 'Tipos de activo criticos'),
            'owner': 'Control interno + Biomédica/Ingenieria',
            'term': '30 dias',
        },
        {
            'priority': '3 - Sostenimiento',
            'action': 'Estandarizar cierres por periodo y trazabilidad por jornada.',
            'focus': 'Gobierno del dato e indicadores',
            'owner': 'Activos fijos + Sistemas',
            'term': 'Trimestral',
        },
    ]

    return {
        'risk_level': risk_level,
        'risk_reason': risk_reason,
        'actions': actions,
    }


def build_executive_conclusion(payload):
    k = payload.get('kpis', {})
    f = payload.get('financial', {})
    c = payload.get('coverage', {})
    by_service = payload.get('by_service', [])
    by_type = payload.get('by_type', [])
    meta = payload.get('meta', {})

    period_name = (meta.get('period') or {}).get('name') or 'el periodo evaluado'
    run_name = (meta.get('run') or {}).get('name') or 'las jornadas del periodo'

    top_service = max(by_service, key=lambda x: x.get('not_found', 0)).get('name') if by_service else 'N/D'
    top_type = max(by_type, key=lambda x: x.get('not_found', 0)).get('name') if by_type else 'N/D'

    found_pct = k.get('found_pct', 0)
    not_found_pct = k.get('not_found_pct', 0)
    scope_assets = c.get('scope_assets', k.get('total', 0))
    base_assets = c.get('base_total_assets', k.get('total', 0))
    not_found_value = money_text(to_number(f.get('not_found_value', 0)))

    technical_line = (
        f"Para {period_name}, considerando {run_name}, el sistema evidencia una efectividad de localizacion "
        f"de {found_pct}% sobre {scope_assets} activos evaluados, con una brecha de no localizacion de "
        f"{not_found_pct}% y un impacto economico asociado de {not_found_value}. "
        f"El mayor foco de atencion se concentra en el servicio '{top_service}' y en el tipo '{top_type}'. "
        f"La cobertura operativa alcanzada frente a la base institucional es de {scope_assets}/{base_assets} activos."
    )

    governance_line = (
        "En consecuencia, este informe consolida una base tecnica confiable para la toma de decisiones "
        "estrategicas y la priorizacion de intervenciones por riesgo operativo, economico y asistencial."
    )

    commitment_line = (
        "El equipo de Almacen y Activos Fijos ratifica su compromiso integral con los objetivos institucionales, "
        "fortaleciendo la custodia, trazabilidad y sostenibilidad de los bienes muebles e inmuebles del hospital."
    )

    return f"{technical_line} {governance_line} {commitment_line}"


def classify_asset_group(asset):
    def clean(value):
        text = str(value or '').upper().strip()
        text = unicodedata.normalize('NFD', text)
        return ''.join(ch for ch in text if unicodedata.category(ch) != 'Mn')

    def classify_non_control_type():
        des_subtiac_local = clean(asset.desc_subtiac)
        c_tiac_local = clean(asset.c_tiac)
        text_local = ' '.join([
            des_subtiac_local,
            clean(asset.desc_tiac),
            clean(asset.nom),
            clean(asset.modelo),
            clean(asset.ref),
            clean(asset.estado_inventario),
        ])

        if des_subtiac_local:
            if any(k in des_subtiac_local for k in ['BIOMED', 'MEDIC', 'HOSPITAL', 'CLINIC']):
                return 'BIOMEDICO'
            if any(k in des_subtiac_local for k in ['MUEBLE', 'ENSER']):
                return 'MUEBLE Y ENSER'
            if any(k in des_subtiac_local for k in ['INDUSTR']):
                return 'INDUSTRIAL'
            if any(k in des_subtiac_local for k in ['TECNOLOG']):
                return 'TECNOLOGICO'

        if c_tiac_local == '2':
            return 'MUEBLE Y ENSER'

        biomed_keywords = [
            'BIOMED', 'VENTIL', 'MONITOR', 'DESFIB', 'INFUS', 'BOMBA DE INFUS',
            'RESPIR', 'ANESTES', 'ECOGRAF', 'TOMOGRAF', 'RAYOS X', 'RAYOS',
            'ELECTRO', 'ELECTROCARD', 'ELECTROBIST', 'ELECTROESTIM', 'ELECTROTERAP',
            'ECG', 'EKG', 'SIGNOS VITALES', 'INCUBAD', 'SUCCION', 'ASPIRADOR QUIRURG',
            'NEONATAL', 'CARDIO', 'DIALIS', 'ULTRASON', 'UCI',
        ]
        if any(k in text_local for k in biomed_keywords):
            return 'BIOMEDICO'

        industrial_keywords = [
            'INDUSTR', 'PLANTA', 'COMPRESOR', 'TABLERO', 'CALDERA', 'MOTOR',
            'GENERADOR', 'BOMBA HIDRAUL', 'TRANSFORMADOR', 'SUBESTACION', 'UPS INDUSTRIAL',
            'ENFRIADOR', 'CHILLER', 'TORRE DE ENFRIAMIENTO',
        ]
        if c_tiac_local == '3' or any(k in text_local for k in industrial_keywords):
            return 'INDUSTRIAL'

        furniture_keywords = [
            'MUEBLE', 'ENSER', 'SILLA', 'ESCRITORIO', 'ARCHIVADOR', 'CAMILLA', 'MESA',
            'GABINETE', 'ESTANTE', 'VITRINA', 'LOCKER',
        ]
        if any(k in text_local for k in furniture_keywords):
            return 'MUEBLE Y ENSER'

        return 'TECNOLOGICO'

    des_subtiac = clean(asset.desc_subtiac)
    deprecia = asset.deprecia
    vida_util = asset.vida_util
    text = ' '.join([
        des_subtiac,
        clean(asset.desc_tiac),
        clean(asset.nom),
        clean(asset.modelo),
        clean(asset.ref),
        clean(asset.estado_inventario),
    ])
    base_type = classify_non_control_type()

    # Regla prioritaria para activos de control:
    # si no deprecia o vida util es 0, va a CONTROL separado por subtipo.
    if is_non_depreciable(deprecia) or is_zero_useful_life(vida_util):
        return f'CONTROL - {base_type}'

    # Regla principal: clasificar desde DES_SUBTIAC.
    if des_subtiac:
        if any(k in des_subtiac for k in ['CONTROL']):
            return f'CONTROL - {base_type}'

    control_keywords = [
        'ACTIVO DE CONTROL', 'CONTROL', 'KIT CONTROL', 'EQUIPO DE CONTROL',
    ]
    if any(k in text for k in control_keywords):
        return f'CONTROL - {base_type}'

    return base_type


def refresh_asset_type_cache(asset):
    asset.tipo_activo_cache = classify_asset_group(asset)
    return asset.tipo_activo_cache


def date_only(value):
    return format_dt_local(value, '%Y-%m-%d')


def normalize_disposal_type_key(type_value):
    txt = str(type_value or '').strip().upper()
    txt = unicodedata.normalize('NFD', txt)
    txt = ''.join(ch for ch in txt if unicodedata.category(ch) != 'Mn')
    if 'CONTROL' in txt:
        return 'CONTROL'
    if 'BIOMED' in txt:
        return 'BIOMEDICO'
    if 'MUEBLE' in txt:
        return 'MUEBLE Y ENSER'
    if 'INDUSTR' in txt:
        return 'INDUSTRIAL'
    if 'TECNOLOG' in txt:
        return 'TECNOLOGICO'
    return ''


def query_disposals(service=None, status=None, period_id=None):
    q = db.session.query(AssetDisposal, Asset).join(Asset, Asset.id == AssetDisposal.asset_id)
    if service:
        q = q.filter(Asset.nom_ccos == service)
    if status:
        q = q.filter(AssetDisposal.status == status)
    if period_id:
        q = q.filter(AssetDisposal.period_id == period_id)
    rows = q.order_by(AssetDisposal.id.desc()).limit(5000).all()
    items = []
    for d, a in rows:
        tipo = classify_asset_group(a)
        item = {
            'id': d.id,
            'code': a.c_act or '',
            'name': a.nom or '',
            'service': a.nom_ccos or '',
            'type': tipo,
            'cost': to_number(a.costo),
            'saldo': to_number(a.saldo),
            'date': date_only(a.fecha_compra),
            'reason': d.reason or '',
            'status': d.status or '',
            'period_id': d.period_id,
        }
        items.append(item)
    return items


def summarize_disposals(rows):
    return {
        'count': len(rows),
        'total_cost': round(sum(r.get('cost', 0) for r in rows), 2),
        'total_saldo': round(sum(r.get('saldo', 0) for r in rows), 2),
    }


def write_disposal_sheet(ws, title, rows, saldo_header='SALDO POR DEPRECIAR', note_text=None):
    headers = [
        'COD ACTIVO FIJO',
        'DESCRIPCION',
        'COSTO INICIAL',
        'SALDO POR DEPRECIAR',
        'FECHA ADQUISICION',
        'MOTIVO DE BAJA',
    ]
    ws.title = title[:31]
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws['A1'].font = Font(bold=True, size=13, color='0B4F6C')
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

    summary = summarize_disposals(rows)
    saldo_resume_label = 'Total saldo por depreciar'
    if 'NO DEPRECIABLE' in str(saldo_header or '').upper() or 'CONTABLE' in str(saldo_header or '').upper():
        saldo_resume_label = 'Total saldo contable'
    ws.append([
        f"Total activos: {summary['count']}  |  Total costo inicial: {money_text(summary['total_cost'])}  |  {saldo_resume_label}: {money_text(summary['total_saldo'])}"
    ])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws['A2'].font = Font(bold=True, color='1E293B')
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')

    header_row = 3
    if note_text:
        ws.append([note_text])
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
        ws['A3'].font = Font(bold=True, color='9A5F00')
        ws['A3'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        header_row = 4

    ws.append(headers)
    header_fill = PatternFill(fill_type='solid', start_color='EAF4FA', end_color='EAF4FA')
    header_font = Font(bold=True, color='0B4F6C')
    thin = Side(style='thin', color='D6E3EC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=header_row, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border

    for row in rows:
        ws.append([
            row.get('code', ''),
            row.get('name', ''),
            row.get('cost', 0),
            row.get('saldo', 0),
            row.get('date', ''),
            row.get('reason', ''),
        ])

    start_data_row = header_row + 1
    last_row = ws.max_row
    for r in range(start_data_row, last_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        ws.cell(row=r, column=3).number_format = '"$"#,##0'
        ws.cell(row=r, column=4).number_format = '"$"#,##0'

    ws.append([
        'TOTALES',
        '',
        summary['total_cost'],
        summary['total_saldo'],
        '',
        '',
    ])
    total_row = ws.max_row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=c)
        cell.font = Font(bold=True, color='0B4F6C')
        cell.fill = PatternFill(fill_type='solid', start_color='F3F9FD', end_color='F3F9FD')
        cell.border = border
    ws.cell(row=total_row, column=3).number_format = '"$"#,##0'
    ws.cell(row=total_row, column=4).number_format = '"$"#,##0'

    widths = [16, 40, 20, 20, 18, 34]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f'A{start_data_row}'


def get_hospital_logo_path():
    for candidate in A22_LOGO_CANDIDATES:
        if os.path.exists(candidate):
            return candidate
    return None


def get_codificacion_path():
    for candidate in CODIFICACION_CANDIDATES:
        if os.path.exists(candidate):
            return candidate
    return None


def append_pdf_header_with_logo(story, title_text, meta_text, include_logo=True):
    logo_path = get_hospital_logo_path()
    title_style = ParagraphStyle(
        'RptTitle',
        fontName='Helvetica-Bold',
        fontSize=17,
        textColor=colors.HexColor('#0B4F6C'),
        leading=20,
    )
    meta_style = ParagraphStyle(
        'RptMeta',
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#5A6B7B'),
        leading=12,
    )
    if include_logo and logo_path:
        logo = RLImage(logo_path, width=18 * mm, height=18 * mm)
        # Usa una tabla para fijar el logo en la esquina superior izquierda.
        head = Table([[logo, Paragraph(f"<b>{title_text}</b><br/>{meta_text}", meta_style)]], colWidths=[22 * mm, 160 * mm])
        head.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        story.append(head)
    else:
        story.append(Paragraph(title_text, title_style))
        story.append(Paragraph(meta_text, meta_style))
    story.append(Spacer(1, 6))


def make_pdf_page_header(
    logo_path,
    right_image_path=None,
    right_width_mm=44,
    right_height_mm=17,
    right_top_mm=17,
):
    logo_reader = ImageReader(logo_path) if logo_path else None
    right_reader = ImageReader(right_image_path) if right_image_path else None

    def _draw_header(c, doc):
        page_w, page_h = doc.pagesize
        c.saveState()
        if logo_reader:
            x = doc.leftMargin
            y = page_h - 16 * mm
            c.drawImage(logo_reader, x, y, width=14 * mm, height=14 * mm, preserveAspectRatio=True, mask='auto')
        if right_reader:
            # Imagen de codificacion en esquina superior derecha.
            right_w = right_width_mm * mm
            right_h = right_height_mm * mm
            x_right = page_w - doc.rightMargin - right_w
            y_right = page_h - right_top_mm * mm
            c.drawImage(right_reader, x_right, y_right, width=right_w, height=right_h, preserveAspectRatio=True, mask='auto')
        c.restoreState()

    return _draw_header


def add_logo_to_excel_sheet(ws, logo_path=None):
    if not logo_path:
        return
    try:
        logo = XLImage(logo_path)
        max_col = max(ws.max_column, 1)
        anchor_col = get_column_letter(max_col)
        logo.width = 90
        logo.height = 36
        # Evita anclar imagen sobre celdas combinadas del encabezado (fila 1),
        # lo cual puede provocar advertencias de reparacion en Excel.
        ws.add_image(logo, f'{anchor_col}2')
        current_h = ws.row_dimensions[2].height or 15
        ws.row_dimensions[2].height = max(current_h, 32)
    except Exception:
        # El reporte no debe fallar por un problema de imagen.
        return


def pdf_cell(text, styles, bold=False, align='LEFT', size=7):
    base = styles['Normal']
    style = ParagraphStyle(
        f'Cell_{align}_{size}_{1 if bold else 0}',
        parent=base,
        fontName='Helvetica-Bold' if bold else 'Helvetica',
        fontSize=size,
        leading=size + 1.5,
        alignment={'LEFT': 0, 'CENTER': 1, 'RIGHT': 2}.get(align, 0),
        wordWrap='CJK',
    )
    return Paragraph(escape(str(text or '')), style)


def reference_serial(asset):
    ref = (asset.ref or '').strip() if asset.ref else ''
    serial = (asset.serie or '').strip() if asset.serie else ''
    if ref and serial:
        return f'{ref} / {serial}'
    return ref or serial


def get_a22_scope(service=None, run_id=None, period_id=None):
    run = None
    q = Asset.query
    period = None
    if period_id:
        period = InventoryPeriod.query.get(period_id)
        if not period:
            return None, None, (jsonify({'error': 'Periodo no encontrado'}), 404)
    if run_id:
        run = InventoryRun.query.get(run_id)
        if not run:
            return None, None, (jsonify({'error': 'Jornada no encontrada'}), 404)
        if not period and run.period_id:
            period = InventoryPeriod.query.get(run.period_id)
        if period and run.period_id != period.id:
            return None, None, (jsonify({'error': 'La jornada no pertenece al periodo seleccionado'}), 400)
        q = apply_run_scope_filter(q, run)
    if service:
        q = q.filter(Asset.nom_ccos == service)
    if run:
        found_asset_ids = [
            row.asset_id for row in RunAssetStatus.query.filter(
                RunAssetStatus.run_id == run.id,
                RunAssetStatus.status == 'Encontrado'
            ).all()
        ]
        if not found_asset_ids:
            assets_scope = []
        else:
            assets_scope = q.filter(Asset.id.in_(found_asset_ids)).order_by(Asset.c_act.asc()).all()
    elif period:
        runs_in_period_q = InventoryRun.query.filter(InventoryRun.period_id == period.id)
        runs_in_period = runs_in_period_q.all()
        if service:
            svc_cf = str(service).casefold()
            runs_in_period = [
                r for r in runs_in_period
                if any(str(s).casefold() == svc_cf for s in run_scope_services(r))
            ]
        run_ids = [r.id for r in runs_in_period]
        scoped_assets = q.order_by(Asset.c_act.asc()).all()
        if not run_ids or not scoped_assets:
            assets_scope = []
        else:
            asset_ids = [a.id for a in scoped_assets]
            statuses = RunAssetStatus.query.filter(
                RunAssetStatus.run_id.in_(run_ids),
                RunAssetStatus.asset_id.in_(asset_ids)
            ).order_by(RunAssetStatus.id.desc()).all()
            latest_by_asset = {}
            for st in statuses:
                if st.asset_id not in latest_by_asset:
                    latest_by_asset[st.asset_id] = st.status
            allowed_ids = {aid for aid, st in latest_by_asset.items() if st == 'Encontrado'}
            assets_scope = [a for a in scoped_assets if a.id in allowed_ids]
    else:
        # Sin jornada: toma solo activos encontrados (escaneados/verificados como encontrados).
        assets_scope = q.filter(Asset.estado_inventario == 'Encontrado').order_by(Asset.c_act.asc()).all()
    return run, assets_scope, None


def normalize_inventory_status(value):
    txt = str(value or '').strip().upper()
    if txt == 'ENCONTRADO':
        return 'Encontrado'
    if txt == 'NO ENCONTRADO':
        return 'No encontrado'
    return 'Pendiente'


def build_reconciliation_rows(service=None, run_id=None, period_id=None):
    q = Asset.query
    run = None
    period = None
    if period_id:
        period = InventoryPeriod.query.get(period_id)
        if not period:
            return None, (jsonify({'error': 'Periodo no encontrado'}), 404)
    if run_id:
        run = InventoryRun.query.get(run_id)
        if not run:
            return None, (jsonify({'error': 'Jornada no encontrada'}), 404)
        if period and run.period_id != period.id:
            # Tolerancia ante desfasajes temporales de UI: si llega una combinacion
            # periodo/jornada invalida, usa la jornada como fuente de verdad.
            period = InventoryPeriod.query.get(run.period_id) if run.period_id else None
        q = apply_run_scope_filter(q, run)
    if service:
        q = q.filter(Asset.nom_ccos == service)
    assets = q.order_by(Asset.nom_ccos.asc(), Asset.c_act.asc()).all()

    status_map = {}
    if run and assets:
        statuses = RunAssetStatus.query.filter(
            RunAssetStatus.run_id == run.id,
            RunAssetStatus.asset_id.in_([a.id for a in assets])
        ).all()
        status_map = {s.asset_id: s.status for s in statuses}
    elif period and assets:
        runs_q = InventoryRun.query.filter(InventoryRun.period_id == period.id)
        runs_in_period = runs_q.all()
        if service:
            svc_cf = str(service).casefold()
            runs_in_period = [
                r for r in runs_in_period
                if any(str(s).casefold() == svc_cf for s in run_scope_services(r))
            ]
        run_ids = [r.id for r in runs_in_period]
        if run_ids:
            statuses = RunAssetStatus.query.filter(
                RunAssetStatus.run_id.in_(run_ids),
                RunAssetStatus.asset_id.in_([a.id for a in assets])
            ).order_by(RunAssetStatus.id.desc()).all()
            for st in statuses:
                if st.asset_id not in status_map:
                    status_map[st.asset_id] = st.status

    rows = []
    for a in assets:
        if run or period:
            status_value = status_map.get(a.id, 'Pendiente')
        else:
            status_value = a.estado_inventario
        rows.append({
            'C_ACT': a.c_act or '',
            'NOM': a.nom or '',
            'SERVICIO': a.nom_ccos or '',
            'UBICACION': a.des_ubi or '',
            'RESPONSABLE': a.nom_resp or '',
            'TIPO': classify_asset_group(a),
            'ESTADO_INVENTARIO': normalize_inventory_status(status_value),
            'FECHA_VERIFICACION': date_only(a.fecha_verificacion),
            'USUARIO_VERIFICADOR': a.usuario_verificador or '',
            'COSTO': to_number(a.costo),
            'SALDO': to_number(a.saldo),
        })
    return rows, None


def write_reconciliation_sheet(ws, title, rows):
    headers = [
        'CODIGO',
        'DESCRIPCION',
        'SERVICIO',
        'UBICACION',
        'RESPONSABLE',
        'TIPO',
        'ESTADO',
        'FECHA VERIFICACION',
        'USUARIO',
        'COSTO',
        'SALDO',
    ]
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws['A1'].font = Font(bold=True, size=13, color='0B4F6C')
    ws.append(headers)
    header_fill = PatternFill(fill_type='solid', start_color='EAF4FA', end_color='EAF4FA')
    header_font = Font(bold=True, color='0B4F6C')
    thin = Side(style='thin', color='D6E3EC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=2, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border

    for r in rows:
        ws.append([
            r['C_ACT'], r['NOM'], r['SERVICIO'], r['UBICACION'], r['RESPONSABLE'],
            r['TIPO'], r['ESTADO_INVENTARIO'], r['FECHA_VERIFICACION'], r['USUARIO_VERIFICADOR'],
            r['COSTO'], r['SALDO'],
        ])

    for i in range(3, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(i, c)
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        ws.cell(i, 10).number_format = '"$"#,##0'
        ws.cell(i, 11).number_format = '"$"#,##0'

    widths = [14, 36, 22, 24, 24, 18, 14, 16, 16, 14, 14]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = 'A3'


def excel_safe_sheet_name(base_name, used_names):
    txt = str(base_name or '').strip()
    if not txt:
        txt = 'SIN SERVICIO'
    for ch in ['\\', '/', '?', '*', '[', ']', ':']:
        txt = txt.replace(ch, ' ')
    txt = ' '.join(txt.split())
    if not txt:
        txt = 'SIN SERVICIO'
    txt = txt[:31]
    candidate = txt
    suffix = 2
    while candidate in used_names:
        base_trim = txt[: max(1, 31 - len(str(suffix)) - 1)]
        candidate = f'{base_trim}-{suffix}'
        suffix += 1
    used_names.add(candidate)
    return candidate


def fit_logo_to_a22_box(sheet, img, from_col=1, to_col=2, from_row=2, to_row=5, padding_px=8, shrink=0.88):
    # Aproximacion de tamaño de columnas/filas de Excel a pixeles.
    def col_px(col_index):
        letter = get_column_letter(col_index)
        width = sheet.column_dimensions[letter].width
        if width is None:
            width = 8.43
        return max(10, int(width * 7 + 5))

    def row_px(row_index):
        height = sheet.row_dimensions[row_index].height
        if height is None:
            height = 15
        return max(8, int(height * 96 / 72))

    target_w = sum(col_px(c) for c in range(from_col, to_col + 1)) - (padding_px * 2)
    target_h = sum(row_px(r) for r in range(from_row, to_row + 1)) - (padding_px * 2)
    if target_w <= 0 or target_h <= 0:
        return

    if img.width and img.height:
        scale = min(target_w / img.width, target_h / img.height) * shrink
        if scale > 0:
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)
            offset_x = max(0, int((target_w - img.width) / 2) + padding_px)
            offset_y = max(0, int((target_h - img.height) / 2) + padding_px)

            marker = AnchorMarker(
                col=from_col - 1,
                colOff=pixels_to_EMU(offset_x),
                row=from_row - 1,
                rowOff=pixels_to_EMU(offset_y),
            )
            ext = XDRPositiveSize2D(pixels_to_EMU(img.width), pixels_to_EMU(img.height))
            img.anchor = OneCellAnchor(_from=marker, ext=ext)


def build_dashboard_payload(service=None, run_id=None, period_id=None):
    q = Asset.query
    base_q = Asset.query
    run = None
    period = None
    period_runs = []
    if period_id:
        period = InventoryPeriod.query.get(period_id)
        if not period:
            return None, 'Periodo no encontrado'
        period_runs_q = InventoryRun.query.filter(InventoryRun.period_id == period.id)
        period_runs = period_runs_q.all()
    if run_id:
        run = InventoryRun.query.get(run_id)
        if not run:
            return None, 'Jornada no encontrada'
        if not period and run.period_id:
            period = InventoryPeriod.query.get(run.period_id)
        if period and run.period_id != period.id:
            return None, 'La jornada no pertenece al periodo seleccionado'
        q = apply_run_scope_filter(q, run)
    if service:
        q = q.filter(Asset.nom_ccos == service)
        base_q = base_q.filter(Asset.nom_ccos == service)
    elif period and not run:
        # En vista por periodo (sin jornada), limita a servicios que realmente
        # tuvieron jornadas en ese periodo para evitar mezclar servicios externos.
        services_in_period = sorted({
            svc
            for r in period_runs
            for svc in run_scope_services(r)
            if svc
        })
        if services_in_period:
            q = q.filter(Asset.nom_ccos.in_(services_in_period))
        else:
            q = q.filter(text('1=0'))

    base_assets = base_q.all()
    base_total_assets = len(base_assets)
    base_total_value = sum(asset_book_value(a) for a in base_assets)

    assets_scope = q.all()
    if not assets_scope:
        payload = {
            'kpis': summarize_status([]),
            'financial': {
                'total_value': 0,
                'found_value': 0,
                'not_found_value': 0,
                'pending_value': 0,
                'not_found_value_pct': 0,
                'critical_not_found_value': 0,
            },
            'by_service': [],
            'by_type': [],
            'by_area': [],
            'critical_not_found': [],
            'not_found_assets': [],
            'not_found_assets_total': 0,
            'not_found_assets_capped': False,
            'top_not_found_by_service_value': [],
            'top_not_found_by_type_value': [],
            'coverage': {
                'base_total_assets': base_total_assets,
                'base_total_value': round(base_total_value, 2),
                'scope_assets': 0,
                'scope_value': 0,
                'scope_assets_pct': 0,
                'scope_value_pct': 0,
                'base_not_in_scope_assets': max(base_total_assets, 0),
                'base_not_in_scope_value': round(base_total_value, 2),
            },
            'meta': {
                'run': run.to_dict() if run else None,
                'period': period.to_dict() if period else None,
                'service_filter': service or '',
                'generated_at': now_iso(),
            }
        }
        payload['insights'] = build_management_insights(payload)
        return payload, None

    status_map = {}
    if run:
        statuses = RunAssetStatus.query.filter(
            RunAssetStatus.run_id == run.id,
            RunAssetStatus.asset_id.in_([a.id for a in assets_scope])
        ).all()
        status_map = {s.asset_id: s.status for s in statuses}
    elif period:
        run_ids = [r.id for r in period_runs]
        asset_ids = [a.id for a in assets_scope]
        if run_ids and asset_ids:
            statuses = RunAssetStatus.query.filter(
                RunAssetStatus.run_id.in_(run_ids),
                RunAssetStatus.asset_id.in_(asset_ids)
            ).order_by(RunAssetStatus.id.desc()).all()
            for st in statuses:
                if st.asset_id not in status_map:
                    status_map[st.asset_id] = st.status

    records = []
    critical_not_found = []
    not_found_assets = []
    for a in assets_scope:
        if run or period:
            status = status_map.get(a.id, 'Pendiente')
        else:
            status = a.estado_inventario
        if status not in {'Encontrado', 'No encontrado'}:
            status = 'Pendiente'
        value = asset_book_value(a)
        critical_info = classify_critical_asset(a)
        records.append({
            'asset_id': a.id,
            'code': a.c_act,
            'asset_name': a.nom or '',
            'service': a.nom_ccos or 'SIN SERVICIO',
            'type': classify_asset_group(a),
            'area': classify_area(a.nom_ccos),
            'status': status,
            'value': value,
            'is_critical': critical_info['is_critical'],
            'critical_score': critical_info['score'],
            'critical_reasons': critical_info['reasons'],
        })
        if status == 'No encontrado' and critical_info['is_critical']:
            critical_not_found.append({
                'code': a.c_act,
                'name': a.nom or '',
                'service': a.nom_ccos or 'SIN SERVICIO',
                'type': classify_asset_group(a),
                'value': value,
                'critical_score': critical_info['score'],
                'critical_reasons': critical_info['reasons'],
                'model': a.modelo or '',
                'serial': a.serie or '',
                'responsible': a.nom_resp or '',
                'location': a.des_ubi or '',
            })
        if status == 'No encontrado':
            not_found_assets.append({
                'code': a.c_act,
                'name': a.nom or '',
                'service': a.nom_ccos or 'SIN SERVICIO',
                'type': classify_asset_group(a),
                'value': value,
                'model': a.modelo or '',
                'serial': a.serie or '',
                'responsible': a.nom_resp or '',
                'location': a.des_ubi or '',
            })

    by_service_map = {}
    by_type_map = {}
    by_area_map = {}
    for r in records:
        by_service_map.setdefault(r['service'], []).append(r)
        by_type_map.setdefault(r['type'], []).append(r)
        by_area_map.setdefault(r['area'], []).append(r)

    by_service = [{
        'name': name,
        **summarize_status(items),
    } for name, items in by_service_map.items()]
    by_type = [{
        'name': name,
        **summarize_status(items),
    } for name, items in by_type_map.items()]
    by_area = [{
        'name': name,
        **summarize_status(items),
    } for name, items in by_area_map.items()]

    by_service.sort(key=lambda x: x['total'], reverse=True)
    by_type.sort(key=lambda x: x['total'], reverse=True)
    by_area.sort(key=lambda x: x['total'], reverse=True)

    total_value = sum(r['value'] for r in records)
    found_value = sum(r['value'] for r in records if r['status'] == 'Encontrado')
    not_found_value = sum(r['value'] for r in records if r['status'] == 'No encontrado')
    pending_value = max(total_value - found_value - not_found_value, 0)
    critical_not_found_value = sum(x['value'] for x in critical_not_found)
    not_found_value_pct = round((not_found_value / total_value) * 100, 2) if total_value else 0
    scope_assets_pct = round((len(records) / base_total_assets) * 100, 2) if base_total_assets else 0
    scope_value_pct = round((total_value / base_total_value) * 100, 2) if base_total_value else 0

    by_service_value_map = {}
    by_type_value_map = {}
    for r in records:
        if r['status'] != 'No encontrado':
            continue
        by_service_value_map[r['service']] = by_service_value_map.get(r['service'], 0) + r['value']
        by_type_value_map[r['type']] = by_type_value_map.get(r['type'], 0) + r['value']

    top_not_found_by_service_value = sorted(
        [{'name': k, 'not_found_value': v} for k, v in by_service_value_map.items()],
        key=lambda x: x['not_found_value'],
        reverse=True
    )
    top_not_found_by_type_value = sorted(
        [{'name': k, 'not_found_value': v} for k, v in by_type_value_map.items()],
        key=lambda x: x['not_found_value'],
        reverse=True
    )
    critical_not_found.sort(key=lambda x: (x['critical_score'], x['value']), reverse=True)
    not_found_assets.sort(key=lambda x: x['value'], reverse=True)
    not_found_assets_total = len(not_found_assets)
    not_found_assets_cap = 500
    not_found_assets = not_found_assets[:not_found_assets_cap]

    payload = {
        'kpis': summarize_status(records),
        'financial': {
            'total_value': round(total_value, 2),
            'found_value': round(found_value, 2),
            'not_found_value': round(not_found_value, 2),
            'pending_value': round(pending_value, 2),
            'not_found_value_pct': not_found_value_pct,
            'critical_not_found_value': round(critical_not_found_value, 2),
        },
        'by_service': by_service,
        'by_type': by_type,
        'by_area': by_area,
        'critical_not_found': critical_not_found,
        'not_found_assets': not_found_assets,
        'not_found_assets_total': not_found_assets_total,
        'not_found_assets_capped': not_found_assets_total > not_found_assets_cap,
        'top_not_found_by_service_value': top_not_found_by_service_value,
        'top_not_found_by_type_value': top_not_found_by_type_value,
        'coverage': {
            'base_total_assets': base_total_assets,
            'base_total_value': round(base_total_value, 2),
            'scope_assets': len(records),
            'scope_value': round(total_value, 2),
            'scope_assets_pct': scope_assets_pct,
            'scope_value_pct': scope_value_pct,
            'base_not_in_scope_assets': max(base_total_assets - len(records), 0),
            'base_not_in_scope_value': round(max(base_total_value - total_value, 0), 2),
        },
        'meta': {
            'run': run.to_dict() if run else None,
            'period': period.to_dict() if period else None,
            'service_filter': service or '',
            'generated_at': now_iso(),
        }
    }
    payload['insights'] = build_management_insights(payload)
    return payload, None


@app.route('/services')
def services():
    ensure_db()
    raw_services = db.session.query(Asset.nom_ccos).all()
    cleaned = set()
    for row in raw_services:
        value = row[0]
        if value is None:
            continue
        text = str(value).strip()
        if not text:
            continue
        if is_excluded_service_name(text):
            continue
        cleaned.add(text)
    services = sorted(cleaned, key=lambda x: x.casefold())
    return jsonify({'services': services, 'total': len(services)})


@app.route('/responsibles')
def responsibles():
    ensure_db()
    raw = db.session.query(Asset.nom_resp).all()
    cleaned = set()
    for row in raw:
        value = row[0]
        if value is None:
            continue
        text = str(value).strip()
        if not text:
            continue
        cleaned.add(text)
    items = sorted(cleaned, key=lambda x: x.casefold())
    return jsonify({'responsibles': items, 'total': len(items)})


@app.route('/periods', methods=['GET'])
def list_periods():
    ensure_db()
    status = (request.args.get('status') or '').strip().lower()
    q = InventoryPeriod.query
    if status in {'open', 'closed', 'cancelled'}:
        q = q.filter(InventoryPeriod.status == status)
    periods = q.order_by(InventoryPeriod.id.desc()).all()
    return jsonify({'periods': [p.to_dict() for p in periods]})


@app.route('/periods', methods=['POST'])
def create_period():
    ensure_db()
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    period_type = (data.get('period_type') or 'semestral').strip().lower()
    start_date = (data.get('start_date') or '').strip() or None
    end_date = (data.get('end_date') or '').strip() or None
    notes = (data.get('notes') or '').strip() or None

    if not name:
        return jsonify({'error': 'Nombre del periodo es obligatorio'}), 400
    if period_type not in {'semestral', 'aleatorio', 'historico'}:
        return jsonify({'error': 'Tipo de periodo invalido'}), 400
    if InventoryPeriod.query.filter(db.func.upper(InventoryPeriod.name) == name.upper()).first():
        return jsonify({'error': 'Ya existe un periodo con ese nombre'}), 400

    period = InventoryPeriod(
        name=name,
        period_type=period_type,
        start_date=start_date,
        end_date=end_date,
        status='open',
        notes=notes,
        created_at=now_iso(),
    )
    db.session.add(period)
    db.session.commit()
    return jsonify({'period': period.to_dict()})


@app.route('/periods/<int:period_id>/close', methods=['POST'])
def close_period(period_id):
    ensure_db()
    period = InventoryPeriod.query.get(period_id)
    if not period:
        return jsonify({'error': 'Periodo no encontrado'}), 404
    if period.status == 'cancelled':
        return jsonify({'error': 'No puedes cerrar un periodo anulado'}), 400
    total_runs = InventoryRun.query.filter_by(period_id=period.id).count()
    if total_runs <= 0:
        return jsonify({'error': 'No puedes cerrar un periodo sin jornadas registradas'}), 400
    active_runs = InventoryRun.query.filter_by(period_id=period.id, status='active').count()
    if active_runs:
        return jsonify({'error': 'No puedes cerrar el periodo con jornadas activas'}), 400
    period.status = 'closed'
    db.session.commit()
    return jsonify({'period': period.to_dict()})


@app.route('/periods/<int:period_id>/cancel', methods=['POST'])
def cancel_period(period_id):
    ensure_db()
    period = InventoryPeriod.query.get(period_id)
    if not period:
        return jsonify({'error': 'Periodo no encontrado'}), 404
    if period.status == 'cancelled':
        return jsonify({'error': 'El periodo ya esta anulado'}), 400

    data = request.get_json() or {}
    reason = (data.get('reason') or '').strip()
    user = (data.get('user') or '').strip() or 'usuario_movil'
    if not reason:
        return jsonify({'error': 'Debes indicar el motivo de anulacion del periodo'}), 400

    runs = InventoryRun.query.filter_by(period_id=period.id).all()
    if any((r.status or '').strip().lower() == 'active' for r in runs):
        return jsonify({'error': 'No puedes anular el periodo porque tiene jornadas activas'}), 400

    run_ids = [r.id for r in runs]
    has_scan_trace = False
    if run_ids:
        has_scan_trace = db.session.query(RunAssetStatus.id).filter(
            RunAssetStatus.run_id.in_(run_ids)
        ).first() is not None
    if has_scan_trace:
        return jsonify({'error': 'No puedes anular el periodo porque ya tiene trazabilidad de escaneo'}), 400

    has_issues = AssetIssue.query.filter_by(period_id=period.id).count() > 0
    if has_issues:
        return jsonify({'error': 'No puedes anular el periodo porque tiene novedades registradas'}), 400

    has_disposals = AssetDisposal.query.filter_by(period_id=period.id).count() > 0
    if has_disposals:
        return jsonify({'error': 'No puedes anular el periodo porque tiene bajas asociadas'}), 400

    for run in runs:
        run.status = 'cancelled'
        run.closed_at = run.closed_at or now_iso()
        run.cancelled_at = now_iso()
        run.cancelled_by = user
        run.cancel_reason = f'Anulada por anulacion de periodo: {reason}'

    period.status = 'cancelled'
    period.cancelled_at = now_iso()
    period.cancelled_by = user
    period.cancel_reason = reason
    db.session.commit()
    return jsonify({'period': period.to_dict(), 'cancelled_runs': len(runs)})


@app.route('/periods/<int:period_id>/service_coverage', methods=['GET'])
def period_service_coverage(period_id):
    ensure_db()
    period = InventoryPeriod.query.get(period_id)
    if not period:
        return jsonify({'error': 'Periodo no encontrado'}), 404

    raw_services = db.session.query(Asset.nom_ccos).all()
    base_services = sorted({
        str(row[0]).strip()
        for row in raw_services
        if row[0] is not None and str(row[0]).strip() and not is_excluded_service_name(str(row[0]).strip())
    }, key=lambda x: x.casefold())

    runs = InventoryRun.query.filter_by(period_id=period_id).all()
    run_ids = [r.id for r in runs]
    run_service_map = {r.id: run_scope_services(r) for r in runs}

    done_services = set()
    if run_ids:
        status_rows = db.session.query(RunAssetStatus.run_id).filter(
            RunAssetStatus.run_id.in_(run_ids)
        ).distinct().all()
        for row in status_rows:
            for svc in run_service_map.get(row[0], []):
                if svc and not is_excluded_service_name(svc):
                    done_services.add(svc)

    # Si no hay registros de escaneo aun, considera servicios con jornada cerrada como gestionados.
    if not done_services:
        for r in runs:
            for svc in run_scope_services(r):
                if svc and (not is_excluded_service_name(svc)) and r.status == 'closed':
                    done_services.add(svc)

    pending_services = [s for s in base_services if s not in done_services]
    total_services = len(base_services)
    done_count = len(done_services)
    pending_count = len(pending_services)
    done_pct = round((done_count / total_services) * 100, 2) if total_services else 0.0
    pending_pct = round((pending_count / total_services) * 100, 2) if total_services else 0.0

    workload_rows = db.session.query(
        Asset.nom_ccos,
        db.func.count(Asset.id),
        db.func.coalesce(db.func.sum(Asset.costo), 0),
    ).filter(
        Asset.nom_ccos.isnot(None)
    ).group_by(
        Asset.nom_ccos
    ).all()
    workload_map = {}
    for svc, cnt, total_cost in workload_rows:
        svc_name = str(svc or '').strip()
        if not svc_name or is_excluded_service_name(svc_name):
            continue
        workload_map[svc_name] = {
            'asset_count': int(cnt or 0),
            'total_cost': float(total_cost or 0),
        }

    service_rows = []
    for svc in base_services:
        # Obtener todos los activos de este servicio
        assets = Asset.query.filter(Asset.nom_ccos == svc).all()
        asset_ids = [a.id for a in assets]
        total_assets = len(asset_ids)
        # Buscar el último status de cada activo en las jornadas de este periodo
        found_count = 0
        if asset_ids and run_ids:
            # Buscar el último status registrado para cada activo en este periodo
            subq = db.session.query(
                RunAssetStatus.asset_id,
                db.func.max(RunAssetStatus.id).label('max_id')
            ).filter(
                RunAssetStatus.run_id.in_(run_ids),
                RunAssetStatus.asset_id.in_(asset_ids)
            ).group_by(RunAssetStatus.asset_id).subquery()
            latest_statuses = db.session.query(RunAssetStatus).join(
                subq, (RunAssetStatus.asset_id == subq.c.asset_id) & (RunAssetStatus.id == subq.c.max_id)
            ).all()
            found_count = sum(1 for s in latest_statuses if s.status == 'Encontrado')
        status_pct = round((found_count / total_assets) * 100, 2) if total_assets else 0

        # Determinar el estado: 'Inventariado', 'En proceso' o 'Pendiente'
        # Un servicio solo es 'Inventariado' si todas sus jornadas están cerradas y está en done_services
        # Si tiene una jornada activa, debe decir 'En proceso' aunque esté en done_services
        in_active_run = any(
            r.status == 'active' and svc in run_scope_services(r)
            for r in runs
        )
        if in_active_run:
            status_label = 'En proceso'
        elif svc in done_services:
            status_label = 'Inventariado'
        else:
            status_label = 'Pendiente'

        service_rows.append({
            'service': svc,
            'status': status_label,
            'status_pct': status_pct,
            'asset_count': int(workload_map.get(svc, {}).get('asset_count', 0)),
            'total_cost': float(workload_map.get(svc, {}).get('total_cost', 0)),
        })

    # Recomendaciones operativas por carga (mayor cantidad de activos primero)
    pending_ranked = sorted(
        [r for r in service_rows if r['status'] == 'Pendiente'],
        key=lambda x: (x.get('asset_count', 0), x.get('total_cost', 0)),
        reverse=True
    )

    recommendations = []
    for idx, row in enumerate(pending_ranked[:5], start=1):
        recommendations.append({
            'priority': idx,
            'service': row['service'],
            'reason': f"Alta carga operativa ({row.get('asset_count', 0)} activos).",
        })

    def cluster_key(service_name):
        txt = str(service_name or '').upper().strip()
        if 'URGEN' in txt:
            return 'URGENCIAS'
        if 'HOSPITAL' in txt:
            return 'HOSPITALIZACION'
        if 'CIRUG' in txt or 'QUIROF' in txt:
            return 'CIRUGIA'
        return txt.split(' ')[0] if txt else 'OTROS'

    cluster_counts = {}
    for row in pending_ranked:
        key = cluster_key(row['service'])
        cluster_counts[key] = cluster_counts.get(key, 0) + 1

    grouped_tips = []
    for key, count in sorted(cluster_counts.items(), key=lambda x: x[1], reverse=True):
        if count >= 2:
            grouped_tips.append(
                f"{key}: {count} subservicios pendientes. Recomendado ejecutarlos el mismo dia con jornadas separadas por subservicio."
            )

    return jsonify({
        'period': period.to_dict(),
        'summary': {
            'total_services': total_services,
            'done_services': done_count,
            'pending_services': pending_count,
            'done_pct': done_pct,
            'pending_pct': pending_pct,
        },
        'services': service_rows,
        'recommendations': recommendations,
        'grouped_tips': grouped_tips,
    })


def detect_asset_issues_for_period(period_id, analyze_base=False):
    period = InventoryPeriod.query.get(period_id)
    if not period:
        return None, 'Periodo no encontrado'

    runs = InventoryRun.query.filter_by(period_id=period_id).all()
    run_ids = [r.id for r in runs]
    run_by_id = {r.id: r for r in runs}
    run_services = sorted({
        svc
        for r in runs
        for svc in run_scope_services(r)
        if svc
    })

    q_assets = Asset.query
    if analyze_base:
        assets = q_assets.all()
    else:
        if not run_services:
            assets = []
        else:
            assets = q_assets.filter(Asset.nom_ccos.in_(run_services)).all()
    assets_by_id = {a.id: a for a in assets}

    latest_status_map = {}
    status_rows = []
    if run_ids and assets:
        status_rows = RunAssetStatus.query.filter(
            RunAssetStatus.run_id.in_(run_ids),
            RunAssetStatus.asset_id.in_([a.id for a in assets])
        ).order_by(RunAssetStatus.id.desc()).all()
        for s in status_rows:
            if s.asset_id not in latest_status_map:
                latest_status_map[s.asset_id] = s

    duplicate_codes = {}
    rows_dup = db.session.query(Asset.c_act, db.func.count(Asset.id)).group_by(Asset.c_act).having(db.func.count(Asset.id) > 1).all()
    for code, count in rows_dup:
        duplicate_codes[str(code or '').strip()] = int(count or 0)

    duplicate_intelligent = {}
    for a in assets:
        payload = asset_raw_payload(a)
        c_int = str(payload.get('CODINTELIGENTE') or '').strip()
        if c_int:
            duplicate_intelligent[c_int] = duplicate_intelligent.get(c_int, 0) + 1

    disposal_by_asset = {d.asset_id: d for d in AssetDisposal.query.all()}

    now_iso_value = now_iso()
    AssetIssue.query.filter_by(period_id=period_id, source='auto').delete()
    db.session.flush()

    created = 0
    for a in assets:
        latest = latest_status_map.get(a.id)
        status = normalize_inventory_status(latest.status if latest else a.estado_inventario)
        value = asset_book_value(a)
        critical = classify_critical_asset(a)

        def add_issue(issue_type, severity, title, description, run_id=None):
            nonlocal created
            db.session.add(AssetIssue(
                issue_type=issue_type,
                title=title,
                severity=severity,
                status='Nuevo',
                source='auto',
                period_id=period_id,
                run_id=run_id,
                asset_id=a.id,
                service=a.nom_ccos or '',
                detected_value=value,
                description=description,
                created_at=now_iso_value,
                updated_at=now_iso_value,
            ))
            created += 1

        if status == 'No encontrado' and critical.get('is_critical'):
            add_issue(
                'NOT_FOUND_CRITICAL',
                'Alta',
                'Activo critico no encontrado',
                f"Estado='{status}' | Criticidad='{critical.get('critical_reasons')}' | Valor aprox={money_text(value)}."
            )
        if status == 'No encontrado' and value >= 20_000_000:
            add_issue(
                'NOT_FOUND_HIGH_VALUE',
                'Alta',
                'Activo no encontrado de alto valor',
                f"Estado='{status}' | Valor contable aprox={money_text(value)}."
            )

        if not str(a.serie or '').strip() and not str(a.ref or '').strip():
            add_issue(
                'MISSING_SERIAL_REF',
                'Media',
                'Activo sin serial ni referencia',
                f"SERIE='{str(a.serie or '').strip() or 'vacio'}' | REF='{str(a.ref or '').strip() or 'vacio'}'."
            )
        if not str(a.modelo or '').strip() and not str(a.nom_marca or '').strip():
            add_issue(
                'MISSING_MODEL_BRAND',
                'Baja',
                'Activo sin marca y modelo',
                f"MARCA='{str(a.nom_marca or '').strip() or 'vacio'}' | MODELO='{str(a.modelo or '').strip() or 'vacio'}'."
            )
        if not str(a.nom_resp or '').strip() or not str(a.des_ubi or '').strip():
            add_issue(
                'MISSING_CUSTODY_DATA',
                'Media',
                'Activo con datos de custodia incompletos',
                f"RESPONSABLE='{str(a.nom_resp or '').strip() or 'vacio'}' | UBICACION='{str(a.des_ubi or '').strip() or 'vacio'}'."
            )
        if to_number(a.costo) <= 0 or to_number(a.saldo) < 0:
            add_issue(
                'INVALID_FINANCIAL_VALUES',
                'Alta',
                'Valores financieros inconsistentes',
                f"Costo={to_number(a.costo)} | Saldo={to_number(a.saldo)}."
            )
        dep_no = is_non_depreciable(a.deprecia)
        vida_zero = is_zero_useful_life(a.vida_util)
        is_control_asset = (classify_asset_group(a) == 'CONTROL')
        if (not is_control_asset) and ((dep_no and (not vida_zero)) or ((not dep_no) and vida_zero)):
            add_issue(
                'DEPRECIATION_INCONSISTENT',
                'Media',
                'Inconsistencia entre deprecia y vida util',
                f"DEPRECIA='{a.deprecia or ''}' | VIDA_UTIL='{a.vida_util or ''}'."
            )
        if run_ids and status == 'Pendiente':
            add_issue(
                'PENDING_UNSCANNED',
                'Media',
                'Activo pendiente sin escaneo',
                f"Estado inventario='{status}' sin verificacion en jornada del periodo."
            )

        if str(a.c_act or '').strip() in duplicate_codes:
            add_issue(
                'DUPLICATE_CODE',
                'Alta',
                'Codigo de activo duplicado',
                f"Existen {duplicate_codes[str(a.c_act or '').strip()]} registros con el mismo codigo."
            )
        else:
            payload = asset_raw_payload(a)
            c_int = str(payload.get('CODINTELIGENTE') or '').strip()
            if c_int and duplicate_intelligent.get(c_int, 0) > 1:
                add_issue(
                    'DUPLICATE_CODE',
                    'Media',
                    'Codificacion inteligente repetida',
                    f"CODINTELIGENTE '{c_int}' repetido en {duplicate_intelligent.get(c_int, 0)} activos."
                )

        disp = disposal_by_asset.get(a.id)
        if disp and str(disp.status or '').strip().lower() in {'pendiente baja', 'en analisis', 'pendiente'}:
            sev = 'Alta' if value >= 10_000_000 else 'Media'
            add_issue(
                'CANDIDATE_DISPOSAL',
                sev,
                'Activo con baja pendiente',
                f"Estado baja='{disp.status}' | Valor aprox={money_text(value)}."
            )

    # Revisiones por escaneo en servicio distinto
    for s in status_rows:
        if normalize_inventory_status(s.status) != 'Encontrado':
            continue
        run = run_by_id.get(s.run_id)
        a = assets_by_id.get(s.asset_id)
        if not run or not a:
            continue
        run_scope = run_scope_services(run)
        run_service = str(run.service or '').strip()
        asset_service = str(a.nom_ccos or '').strip()
        scope_cf = {x.casefold() for x in run_scope}
        if run_scope and asset_service and asset_service.casefold() not in scope_cf:
            run_label = ', '.join(run_scope[:3]) + (' ...' if len(run_scope) > 3 else '')
            for issue_type, title in [
                ('SCANNED_OTHER_SERVICE', 'Escaneado en servicio distinto'),
                ('LOCATION_REVIEW', 'Revision de ubicacion requerida'),
                ('RESPONSIBLE_REVIEW', 'Revision de responsable requerida'),
            ]:
                db.session.add(AssetIssue(
                    issue_type=issue_type,
                    title=title,
                    severity='Media',
                    status='Nuevo',
                    source='auto',
                    period_id=period_id,
                    run_id=run.id,
                    asset_id=a.id,
                    service=run_label or run_service,
                    detected_value=asset_book_value(a),
                    description=f"Escaneado en alcance '{run_label or run_service}' pero base actual indica '{asset_service}'.",
                    created_at=now_iso_value,
                    updated_at=now_iso_value,
                ))
                created += 1

    db.session.commit()
    return {'created': created}, None


@app.route('/issues/scan', methods=['POST'])
def issues_scan():
    ensure_db()
    data = request.get_json() or {}
    period_id = data.get('period_id')
    try:
        period_id = int(period_id)
    except Exception:
        period_id = None
    if not period_id:
        return jsonify({'error': 'Periodo es obligatorio'}), 400
    analyze_base = parse_bool(data.get('analyze_base'), default=False)
    result, err = detect_asset_issues_for_period(period_id, analyze_base=analyze_base)
    if err:
        return jsonify({'error': err}), 400
    return jsonify({'ok': True, **result})


@app.route('/issues', methods=['GET'])
def issues_list():
    ensure_db()
    period_id = request.args.get('period_id', type=int)
    status = (request.args.get('status') or '').strip()
    severity = (request.args.get('severity') or '').strip()
    issue_type = (request.args.get('issue_type') or '').strip()

    q = AssetIssue.query
    if period_id:
        q = q.filter(AssetIssue.period_id == period_id)
    if status:
        q = q.filter(AssetIssue.status == status)
    if severity:
        q = q.filter(AssetIssue.severity == severity)
    if issue_type:
        q = q.filter(AssetIssue.issue_type == issue_type)

    rows = q.order_by(AssetIssue.severity.asc(), AssetIssue.id.desc()).all()
    items = [r.to_dict() for r in rows]
    total_value_risk = sum(to_number(x.get('detected_value')) for x in items if x.get('status') != 'Cerrado')

    by_status = {}
    by_severity = {}
    for x in items:
        by_status[x['status']] = by_status.get(x['status'], 0) + 1
        by_severity[x['severity']] = by_severity.get(x['severity'], 0) + 1

    return jsonify({
        'items': items,
        'summary': {
            'total': len(items),
            'open': sum(1 for x in items if x.get('status') != 'Cerrado'),
            'value_risk': round(total_value_risk, 2),
            'by_status': by_status,
            'by_severity': by_severity,
        },
        'meta': {
            'statuses': ISSUE_STATUSES,
            'severities': ISSUE_SEVERITIES,
            'issue_types': [{'key': k, 'label': v} for k, v in ISSUE_TYPE_LABELS.items()],
        }
    })


@app.route('/issues/<int:issue_id>', methods=['PATCH'])
def issues_update(issue_id):
    ensure_db()
    row = AssetIssue.query.get(issue_id)
    if not row:
        return jsonify({'error': 'Novedad no encontrada'}), 404
    data = request.get_json() or {}
    status = str(data.get('status') or '').strip()
    assigned_to = str(data.get('assigned_to') or '').strip()
    due_date = str(data.get('due_date') or '').strip()
    resolution_notes = str(data.get('resolution_notes') or '').strip()
    severity = str(data.get('severity') or '').strip()

    if status and status in ISSUE_STATUSES:
        row.status = status
    if severity and severity in ISSUE_SEVERITIES:
        row.severity = severity
    row.assigned_to = assigned_to or row.assigned_to
    row.due_date = due_date or row.due_date
    row.resolution_notes = resolution_notes or row.resolution_notes
    row.updated_at = now_iso()
    db.session.commit()
    return jsonify({'item': row.to_dict()})


@app.route('/issues/report_pdf', methods=['GET'])
def issues_report_pdf():
    ensure_db()
    period_id = request.args.get('period_id', type=int)
    if not period_id:
        return jsonify({'error': 'Periodo es obligatorio'}), 400
    period = InventoryPeriod.query.get(period_id)
    if not period:
        return jsonify({'error': 'Periodo no encontrado'}), 404

    rows = AssetIssue.query.filter_by(period_id=period_id).order_by(AssetIssue.id.desc()).all()
    out = BytesIO()
    doc = SimpleDocTemplate(out, pagesize=letter, leftMargin=14 * mm, rightMargin=14 * mm, topMargin=18 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    brand_blue = colors.HexColor('#0A6FB3')
    brand_blue_dark = colors.HexColor('#07507F')
    brand_green = colors.HexColor('#1E9E57')
    brand_yellow = colors.HexColor('#F2C94C')
    brand_red = colors.HexColor('#C0392B')
    brand_blue = colors.HexColor('#0A6FB3')
    brand_blue_dark = colors.HexColor('#07507F')
    brand_green = colors.HexColor('#1E9E57')
    brand_yellow = colors.HexColor('#F2C94C')
    brand_red = colors.HexColor('#C0392B')
    brand_blue = colors.HexColor('#0A6FB3')
    brand_blue_dark = colors.HexColor('#07507F')
    brand_green = colors.HexColor('#1E9E57')
    brand_yellow = colors.HexColor('#F2C94C')
    brand_red = colors.HexColor('#C0392B')
    title_style = ParagraphStyle('it', parent=styles['Heading2'], fontSize=14, textColor=colors.HexColor('#0B4F6C'))
    normal = ParagraphStyle('in', parent=styles['Normal'], fontSize=8, leading=10)
    story = []
    story.append(Paragraph(f'Informe de novedades y saneamiento - {period.name}', title_style))
    story.append(Spacer(1, 6))
    summary_data = [
        ['Total novedades', str(len(rows))],
        ['Abiertas', str(sum(1 for r in rows if r.status != 'Cerrado'))],
        ['Valor en riesgo', money_text(sum(to_number(r.detected_value) for r in rows if r.status != 'Cerrado'))],
    ]
    t = Table(summary_data, colWidths=[60 * mm, 120 * mm])
    t.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#C9DCE8')),
        ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#DCE8F0')),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F2F8FD')),
    ]))
    story.append(t)
    story.append(Spacer(1, 8))

    data = [[
        Paragraph('<b>ID</b>', normal),
        Paragraph('<b>TIPO</b>', normal),
        Paragraph('<b>ACTIVO</b>', normal),
        Paragraph('<b>SERVICIO</b>', normal),
        Paragraph('<b>SEVERIDAD</b>', normal),
        Paragraph('<b>ESTADO</b>', normal),
        Paragraph('<b>ASIGNADO</b>', normal),
    ]]
    for r in rows[:400]:
        info = r.to_dict()
        data.append([
            Paragraph(str(info.get('id', '')), normal),
            Paragraph(str(info.get('issue_type_label', '')), normal),
            Paragraph(f"{info.get('asset_code', '')} - {info.get('asset_name', '')}", normal),
            Paragraph(str(info.get('service', '') or ''), normal),
            Paragraph(str(info.get('severity', '')), normal),
            Paragraph(str(info.get('status', '')), normal),
            Paragraph(str(info.get('assigned_to', '') or ''), normal),
        ])
    tb = Table(data, colWidths=[10 * mm, 36 * mm, 56 * mm, 34 * mm, 18 * mm, 20 * mm, 24 * mm], repeatRows=1)
    tb.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0B4F6C')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#C9DCE8')),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#DCE8F0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7FBFD')]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(tb)
    doc.build(story, onFirstPage=make_pdf_page_header(get_hospital_logo_path()), onLaterPages=make_pdf_page_header(get_hospital_logo_path()))
    out.seek(0)
    name = f"novedades_saneamiento_{period.name.replace(' ', '_')}.pdf"
    return send_file(out, as_attachment=True, download_name=name, mimetype='application/pdf')


@app.route('/assets')
def assets():
    ensure_db()
    service = request.args.get('service')
    run_id = request.args.get('run_id', type=int)
    q = Asset.query

    run = None
    if run_id:
        run = InventoryRun.query.get(run_id)
        if not run:
            return jsonify({'error': 'Jornada no encontrada'}), 404
        q = apply_run_scope_filter(q, run)

    if service:
        q = q.filter(Asset.nom_ccos == service)

    assets_list = q.limit(5000).all()
    items = [a.to_dict() for a in assets_list]

    ids = [a.id for a in assets_list]
    status_by_asset = {}
    if run and ids:
        statuses = RunAssetStatus.query.filter(
            RunAssetStatus.run_id == run.id,
            RunAssetStatus.asset_id.in_(ids)
        ).all()
        status_by_asset = {s.asset_id: s.status for s in statuses}

    for i, asset in enumerate(assets_list):
        items[i]['estado_jornada'] = status_by_asset.get(asset.id, '')
        items[i]['gestionado_jornada'] = asset.id in status_by_asset

    if assets_list:
        disposal_rows = AssetDisposal.query.filter(
            AssetDisposal.asset_id.in_([a.id for a in assets_list])
        ).all()
        disposal_by_asset = {d.asset_id: d.status for d in disposal_rows}
        for i, asset in enumerate(assets_list):
            items[i]['estado_baja'] = disposal_by_asset.get(asset.id, '')

    return jsonify({'assets': items})


@app.route('/assets/<int:asset_id>/classification', methods=['PATCH'])
def update_asset_classification(asset_id):
    ensure_db()
    asset = Asset.query.get(asset_id)
    if not asset:
        return jsonify({'error': 'Activo no encontrado'}), 404

    data = request.get_json() or {}
    classification = (data.get('classification') or '').strip()
    disposal_reason = (data.get('disposal_reason') or '').strip()
    period_id = parse_int(data.get('period_id'))
    run_id = parse_int(data.get('run_id'))
    user = (data.get('user') or '').strip() or 'usuario_movil'
    if not classification:
        return jsonify({'error': 'Debe enviar la clasificacion'}), 400

    allowed = {
        'Pendiente verificacion',
        'En mantenimiento',
        'Prestado',
        'Activo de control',
        'Para baja',
        'Baja aprobada',
    }
    if classification not in allowed:
        return jsonify({'error': 'Clasificacion invalida'}), 400
    if classification in {'Para baja', 'Baja aprobada'} and not disposal_reason:
        return jsonify({'error': 'Debes registrar el motivo real de baja'}), 400

    now_iso_value = now_iso()
    resolved_period_id = None
    if run_id:
        run = InventoryRun.query.get(run_id)
        if run and run.period_id:
            resolved_period_id = run.period_id
    if resolved_period_id is None and period_id:
        period = InventoryPeriod.query.get(period_id)
        if not period:
            return jsonify({'error': 'Periodo no encontrado'}), 404
        resolved_period_id = period.id
    if resolved_period_id is None:
        resolved_period_id = get_or_create_default_period().id

    asset.estado_inventario = classification
    asset.fecha_verificacion = now_iso_value
    asset.usuario_verificador = user

    if classification in {'Para baja', 'Baja aprobada'}:
        disposal = AssetDisposal.query.filter_by(asset_id=asset.id).first()
        if not disposal:
            disposal = AssetDisposal(
                asset_id=asset.id,
                period_id=resolved_period_id,
                status='Pendiente baja',
                reason=disposal_reason,
                requested_by=user,
                requested_at=now_iso_value,
            )
            db.session.add(disposal)
        disposal.period_id = resolved_period_id
        if classification == 'Para baja':
            disposal.status = 'Pendiente baja'
            disposal.reason = disposal_reason
            disposal.requested_by = user
            disposal.requested_at = now_iso_value
            disposal.reviewed_by = None
            disposal.reviewed_at = None
            disposal.review_notes = None
        if classification == 'Baja aprobada':
            disposal.status = 'Aprobada para baja'
            disposal.reason = disposal_reason
            disposal.reviewed_by = user
            disposal.reviewed_at = now_iso_value
            disposal.review_notes = disposal.review_notes or 'Aprobada desde inventario'

    refresh_asset_type_cache(asset)
    db.session.commit()
    return jsonify({'ok': True, 'asset': asset.to_dict(), 'classification': classification})


@app.route('/assets/<int:asset_id>/service', methods=['PATCH'])
def update_asset_service(asset_id):
    ensure_db()
    asset = Asset.query.get(asset_id)
    if not asset:
        return jsonify({'error': 'Activo no encontrado'}), 404

    data = request.get_json() or {}
    service = (data.get('service') or '').strip()
    user = (data.get('user') or '').strip() or 'usuario_movil'
    run_id = data.get('run_id')

    if not service:
        return jsonify({'error': 'Debe enviar el servicio destino'}), 400

    if run_id is not None:
        run = InventoryRun.query.get(run_id)
        if not run:
            return jsonify({'error': 'Jornada no encontrada'}), 404
        if run.status != 'active':
            return jsonify({'error': 'La jornada ya esta cerrada'}), 400
        run_scope = run_scope_services(run)
        run_scope_cf = {s.casefold() for s in run_scope}
        if run_scope and service.casefold() not in run_scope_cf:
            return jsonify({'error': 'El servicio destino debe estar dentro del alcance de la jornada activa'}), 400

    now_iso_value = now_iso()
    old_service = str(asset.nom_ccos or '').strip()
    asset.nom_ccos = service
    asset.fecha_verificacion = now_iso_value
    asset.usuario_verificador = user
    db.session.commit()
    return jsonify({
        'ok': True,
        'asset': asset.to_dict(),
        'old_service': old_service,
        'new_service': service,
    })


@app.route('/scan', methods=['POST'])
def scan():
    ensure_db()
    data = request.get_json() or {}
    code = data.get('code')
    user = data.get('user') or 'unknown'
    run_id = data.get('run_id')
    if not code:
        return jsonify({'error': 'No code provided'}), 400
    scanned_code = normalize_scan_code(code)
    asset, matched_by = get_asset_by_code(scanned_code)
    if not asset:
        return jsonify({'found': False, 'scanned_code': scanned_code}), 200

    if run_id is None:
        return jsonify({'error': 'Debes iniciar una jornada activa para escanear'}), 400

    run = InventoryRun.query.get(run_id)
    if not run:
        return jsonify({'error': 'Jornada no encontrada'}), 404
    if run.status != 'active':
        return jsonify({'error': 'La jornada ya esta cerrada'}), 400
    run_scope = run_scope_services(run)
    run_service = str(run.service or '').strip()
    asset_service = str(asset.nom_ccos or '').strip()
    run_scope_cf = {s.casefold() for s in run_scope}
    if run_scope and asset_service.casefold() not in run_scope_cf:
        expected_label = ', '.join(run_scope[:3]) + (' ...' if len(run_scope) > 3 else '')
        return jsonify({
            'error': f'Escaneado fuera del alcance de la jornada. Base actual: {asset_service or "sin servicio"} | Alcance: {expected_label}',
            'code': 'SERVICE_MISMATCH',
            'expected_service': run_service or (run_scope[0] if run_scope else ''),
            'expected_services': run_scope,
            'expected_service_label': expected_label,
            'current_service': asset_service,
            'matched_by': matched_by or 'C_ACT',
            'scanned_code': scanned_code,
            'asset': asset.to_dict(),
            'run': run.to_dict(),
        }), 409

    run_status = RunAssetStatus.query.filter_by(run_id=run.id, asset_id=asset.id).first()
    if not run_status:
        run_status = RunAssetStatus(
            run_id=run.id,
            asset_id=asset.id,
            status='Encontrado',
            scanned_at=now_iso(),
            scanned_by=user,
        )
        db.session.add(run_status)
    else:
        run_status.status = 'Encontrado'
        run_status.scanned_at = now_iso()
        run_status.scanned_by = user

    asset.estado_inventario = 'Encontrado'
    asset.fecha_verificacion = now_iso()
    asset.usuario_verificador = user
    db.session.commit()
    return jsonify({
        'found': True,
        'asset': asset.to_dict(),
        'run_id': run.id if run else None,
        'matched_by': matched_by or 'C_ACT',
        'scanned_code': scanned_code,
    })


@app.route('/runs', methods=['GET'])
def list_runs():
    ensure_db()
    period_id = request.args.get('period_id', type=int)
    status = (request.args.get('status') or '').strip().lower()
    q = InventoryRun.query
    if period_id:
        q = q.filter(InventoryRun.period_id == period_id)
    if status in {'active', 'closed', 'cancelled'}:
        q = q.filter(InventoryRun.status == status)
    runs = q.order_by(InventoryRun.id.desc()).limit(300).all()
    run_ids = [r.id for r in runs]
    period_ids = sorted({r.period_id for r in runs if r.period_id})
    periods_map = {}
    if period_ids:
        periods_map = {p.id: p for p in InventoryPeriod.query.filter(InventoryPeriod.id.in_(period_ids)).all()}
    found_by_run = {}
    not_found_by_run = {}
    if run_ids:
        statuses = db.session.query(
            RunAssetStatus.run_id,
            RunAssetStatus.status,
            db.func.count(RunAssetStatus.id)
        ).filter(
            RunAssetStatus.run_id.in_(run_ids)
        ).group_by(
            RunAssetStatus.run_id,
            RunAssetStatus.status
        ).all()
        for run_id, status, count in statuses:
            if status == 'Encontrado':
                found_by_run[run_id] = int(count or 0)
            elif status == 'No encontrado':
                not_found_by_run[run_id] = int(count or 0)

    payload = []
    for r in runs:
        row = r.to_dict()
        period = periods_map.get(r.period_id)
        row['period_name'] = period.name if period else None
        row['period_status'] = period.status if period else None
        row['found'] = found_by_run.get(r.id, 0)
        row['not_found'] = not_found_by_run.get(r.id, 0)
        payload.append(row)
    return jsonify({'runs': payload})


@app.route('/disposals', methods=['GET'])
def list_disposals():
    ensure_db()
    service = request.args.get('service')
    status = request.args.get('status')
    period_id = request.args.get('period_id', type=int)
    rows = query_disposals(service=service, status=status, period_id=period_id)
    items = []
    for row in rows:
        items.append({
            'id': row['id'],
            'period_id': row.get('period_id'),
            'reason': row['reason'],
            'status': row['status'],
            'asset': {
                'C_ACT': row['code'],
                'NOM': row['name'],
                'NOM_CCOS': row['service'],
                'TIPO_ACTIVO': row['type'],
                'COSTO': row['cost'],
                'SALDO': row['saldo'],
                'FECHA_COMPRA': row['date'],
            }
        })
    return jsonify({'disposals': items})


@app.route('/disposals/export_excel', methods=['GET'])
def export_disposals_excel():
    ensure_db()
    service = request.args.get('service')
    status = request.args.get('status')
    period_id = request.args.get('period_id', type=int)
    type_exact = (request.args.get('type_exact') or '').strip()
    type_key = normalize_disposal_type_key(request.args.get('type'))
    rows = query_disposals(service=service, status=status, period_id=period_id)
    if type_exact:
        rows = [r for r in rows if str(r.get('type', '')).strip().upper() == type_exact.upper()]
    elif type_key:
        rows = [r for r in rows if normalize_disposal_type_key(r.get('type')) == type_key]
    if not rows:
        return jsonify({'error': 'No hay activos para exportar con ese filtro'}), 400

    is_control_report = (type_exact and 'CONTROL' in type_exact.upper()) or (type_key == 'CONTROL')
    logo_path = get_hospital_logo_path()
    wb = Workbook()
    ws = wb.active
    title = f'Bajas {type_exact}' if type_exact else (f'Bajas {type_key}' if type_key else 'Bajas - Todos los tipos')
    write_disposal_sheet(
        ws,
        title,
        rows,
        saldo_header='SALDO CONTABLE (NO DEPRECIABLE)' if is_control_report else 'SALDO POR DEPRECIAR',
        note_text=(
            'Nota: los activos de control no se deprecian; por politica contable su saldo contable suele coincidir con el costo inicial.'
            if is_control_report else None
        ),
    )
    add_logo_to_excel_sheet(ws, logo_path=logo_path)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    filename_type = (type_exact or type_key or 'todos').lower().replace(' ', '_')
    filename = clean_filename(f"bajas_{filename_type}.xlsx")
    return send_file(
        out,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/disposals/export_pdf', methods=['GET'])
def export_disposals_pdf():
    ensure_db()
    service = request.args.get('service')
    status = request.args.get('status')
    period_id = request.args.get('period_id', type=int)
    type_exact = (request.args.get('type_exact') or '').strip()
    type_key = normalize_disposal_type_key(request.args.get('type'))
    rows = query_disposals(service=service, status=status, period_id=period_id)
    if type_exact:
        rows = [r for r in rows if str(r.get('type', '')).strip().upper() == type_exact.upper()]
    elif type_key:
        rows = [r for r in rows if normalize_disposal_type_key(r.get('type')) == type_key]
    if not rows:
        return jsonify({'error': 'No hay activos para exportar con ese filtro'}), 400

    is_control_report = (type_exact and 'CONTROL' in type_exact.upper()) or (type_key == 'CONTROL')
    summary = summarize_disposals(rows)
    logo_path = get_hospital_logo_path()
    out = BytesIO()
    doc = SimpleDocTemplate(
        out, pagesize=letter, leftMargin=16 * mm, rightMargin=16 * mm, topMargin=22 * mm, bottomMargin=14 * mm
    )
    styles = getSampleStyleSheet()
    story = []
    report_title = f"Bajas - {type_exact}" if type_exact else (f"Bajas - {type_key}" if type_key else "Bajas - Todos los tipos")
    append_pdf_header_with_logo(
        story,
        report_title,
        f"Generado: {now_local_dt().strftime('%Y-%m-%d %H:%M')} | Servicio: {service or 'TODOS'} | Estado: {status or 'TODOS'}",
        include_logo=False,
    )
    story.append(Paragraph(
        f"<b>Total activos:</b> {summary['count']} &nbsp;&nbsp; <b>Total costo inicial:</b> {money_text(summary['total_cost'])} "
        f"&nbsp;&nbsp; <b>{'Total saldo contable' if is_control_report else 'Total saldo por depreciar'}:</b> {money_text(summary['total_saldo'])}",
        styles['Normal']
    ))
    if is_control_report:
        story.append(Paragraph(
            '<b>Nota:</b> Los activos de control no se deprecian; por politica contable su saldo contable suele coincidir con el costo inicial.',
            ParagraphStyle('CtrlNote', parent=styles['Normal'], textColor=colors.HexColor('#9A5F00'), fontSize=9)
        ))
    story.append(Spacer(1, 8))

    table_data = [[
        pdf_cell('COD ACTIVO FIJO', styles, bold=True, align='CENTER'),
        pdf_cell('DESCRIPCION', styles, bold=True, align='CENTER'),
        pdf_cell('COSTO INICIAL', styles, bold=True, align='CENTER'),
        pdf_cell('SALDO POR DEPRECIAR', styles, bold=True, align='CENTER'),
        pdf_cell('FECHA ADQUISICION', styles, bold=True, align='CENTER'),
        pdf_cell('MOTIVO DE BAJA', styles, bold=True, align='CENTER'),
    ]]
    for r in rows:
        table_data.append([
            pdf_cell(r['code'], styles, align='CENTER'),
            pdf_cell(r['name'], styles),
            pdf_cell(money_text(r['cost']), styles, align='RIGHT'),
            pdf_cell(money_text(r['saldo']), styles, align='RIGHT'),
            pdf_cell(r['date'], styles, align='CENTER'),
            pdf_cell(r['reason'], styles),
        ])
    table = Table(table_data, colWidths=[24*mm, 60*mm, 24*mm, 24*mm, 24*mm, 42*mm], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#EAF4FA')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#0B4F6C')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#C8D8E4')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FCFF')]),
    ]))
    story.append(table)
    page_header = make_pdf_page_header(logo_path)
    doc.build(story, onFirstPage=page_header, onLaterPages=page_header)
    out.seek(0)
    filename_type = (type_exact or type_key or 'todos').lower().replace(' ', '_')
    filename = clean_filename(f"bajas_{filename_type}.pdf")
    return send_file(out, as_attachment=True, download_name=filename, mimetype='application/pdf')


@app.route('/disposals/export_general_excel', methods=['GET'])
def export_disposals_general_excel():
    ensure_db()
    service = request.args.get('service')
    status = request.args.get('status')
    period_id = request.args.get('period_id', type=int)
    all_rows = query_disposals(service=service, status=status, period_id=period_id)
    rows = [r for r in all_rows if normalize_disposal_type_key(r.get('type')) != 'CONTROL']
    if not rows:
        return jsonify({'error': 'No hay activos para exportar en reporte general'}), 400

    logo_path = get_hospital_logo_path()
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = 'Resumen General'

    grouped = {k: [] for k in DISPOSAL_TYPE_KEYS if k != 'CONTROL'}
    for r in rows:
        key = normalize_disposal_type_key(r.get('type'))
        if key in grouped:
            grouped[key].append(r)

    ws_summary.append(['REPORTE GENERAL DE BAJAS (SIN CONTROL)'])
    ws_summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws_summary['A1'].font = Font(bold=True, size=14, color='0B4F6C')
    ws_summary.append([
        f"Generado: {now_local_dt().strftime('%Y-%m-%d %H:%M')} | Servicio: {service or 'TODOS'} | Estado: {status or 'TODOS'}"
    ])
    ws_summary.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    ws_summary.append(['TIPO', 'CANTIDAD', 'TOTAL COSTO INICIAL', 'TOTAL SALDO POR DEPRECIAR', '% PARTICIPACION'])

    total_cost = sum(r['cost'] for r in rows)
    total_saldo = sum(r['saldo'] for r in rows)
    total_count = len(rows)
    for t in ['BIOMEDICO', 'MUEBLE Y ENSER', 'INDUSTRIAL', 'TECNOLOGICO']:
        sub = grouped.get(t, [])
        sub_cost = sum(r['cost'] for r in sub)
        sub_saldo = sum(r['saldo'] for r in sub)
        pct = round((len(sub) / total_count) * 100, 2) if total_count else 0
        ws_summary.append([t, len(sub), sub_cost, sub_saldo, pct])

    ws_summary.append(['TOTAL GENERAL', total_count, total_cost, total_saldo, 100 if total_count else 0])
    for c in ['A', 'B', 'C', 'D', 'E']:
        ws_summary.column_dimensions[c].width = [26, 12, 22, 24, 16][ord(c) - ord('A')]
    for row in ws_summary.iter_rows(min_row=3, max_row=ws_summary.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = Alignment(vertical='center', horizontal='center')
    for r in range(4, ws_summary.max_row + 1):
        ws_summary.cell(r, 3).number_format = '"$"#,##0'
        ws_summary.cell(r, 4).number_format = '"$"#,##0'
        ws_summary.cell(r, 5).number_format = '0.00"%"'

    for t in ['BIOMEDICO', 'MUEBLE Y ENSER', 'INDUSTRIAL', 'TECNOLOGICO']:
        ws = wb.create_sheet(title=t[:31])
        write_disposal_sheet(ws, f'Bajas {t}', grouped.get(t, []))
        add_logo_to_excel_sheet(ws, logo_path=logo_path)
    add_logo_to_excel_sheet(ws_summary, logo_path=logo_path)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    filename = clean_filename('bajas_general_sin_control.xlsx')
    return send_file(
        out,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/disposals/export_general_pdf', methods=['GET'])
def export_disposals_general_pdf():
    ensure_db()
    service = request.args.get('service')
    status = request.args.get('status')
    period_id = request.args.get('period_id', type=int)
    all_rows = query_disposals(service=service, status=status, period_id=period_id)
    rows = [r for r in all_rows if normalize_disposal_type_key(r.get('type')) != 'CONTROL']
    if not rows:
        return jsonify({'error': 'No hay activos para exportar en reporte general'}), 400

    grouped = {k: [] for k in ['BIOMEDICO', 'MUEBLE Y ENSER', 'INDUSTRIAL', 'TECNOLOGICO']}
    for r in rows:
        key = normalize_disposal_type_key(r.get('type'))
        if key in grouped:
            grouped[key].append(r)

    total = summarize_disposals(rows)
    logo_path = get_hospital_logo_path()
    out = BytesIO()
    doc = SimpleDocTemplate(
        out, pagesize=letter, leftMargin=16 * mm, rightMargin=16 * mm, topMargin=22 * mm, bottomMargin=14 * mm
    )
    styles = getSampleStyleSheet()
    story = []
    append_pdf_header_with_logo(
        story,
        'Reporte General de Bajas',
        f"Generado: {now_local_dt().strftime('%Y-%m-%d %H:%M')} | Servicio: {service or 'TODOS'} | Estado: {status or 'TODOS'}",
        include_logo=False,
    )
    story.append(Paragraph(
        f"<b>Total activos:</b> {total['count']} &nbsp;&nbsp; <b>Total costo inicial:</b> {money_text(total['total_cost'])} "
        f"&nbsp;&nbsp; <b>Total saldo por depreciar:</b> {money_text(total['total_saldo'])}",
        styles['Normal']
    ))
    story.append(Spacer(1, 8))

    res = [['TIPO', 'CANTIDAD', 'TOTAL COSTO', 'TOTAL SALDO']]
    for t in ['BIOMEDICO', 'MUEBLE Y ENSER', 'INDUSTRIAL', 'TECNOLOGICO']:
        sub = grouped[t]
        sum_sub = summarize_disposals(sub)
        res.append([t, sum_sub['count'], money_text(sum_sub['total_cost']), money_text(sum_sub['total_saldo'])])
    summary_table = Table(res, colWidths=[48*mm, 26*mm, 48*mm, 48*mm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#EAF4FA')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#C8D8E4')),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 10))

    chart_d = Drawing(170 * mm, 45 * mm)
    chart = HorizontalBarChart()
    chart.x = 22
    chart.y = 4
    chart.width = 132 * mm
    chart.height = 34 * mm
    chart.data = [[len(grouped['BIOMEDICO']), len(grouped['MUEBLE Y ENSER']), len(grouped['INDUSTRIAL']), len(grouped['TECNOLOGICO'])]]
    chart.categoryAxis.categoryNames = ['Biomedico', 'Mueble y enser', 'Industrial', 'Tecnologico']
    chart.valueAxis.valueMin = 0
    chart.bars[0].fillColor = colors.HexColor('#1E88E5')
    chart_d.add(chart)
    chart_d.add(String(6, 42 * mm, 'Distribucion de activos por tipo', fontSize=9, fillColor=colors.HexColor('#0B4F6C')))
    story.append(chart_d)

    for t in ['BIOMEDICO', 'MUEBLE Y ENSER', 'INDUSTRIAL', 'TECNOLOGICO']:
        sub = grouped[t]
        if not sub:
            continue
        story.append(PageBreak())
        story.append(Paragraph(f'Detalle {t}', ParagraphStyle(
            'Sec', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=13, textColor=colors.HexColor('#0B4F6C')
        )))
        sum_sub = summarize_disposals(sub)
        story.append(Paragraph(
            f"<b>Activos:</b> {sum_sub['count']} &nbsp;&nbsp; <b>Costo:</b> {money_text(sum_sub['total_cost'])} "
            f"&nbsp;&nbsp; <b>Saldo:</b> {money_text(sum_sub['total_saldo'])}",
            styles['Normal']
        ))
        story.append(Spacer(1, 6))
        detail = [[
            pdf_cell('COD ACTIVO FIJO', styles, bold=True, align='CENTER', size=7.5),
            pdf_cell('DESCRIPCION', styles, bold=True, align='CENTER', size=7.5),
            pdf_cell('COSTO INICIAL', styles, bold=True, align='CENTER', size=7.5),
            pdf_cell('SALDO POR DEPRECIAR', styles, bold=True, align='CENTER', size=7.5),
            pdf_cell('FECHA ADQUISICION', styles, bold=True, align='CENTER', size=7.5),
            pdf_cell('MOTIVO DE BAJA', styles, bold=True, align='CENTER', size=7.5),
        ]]
        for r in sub:
            detail.append([
                pdf_cell(r['code'], styles, align='CENTER'),
                pdf_cell(r['name'], styles),
                pdf_cell(money_text(r['cost']), styles, align='RIGHT'),
                pdf_cell(money_text(r['saldo']), styles, align='RIGHT'),
                pdf_cell(r['date'], styles, align='CENTER'),
                pdf_cell(r['reason'], styles),
            ])
        t_detail = Table(detail, colWidths=[16*mm, 58*mm, 22*mm, 22*mm, 18*mm, 48*mm], repeatRows=1)
        t_detail.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F3F9FD')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#D4E2EC')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(t_detail)

    page_header = make_pdf_page_header(logo_path)
    doc.build(story, onFirstPage=page_header, onLaterPages=page_header)
    out.seek(0)
    filename = clean_filename('bajas_general_sin_control.pdf')
    return send_file(out, as_attachment=True, download_name=filename, mimetype='application/pdf')


@app.route('/disposals/export_general_control_excel', methods=['GET'])
def export_disposals_general_control_excel():
    ensure_db()
    service = request.args.get('service')
    status = request.args.get('status')
    period_id = request.args.get('period_id', type=int)
    all_rows = query_disposals(service=service, status=status, period_id=period_id)
    rows = [r for r in all_rows if normalize_disposal_type_key(r.get('type')) == 'CONTROL']

    grouped = {}
    for r in rows:
        key = str(r.get('type') or 'CONTROL - OTROS').strip().upper()
        grouped.setdefault(key, []).append(r)

    logo_path = get_hospital_logo_path()
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = 'Resumen Control'
    ws_summary.append(['REPORTE GENERAL DE BAJAS - ACTIVOS DE CONTROL'])
    ws_summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws_summary['A1'].font = Font(bold=True, size=14, color='9A5F00')
    ws_summary.append([
        f"Generado: {now_local_dt().strftime('%Y-%m-%d %H:%M')} | Servicio: {service or 'TODOS'} | Estado: {status or 'TODOS'}"
    ])
    ws_summary.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    ws_summary.append([
        'Nota: los activos de control no se deprecian; su saldo contable suele coincidir con el costo inicial.'
    ])
    ws_summary.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5)
    ws_summary['A3'].font = Font(bold=True, color='9A5F00')
    ws_summary.append(['SUBTIPO CONTROL', 'CANTIDAD', 'TOTAL COSTO INICIAL', 'TOTAL SALDO CONTABLE', '% PARTICIPACION'])

    total_cost = sum(r['cost'] for r in rows)
    total_saldo = sum(r['saldo'] for r in rows)
    total_count = len(rows)
    for key in sorted(grouped.keys()):
        sub = grouped.get(key, [])
        sub_cost = sum(r['cost'] for r in sub)
        sub_saldo = sum(r['saldo'] for r in sub)
        pct = round((len(sub) / total_count) * 100, 2) if total_count else 0
        ws_summary.append([key, len(sub), sub_cost, sub_saldo, pct])

    ws_summary.append(['TOTAL GENERAL CONTROL', total_count, total_cost, total_saldo, 100 if total_count else 0])
    for c in ['A', 'B', 'C', 'D', 'E']:
        ws_summary.column_dimensions[c].width = [34, 12, 22, 24, 16][ord(c) - ord('A')]
    for row in ws_summary.iter_rows(min_row=4, max_row=ws_summary.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = Alignment(vertical='center', horizontal='center')
    for r in range(5, ws_summary.max_row + 1):
        ws_summary.cell(r, 3).number_format = '"$"#,##0'
        ws_summary.cell(r, 4).number_format = '"$"#,##0'
        ws_summary.cell(r, 5).number_format = '0.00"%"'

    for key in sorted(grouped.keys()):
        ws = wb.create_sheet(title=key[:31])
        write_disposal_sheet(
            ws,
            f'Bajas {key}',
            grouped.get(key, []),
            saldo_header='SALDO CONTABLE (NO DEPRECIABLE)',
            note_text='Nota: activo de control no depreciable (vida util 0 o marcado como no depreciable).',
        )
        add_logo_to_excel_sheet(ws, logo_path=logo_path)
    add_logo_to_excel_sheet(ws_summary, logo_path=logo_path)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    filename = clean_filename('bajas_general_control.xlsx')
    return send_file(
        out,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/disposals/export_general_control_pdf', methods=['GET'])
def export_disposals_general_control_pdf():
    ensure_db()
    service = request.args.get('service')
    status = request.args.get('status')
    period_id = request.args.get('period_id', type=int)
    all_rows = query_disposals(service=service, status=status, period_id=period_id)
    rows = [r for r in all_rows if normalize_disposal_type_key(r.get('type')) == 'CONTROL']

    grouped = {}
    for r in rows:
        key = str(r.get('type') or 'CONTROL - OTROS').strip().upper()
        grouped.setdefault(key, []).append(r)

    total = summarize_disposals(rows)
    logo_path = get_hospital_logo_path()
    out = BytesIO()
    doc = SimpleDocTemplate(
        out, pagesize=letter, leftMargin=16 * mm, rightMargin=16 * mm, topMargin=22 * mm, bottomMargin=14 * mm
    )
    styles = getSampleStyleSheet()
    story = []
    append_pdf_header_with_logo(
        story,
        'Reporte General de Bajas - Activos de Control',
        f"Generado: {now_local_dt().strftime('%Y-%m-%d %H:%M')} | Servicio: {service or 'TODOS'} | Estado: {status or 'TODOS'}",
        include_logo=False,
    )
    story.append(Paragraph(
        f"<b>Total activos control:</b> {total['count']} &nbsp;&nbsp; <b>Total costo inicial:</b> {money_text(total['total_cost'])} "
        f"&nbsp;&nbsp; <b>Total saldo contable:</b> {money_text(total['total_saldo'])}",
        styles['Normal']
    ))
    control_pct = round((len(rows) / len(all_rows)) * 100, 2) if all_rows else 0
    story.append(Paragraph(
        f"<b>Participacion de control sobre bajas filtradas:</b> {control_pct}%",
        styles['Normal']
    ))
    story.append(Paragraph(
        '<b>Nota:</b> Los activos de control no se deprecian; por politica contable su saldo contable suele coincidir con el costo inicial.',
        ParagraphStyle('CtrlGenNote', parent=styles['Normal'], textColor=colors.HexColor('#9A5F00'), fontSize=9)
    ))
    story.append(Spacer(1, 8))

    res = [['SUBTIPO CONTROL', 'CANTIDAD', 'TOTAL COSTO', 'TOTAL SALDO CONTABLE']]
    for key in sorted(grouped.keys()):
        sub = grouped[key]
        sum_sub = summarize_disposals(sub)
        res.append([key, sum_sub['count'], money_text(sum_sub['total_cost']), money_text(sum_sub['total_saldo'])])
    summary_table = Table(res, colWidths=[62*mm, 22*mm, 44*mm, 44*mm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFF4DE')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#E2C998')),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
    ]))
    story.append(summary_table)
    if not rows:
        story.append(Spacer(1, 8))
        story.append(Paragraph(
            'No hay activos de control para los filtros seleccionados. Se genera reporte en blanco para control documental.',
            ParagraphStyle('CtrlEmpty', parent=styles['Normal'], textColor=colors.HexColor('#7A4A00'), fontSize=9)
        ))
    story.append(Spacer(1, 10))

    chart_keys = sorted(grouped.keys())
    chart_vals = [len(grouped[k]) for k in chart_keys]
    chart_d = Drawing(170 * mm, 45 * mm)
    chart = HorizontalBarChart()
    chart.x = 22
    chart.y = 4
    chart.width = 132 * mm
    chart.height = 34 * mm
    chart.data = [chart_vals] if chart_vals else [[0]]
    chart.categoryAxis.categoryNames = [k.replace('CONTROL - ', '').title() for k in chart_keys] if chart_keys else ['Sin datos']
    chart.valueAxis.valueMin = 0
    chart.bars[0].fillColor = colors.HexColor('#D97706')
    chart_d.add(chart)
    chart_d.add(String(6, 42 * mm, 'Distribucion de activos de control por subtipo', fontSize=9, fillColor=colors.HexColor('#9A5F00')))
    story.append(chart_d)

    for key in sorted(grouped.keys()):
        sub = grouped[key]
        if not sub:
            continue
        story.append(PageBreak())
        story.append(Paragraph(f'Detalle {key}', ParagraphStyle(
            'SecCtrl', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=13, textColor=colors.HexColor('#9A5F00')
        )))
        sum_sub = summarize_disposals(sub)
        story.append(Paragraph(
            f"<b>Activos:</b> {sum_sub['count']} &nbsp;&nbsp; <b>Costo:</b> {money_text(sum_sub['total_cost'])} "
            f"&nbsp;&nbsp; <b>Saldo:</b> {money_text(sum_sub['total_saldo'])}",
            styles['Normal']
        ))
        story.append(Spacer(1, 6))
        detail = [[
            pdf_cell('COD ACTIVO FIJO', styles, bold=True, align='CENTER', size=7.5),
            pdf_cell('DESCRIPCION', styles, bold=True, align='CENTER', size=7.5),
            pdf_cell('COSTO INICIAL', styles, bold=True, align='CENTER', size=7.5),
            pdf_cell('SALDO POR DEPRECIAR', styles, bold=True, align='CENTER', size=7.5),
            pdf_cell('FECHA ADQUISICION', styles, bold=True, align='CENTER', size=7.5),
            pdf_cell('MOTIVO DE BAJA', styles, bold=True, align='CENTER', size=7.5),
        ]]
        for r in sub:
            detail.append([
                pdf_cell(r['code'], styles, align='CENTER'),
                pdf_cell(r['name'], styles),
                pdf_cell(money_text(r['cost']), styles, align='RIGHT'),
                pdf_cell(money_text(r['saldo']), styles, align='RIGHT'),
                pdf_cell(r['date'], styles, align='CENTER'),
                pdf_cell(r['reason'], styles),
            ])
        t_detail = Table(detail, colWidths=[16*mm, 58*mm, 22*mm, 22*mm, 18*mm, 48*mm], repeatRows=1)
        t_detail.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFF7E8')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#E4D1A9')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(t_detail)

    page_header = make_pdf_page_header(logo_path)
    doc.build(story, onFirstPage=page_header, onLaterPages=page_header)
    out.seek(0)
    filename = clean_filename('bajas_general_control.pdf')
    return send_file(out, as_attachment=True, download_name=filename, mimetype='application/pdf')


@app.route('/maintenance/reclassify', methods=['POST'])
def reclassify_assets():
    ensure_db()
    data = request.get_json() or {}
    service = (data.get('service') or '').strip()
    only_disposals = bool(data.get('only_disposals'))

    q = Asset.query
    if service:
        q = q.filter(Asset.nom_ccos == service)
    if only_disposals:
        q = q.join(AssetDisposal, AssetDisposal.asset_id == Asset.id)

    assets = q.all()
    updated = 0
    for asset in assets:
        old = (asset.tipo_activo_cache or '').strip()
        new = refresh_asset_type_cache(asset)
        if old != new:
            updated += 1

    db.session.commit()
    return jsonify({
        'ok': True,
        'total_processed': len(assets),
        'updated': updated,
        'service': service or None,
        'only_disposals': only_disposals,
        'executed_at': now_iso(),
    })


@app.route('/disposals', methods=['POST'])
def create_disposal():
    ensure_db()
    data = request.get_json() or {}
    code = data.get('code')
    reason = (data.get('reason') or '').strip()
    requested_by = (data.get('requested_by') or '').strip() or 'unknown'
    period_id = parse_int(data.get('period_id'))
    if not code:
        return jsonify({'error': 'Debe enviar codigo de activo'}), 400
    if not period_id:
        return jsonify({'error': 'Debes seleccionar un periodo para registrar la baja'}), 400

    period = InventoryPeriod.query.get(period_id)
    if not period:
        return jsonify({'error': 'Periodo no encontrado'}), 404

    asset, _ = get_asset_by_code(code)
    if not asset:
        return jsonify({'error': 'Activo no encontrado'}), 404

    disposal = AssetDisposal.query.filter_by(asset_id=asset.id).first()
    now_iso_value = now_iso()
    if not disposal:
        disposal = AssetDisposal(
            asset_id=asset.id,
            period_id=period.id,
            status='Pendiente baja',
            reason=reason,
            requested_by=requested_by,
            requested_at=now_iso_value,
        )
        db.session.add(disposal)
    else:
        disposal.period_id = period.id
        disposal.status = 'Pendiente baja'
        disposal.reason = reason or disposal.reason
        disposal.requested_by = requested_by
        disposal.requested_at = now_iso_value
        disposal.reviewed_by = None
        disposal.reviewed_at = None
        disposal.review_notes = None

    db.session.commit()
    return jsonify({'disposal': disposal.to_dict(asset=asset)})


@app.route('/disposals/<int:disposal_id>', methods=['PATCH'])
def update_disposal(disposal_id):
    ensure_db()
    disposal = AssetDisposal.query.get(disposal_id)
    if not disposal:
        return jsonify({'error': 'Registro de baja no encontrado'}), 404

    data = request.get_json() or {}
    new_status = (data.get('status') or '').strip()
    reason_raw = data.get('reason', None)
    review_notes = (data.get('review_notes') or '').strip() or None
    reviewed_by = (data.get('reviewed_by') or '').strip() or 'unknown'
    allowed = {'Pendiente baja', 'Aprobada para baja', 'Rechazada'}
    if new_status and new_status not in allowed:
        return jsonify({'error': 'Estado de baja invalido'}), 400
    if reason_raw is not None:
        reason_txt = str(reason_raw).strip()
        if not reason_txt:
            return jsonify({'error': 'El motivo de baja no puede quedar vacio'}), 400
        disposal.reason = reason_txt
    if new_status:
        disposal.status = new_status
        disposal.reviewed_by = reviewed_by
        disposal.reviewed_at = now_iso()
        disposal.review_notes = review_notes

    db.session.commit()
    asset = Asset.query.get(disposal.asset_id)
    return jsonify({'disposal': disposal.to_dict(asset=asset)})


@app.route('/disposals/<int:disposal_id>', methods=['DELETE'])
def delete_disposal(disposal_id):
    ensure_db()
    disposal = AssetDisposal.query.get(disposal_id)
    if not disposal:
        return jsonify({'error': 'Registro de baja no encontrado'}), 404

    asset = Asset.query.get(disposal.asset_id)
    if asset:
        asset.estado_inventario = 'Pendiente verificacion'
        asset.fecha_verificacion = now_iso()
        asset.usuario_verificador = 'usuario_movil'

    db.session.delete(disposal)
    db.session.commit()
    return jsonify({'ok': True, 'asset_id': disposal.asset_id})


@app.route('/dashboard/summary', methods=['GET'])
def dashboard_summary():
    ensure_db()
    service = request.args.get('service')
    run_id = request.args.get('run_id', type=int)
    period_id = request.args.get('period_id', type=int)
    payload, error = build_dashboard_payload(service=service, run_id=run_id, period_id=period_id)
    if error:
        return jsonify({'error': error}), 404
    return jsonify(payload)


@app.route('/dashboard/report_pdf', methods=['GET'])
def dashboard_report_pdf():
    ensure_db()
    service = request.args.get('service')
    run_id = request.args.get('run_id', type=int)
    period_id = request.args.get('period_id', type=int)
    payload, error = build_dashboard_payload(service=service, run_id=run_id, period_id=period_id)
    if error:
        return jsonify({'error': error}), 404

    out = BytesIO()
    logo_path = get_hospital_logo_path()
    doc = SimpleDocTemplate(
        out,
        pagesize=letter,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=16 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    brand_blue = colors.HexColor('#0A6FB3')
    brand_blue_dark = colors.HexColor('#07507F')
    brand_green = colors.HexColor('#1E9E57')
    brand_yellow = colors.HexColor('#F2C94C')
    brand_red = colors.HexColor('#C0392B')
    title_style = ParagraphStyle(
        'DashTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=21,
        textColor=brand_blue_dark,
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        'DashSubTitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#51606F'),
        leading=12,
        spaceAfter=4,
    )
    section_style = ParagraphStyle(
        'DashSection',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=13,
        textColor=brand_blue_dark,
        spaceBefore=6,
        spaceAfter=3,
    )
    cell_style = ParagraphStyle(
        'Cell',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
    )

    k = payload['kpis']
    meta = payload.get('meta', {})
    run_name = (meta.get('run') or {}).get('name') if meta.get('run') else 'Sin jornada'
    period_name = (meta.get('period') or {}).get('name') if meta.get('period') else 'Sin periodo'
    service_filter = meta.get('service_filter') or 'TODOS'
    generated_at = meta.get('generated_at') or now_iso()

    story = []
    hero = Table([[
        Paragraph(
            '<font color="white"><b>Dashboard Institucional de Inventario</b></font><br/>'
            '<font color="white">Hospital Francisco de Paula Santander E.S.E.</font>',
            ParagraphStyle(
                'HeroTitle',
                parent=styles['Normal'],
                fontName='Helvetica-Bold',
                fontSize=15,
                leading=18,
            )
        )
    ]], colWidths=[182 * mm])
    hero.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), brand_blue),
        ('BOX', (0, 0), (-1, -1), 1.0, brand_blue_dark),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    story.append(hero)
    story.append(Spacer(1, 5))
    story.append(Paragraph('Reporte Ejecutivo Dashboard Inventario', title_style))
    story.append(Paragraph(
        f'Fecha: {generated_at} &nbsp;&nbsp;|&nbsp;&nbsp; Periodo: {period_name} '
        f'&nbsp;&nbsp;|&nbsp;&nbsp; Jornada: {run_name} '
        f'&nbsp;&nbsp;|&nbsp;&nbsp; Servicio: {service_filter}',
        subtitle_style
    ))

    narrative = build_executive_narrative(payload)
    plan = build_executive_action_plan(payload)
    story.append(Paragraph('Objetivo General', section_style))
    story.append(Paragraph(narrative.get('objetivo_general', ''), subtitle_style))
    story.append(Paragraph('Objetivos Especificos', section_style))
    obj_lines = '<br/>'.join([f'- {x}' for x in narrative.get('objetivos_especificos', [])])
    story.append(Paragraph(obj_lines or 'Sin objetivos definidos.', subtitle_style))
    story.append(Paragraph('Resumen Ejecutivo', section_style))
    story.append(Paragraph(narrative.get('resumen', ''), subtitle_style))
    story.append(Paragraph('Interpretacion Contextual', section_style))
    int_lines = '<br/>'.join([f'- {x}' for x in narrative.get('interpretacion', [])])
    story.append(Paragraph(int_lines or 'Sin interpretacion disponible.', subtitle_style))
    story.append(Spacer(1, 3))

    risk_color = {
        'ALTO': '#B42318',
        'MEDIO': '#B26A00',
        'BAJO': '#0D7A52',
    }.get(plan.get('risk_level'), '#0D7A52')
    semaforo = Table([[
        Paragraph(f"<b>Semaforo de riesgo:</b> {plan.get('risk_level', 'N/D')}<br/>{plan.get('risk_reason', '')}", styles['Normal'])
    ]], colWidths=[168 * mm])
    semaforo.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#F8FBFF')),
        ('TEXTCOLOR', (0, 0), (0, 0), colors.HexColor(risk_color)),
        ('BOX', (0, 0), (-1, -1), 1.0, colors.HexColor(risk_color)),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(semaforo)
    story.append(Spacer(1, 4))

    kpi_data = [[
        Paragraph('<b>Total activos</b><br/>{}'.format(k.get('total', 0)), styles['Normal']),
        Paragraph('<b>Encontrados</b><br/>{} ({}%)'.format(k.get('found', 0), k.get('found_pct', 0)), styles['Normal']),
        Paragraph('<b>No encontrados</b><br/>{} ({}%)'.format(k.get('not_found', 0), k.get('not_found_pct', 0)), styles['Normal']),
        Paragraph('<b>Pendientes</b><br/>{}'.format(k.get('pending', 0)), styles['Normal']),
    ]]
    kpi_table = Table(kpi_data, colWidths=[42 * mm, 42 * mm, 42 * mm, 42 * mm])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#E8F2FC')),
        ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#E9F7EF')),
        ('BACKGROUND', (2, 0), (2, 0), colors.HexColor('#FFF0F0')),
        ('BACKGROUND', (3, 0), (3, 0), colors.HexColor('#FFF7E6')),
        ('BOX', (0, 0), (-1, -1), 0.6, colors.HexColor('#BFD5E3')),
        ('INNERGRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#D7E5EE')),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 4))

    coverage = payload.get('coverage', {})
    coverage_data = [[
        Paragraph('<b>Base total activos</b><br/>{}'.format(coverage.get('base_total_assets', 0)), styles['Normal']),
        Paragraph('<b>Activos en alcance periodo/jornada</b><br/>{} ({}%)'.format(
            coverage.get('scope_assets', 0), coverage.get('scope_assets_pct', 0)
        ), styles['Normal']),
        Paragraph('<b>Base fuera de alcance</b><br/>{}'.format(coverage.get('base_not_in_scope_assets', 0)), styles['Normal']),
        Paragraph('<b>Cobertura valor</b><br/>{} / {} ({}%)'.format(
            money_text(to_number(coverage.get('scope_value', 0))),
            money_text(to_number(coverage.get('base_total_value', 0))),
            coverage.get('scope_value_pct', 0),
        ), styles['Normal']),
    ]]
    coverage_table = Table(coverage_data, colWidths=[42 * mm, 42 * mm, 42 * mm, 42 * mm])
    coverage_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (3, 0), colors.HexColor('#E9F7EF')),
        ('BOX', (0, 0), (-1, -1), 0.6, colors.HexColor('#C7E5C9')),
        ('INNERGRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#DDF0DF')),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(coverage_table)
    story.append(Spacer(1, 4))

    financial = payload.get('financial', {})
    financial_data = [[
        Paragraph('<b>Valor total inventario</b><br/>{}'.format(money_text(to_number(financial.get('total_value', 0)))), styles['Normal']),
        Paragraph('<b>Valor encontrado</b><br/>{}'.format(money_text(to_number(financial.get('found_value', 0)))), styles['Normal']),
        Paragraph('<b>Valor no encontrado</b><br/>{} ({}%)'.format(
            money_text(to_number(financial.get('not_found_value', 0))),
            financial.get('not_found_value_pct', 0)
        ), styles['Normal']),
        Paragraph('<b>Valor critico no encontrado</b><br/>{}'.format(
            money_text(to_number(financial.get('critical_not_found_value', 0)))
        ), styles['Normal']),
    ]]
    financial_table = Table(financial_data, colWidths=[42 * mm, 42 * mm, 42 * mm, 42 * mm])
    financial_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (3, 0), colors.HexColor('#FFF6E8')),
        ('BOX', (0, 0), (-1, -1), 0.6, colors.HexColor('#E6D0BC')),
        ('INNERGRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#F0E1D2')),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(financial_table)
    story.append(Spacer(1, 4))

    story.append(Paragraph('Distribucion del estado del inventario', section_style))
    pie_drawing = Drawing(170 * mm, 58 * mm)
    pie = Pie()
    pie.x = 18
    pie.y = 6
    pie.width = 70 * mm
    pie.height = 46 * mm
    pie.slices.strokeWidth = 0.5
    found_v = max(0, int(k.get('found', 0) or 0))
    not_found_v = max(0, int(k.get('not_found', 0) or 0))
    pending_v = max(0, int(k.get('pending', 0) or 0))
    total_v = found_v + not_found_v + pending_v
    pie.data = [found_v, not_found_v, pending_v] if total_v > 0 else [1]
    pie.labels = ['Encontrados', 'No encontrados', 'Pendientes'] if total_v > 0 else ['Sin datos']
    if total_v > 0:
        pie.slices[0].fillColor = brand_green
        pie.slices[1].fillColor = brand_red
        pie.slices[2].fillColor = brand_yellow
        pie.slices[1].popout = 2
    else:
        pie.slices[0].fillColor = colors.HexColor('#CBD5E1')
    pie_drawing.add(pie)
    pie_drawing.add(String(95 * mm, 40 * mm, f'Encontrados: {found_v} ({k.get("found_pct", 0)}%)', fontSize=9, fillColor=brand_green))
    pie_drawing.add(String(95 * mm, 30 * mm, f'No encontrados: {not_found_v} ({k.get("not_found_pct", 0)}%)', fontSize=9, fillColor=brand_red))
    pie_drawing.add(String(95 * mm, 20 * mm, f'Pendientes: {pending_v}', fontSize=9, fillColor=colors.HexColor('#9A6700')))
    story.append(pie_drawing)
    story.append(Spacer(1, 4))

    insights = payload.get('insights', [])
    story.append(Paragraph('Mensajes Clave para Alta Gerencia', section_style))
    if insights:
        bullets = ''.join([f'• {i}<br/>' for i in insights])
        story.append(Paragraph(bullets, subtitle_style))
    else:
        story.append(Paragraph('Sin hallazgos relevantes para este corte.', subtitle_style))
    story.append(Spacer(1, 3))

    def make_chart(title, rows, top_n=15):
        drawing = Drawing(520, 200)
        drawing.add(String(0, 186, title, fontName='Helvetica-Bold', fontSize=11, fillColor=brand_blue_dark))
        if not rows:
            drawing.add(String(0, 162, 'Sin datos', fontName='Helvetica', fontSize=9, fillColor=colors.HexColor('#7A8794')))
            return drawing

        selected = rows[:top_n]
        labels = [str(r.get('name', ''))[:46] for r in selected]
        values = [float(r.get('total', 0) or 0) for r in selected]

        chart = HorizontalBarChart()
        chart.x = 110
        chart.y = 10
        chart.height = 155
        chart.width = 390
        chart.data = [values]
        chart.categoryAxis.categoryNames = labels
        chart.categoryAxis.labels.fontName = 'Helvetica'
        chart.categoryAxis.labels.fontSize = 7
        chart.categoryAxis.labels.boxAnchor = 'e'
        chart.categoryAxis.labels.dx = -4
        chart.valueAxis.valueMin = 0
        chart.valueAxis.labels.fontSize = 7
        chart.valueAxis.visibleGrid = 1
        chart.valueAxis.gridStrokeColor = colors.HexColor('#DCE7EE')
        chart.bars[0].fillColor = brand_blue
        chart.barSpacing = 2
        chart.groupSpacing = 4
        palette = [brand_blue, colors.HexColor('#118AB2'), brand_green, colors.HexColor('#2FAE66'), brand_yellow]
        for i in range(len(values)):
            try:
                chart.bars[(0, i)].fillColor = palette[i % len(palette)]
            except Exception:
                pass
        drawing.add(chart)
        return drawing

    story.append(Paragraph('Visualizaciones', section_style))
    story.append(make_chart('Activos por servicio (Top 10)', payload.get('by_service', []), top_n=10))
    story.append(Spacer(1, 3))
    story.append(make_chart('Activos por tipo de equipo (Top 10)', payload.get('by_type', []), top_n=10))
    story.append(Spacer(1, 3))
    story.append(make_chart('Activos por área', payload.get('by_area', []), top_n=15))
    story.append(Spacer(1, 4))

    story.append(Paragraph('Activos Criticos y Costosos No Encontrados', section_style))
    critical_rows = payload.get('critical_not_found', [])
    if not critical_rows:
        story.append(Paragraph('No se identificaron activos criticos no encontrados en este corte.', subtitle_style))
    else:
        critical_data = [[
            Paragraph('<b>Codigo</b>', cell_style),
            Paragraph('<b>Activo</b>', cell_style),
            Paragraph('<b>Servicio</b>', cell_style),
            Paragraph('<b>Tipo</b>', cell_style),
            Paragraph('<b>Valor libro</b>', cell_style),
            Paragraph('<b>Criticidad</b>', cell_style),
            Paragraph('<b>Motivo</b>', cell_style),
        ]]
        for item in critical_rows:
            critical_data.append([
                Paragraph(str(item.get('code', '')), cell_style),
                Paragraph(str(item.get('name', '')), cell_style),
                Paragraph(str(item.get('service', '')), cell_style),
                Paragraph(str(item.get('type', '')), cell_style),
                Paragraph(money_text(to_number(item.get('value', 0))), cell_style),
                Paragraph(str(item.get('critical_score', 0)), cell_style),
                Paragraph(str(item.get('critical_reasons', '')), cell_style),
            ])
        critical_table = Table(
            critical_data,
            colWidths=[20 * mm, 46 * mm, 26 * mm, 28 * mm, 20 * mm, 16 * mm, 24 * mm],
            repeatRows=1
        )
        critical_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7A1F1F')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (4, 1), (5, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOX', (0, 0), (-1, -1), 0.6, colors.HexColor('#D6B5B5')),
            ('INNERGRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#E7CFCF')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FFF7F7')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(critical_table)
    story.append(Spacer(1, 4))

    story.append(Paragraph('Anexo Operativo: Activos No Encontrados', section_style))
    not_found_rows = payload.get('not_found_assets', [])
    if not not_found_rows:
        story.append(Paragraph('No hay activos no encontrados para este corte.', subtitle_style))
    else:
        annex_cap = 120
        not_found_rows = not_found_rows[:annex_cap]
        if payload.get('not_found_assets_capped') or payload.get('not_found_assets_total', 0) > annex_cap:
            story.append(Paragraph(
                f"Se muestran los primeros {len(not_found_rows)} de {payload.get('not_found_assets_total', len(not_found_rows))} activos no encontrados. El detalle completo se gestiona en exportes operativos.",
                subtitle_style
            ))
        nf_data = [[
            Paragraph('<b>Codigo</b>', cell_style),
            Paragraph('<b>Activo</b>', cell_style),
            Paragraph('<b>Servicio</b>', cell_style),
            Paragraph('<b>Responsable</b>', cell_style),
            Paragraph('<b>Ubicacion</b>', cell_style),
            Paragraph('<b>Valor libro</b>', cell_style),
        ]]
        for item in not_found_rows:
            nf_data.append([
                Paragraph(str(item.get('code', '')), cell_style),
                Paragraph(str(item.get('name', '')), cell_style),
                Paragraph(str(item.get('service', '')), cell_style),
                Paragraph(str(item.get('responsible', '')), cell_style),
                Paragraph(str(item.get('location', '')), cell_style),
                Paragraph(money_text(to_number(item.get('value', 0))), cell_style),
            ])
        nf_table = Table(
            nf_data,
            colWidths=[20 * mm, 48 * mm, 36 * mm, 34 * mm, 36 * mm, 22 * mm],
            repeatRows=1,
        )
        nf_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#922020')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (5, 1), (5, -1), 'RIGHT'),
            ('BOX', (0, 0), (-1, -1), 0.6, colors.HexColor('#D5B3B3')),
            ('INNERGRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#E9D7D7')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FFF8F8')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(nf_table)
    story.append(Spacer(1, 4))

    def value_table(title, rows):
        story.append(Paragraph(title, section_style))
        if not rows:
            story.append(Paragraph('Sin datos para esta seccion.', subtitle_style))
            story.append(Spacer(1, 3))
            return
        data = [[
            Paragraph('<b>Nombre</b>', cell_style),
            Paragraph('<b>Valor no encontrado</b>', cell_style),
        ]]
        for r in rows:
            data.append([
                Paragraph(str(r.get('name', '')), cell_style),
                Paragraph(money_text(to_number(r.get('not_found_value', 0))), cell_style),
            ])
        table = Table(data, colWidths=[120 * mm, 60 * mm], repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#244B5A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('BOX', (0, 0), (-1, -1), 0.6, colors.HexColor('#BFD5E3')),
            ('INNERGRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#D7E5EE')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7FBFD')]),
        ]))
        story.append(table)
        story.append(Spacer(1, 4))

    value_table('Impacto Economico de No Encontrados por Servicio', payload.get('top_not_found_by_service_value', []))
    value_table('Impacto Economico de No Encontrados por Tipo de Equipo', payload.get('top_not_found_by_type_value', []))
    story.append(PageBreak())

    def section_table(title, rows):
        story.append(Paragraph(title, section_style))
        if not rows:
            story.append(Paragraph('Sin datos para esta sección.', subtitle_style))
            story.append(Spacer(1, 3))
            return

        data = [[
            Paragraph('<b>Nombre</b>', cell_style),
            Paragraph('<b>Total</b>', cell_style),
            Paragraph('<b>Encontrados</b>', cell_style),
            Paragraph('<b>No encontrados</b>', cell_style),
            Paragraph('<b>Pendientes</b>', cell_style),
            Paragraph('<b>% Encontrados</b>', cell_style),
        ]]
        for r in rows:
            data.append([
                Paragraph(str(r.get('name', '')), cell_style),
                Paragraph(str(r.get('total', 0)), cell_style),
                Paragraph(str(r.get('found', 0)), cell_style),
                Paragraph(str(r.get('not_found', 0)), cell_style),
                Paragraph(str(r.get('pending', 0)), cell_style),
                Paragraph(str(r.get('found_pct', 0)), cell_style),
            ])

        table = Table(data, colWidths=[77 * mm, 16 * mm, 20 * mm, 25 * mm, 19 * mm, 22 * mm], repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#07507F')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOX', (0, 0), (-1, -1), 0.6, colors.HexColor('#BFD5E3')),
            ('INNERGRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#D7E5EE')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7FBFD')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(table)
        story.append(Spacer(1, 4))

    section_table('Detalle Completo por Servicio', payload.get('by_service', []))
    section_table('Detalle Completo por Tipo de Equipo', payload.get('by_type', []))
    section_table('Detalle completo por área', payload.get('by_area', []))

    story.append(PageBreak())
    story.append(Paragraph('Conclusion Final', section_style))
    story.append(Paragraph(build_executive_conclusion(payload), subtitle_style))

    page_header = make_pdf_page_header(
        logo_path=None,
        right_image_path=logo_path,
        right_width_mm=14,
        right_height_mm=14,
        right_top_mm=16,
    )
    doc.build(story, onFirstPage=page_header, onLaterPages=page_header)
    out.seek(0)
    base_name = run_name if run_name and run_name != 'Sin jornada' else service_filter
    safe_name = clean_filename(base_name)
    filename = f'dashboard_{safe_name}.pdf'
    return send_file(out, download_name=filename, as_attachment=True, mimetype='application/pdf')


@app.route('/dashboard/compare_periods', methods=['GET'])
def dashboard_compare_periods():
    ensure_db()
    period_a = request.args.get('period_a', type=int)
    period_b = request.args.get('period_b', type=int)
    service = (request.args.get('service') or '').strip() or None
    if not period_a or not period_b:
        return jsonify({'error': 'Debes indicar ambos periodos para comparar'}), 400

    payload_a, err_a = build_dashboard_payload(service=service, period_id=period_a)
    if err_a:
        return jsonify({'error': err_a}), 400
    payload_b, err_b = build_dashboard_payload(service=service, period_id=period_b)
    if err_b:
        return jsonify({'error': err_b}), 400

    ka = payload_a.get('kpis', {})
    kb = payload_b.get('kpis', {})
    pa = (payload_a.get('meta', {}).get('period') or {})
    pb = (payload_b.get('meta', {}).get('period') or {})
    response = {
        'period_a': {
            'id': pa.get('id'),
            'name': pa.get('name') or f'Periodo {period_a}',
            'total': ka.get('total', 0),
            'found': ka.get('found', 0),
            'not_found': ka.get('not_found', 0),
            'found_pct': ka.get('found_pct', 0),
            'not_found_pct': ka.get('not_found_pct', 0),
        },
        'period_b': {
            'id': pb.get('id'),
            'name': pb.get('name') or f'Periodo {period_b}',
            'total': kb.get('total', 0),
            'found': kb.get('found', 0),
            'not_found': kb.get('not_found', 0),
            'found_pct': kb.get('found_pct', 0),
            'not_found_pct': kb.get('not_found_pct', 0),
        },
        'delta': {
            'total': kb.get('total', 0) - ka.get('total', 0),
            'found': kb.get('found', 0) - ka.get('found', 0),
            'not_found': kb.get('not_found', 0) - ka.get('not_found', 0),
            'found_pct': round((kb.get('found_pct', 0) - ka.get('found_pct', 0)), 2),
            'not_found_pct': round((kb.get('not_found_pct', 0) - ka.get('not_found_pct', 0)), 2),
        },
        'service_filter': service or '',
        'generated_at': now_iso(),
    }
    return jsonify(response)


@app.route('/runs', methods=['POST'])
def create_run():
    ensure_db()
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    services = normalize_service_scope(data.get('services'))
    if not services:
        legacy_service = (data.get('service') or '').strip()
        services = normalize_service_scope([legacy_service] if legacy_service else [])
    service = services[0] if services else None
    period_id = data.get('period_id')
    try:
        period_id = int(period_id) if period_id not in (None, '') else None
    except Exception:
        period_id = None
    created_by = (data.get('created_by') or '').strip() or 'unknown'
    if not name:
        return jsonify({'error': 'Debe indicar nombre de jornada'}), 400
    if not services:
        return jsonify({'error': 'Debes seleccionar al menos un servicio para la jornada'}), 400
    if not period_id:
        return jsonify({'error': 'Debes seleccionar el periodo de inventario'}), 400
    period = InventoryPeriod.query.get(period_id)
    if not period:
        return jsonify({'error': 'Periodo no encontrado'}), 404
    if period.status != 'open':
        return jsonify({'error': 'El periodo seleccionado esta cerrado o anulado'}), 400

    run = InventoryRun(
        name=name,
        period_id=period.id,
        service=service,
        service_scope_json=json.dumps(services, ensure_ascii=False),
        status='active',
        started_at=now_iso(),
        created_by=created_by,
    )
    db.session.add(run)
    db.session.commit()
    row = run.to_dict()
    row['period_name'] = period.name
    row['period_status'] = period.status
    return jsonify({'run': row})


@app.route('/runs/<int:run_id>/summary', methods=['GET'])
def run_summary(run_id):
    ensure_db()
    run, err = get_run_or_404(run_id)
    if err:
        return err

    q = Asset.query
    q = apply_run_scope_filter(q, run)
    total = q.count()

    found = RunAssetStatus.query.filter_by(run_id=run.id, status='Encontrado').count()
    not_found = RunAssetStatus.query.filter_by(run_id=run.id, status='No encontrado').count()
    pending = max(total - found - not_found, 0)
    return jsonify({
        'run': run.to_dict(),
        'summary': {
            'total': total,
            'found': found,
            'not_found': not_found,
            'pending': pending,
        }
    })


@app.route('/runs/<int:run_id>/close', methods=['POST'])
def close_run(run_id):
    ensure_db()
    run, err = get_run_or_404(run_id)
    if err:
        return err
    if run.status == 'cancelled':
        return jsonify({'error': 'La jornada esta anulada'}), 400
    if run.status != 'active':
        return jsonify({'error': 'La jornada ya esta cerrada'}), 400

    data = request.get_json() or {}
    user = (data.get('user') or '').strip() or 'system_close'
    now_iso_value = now_iso()

    q = Asset.query
    q = apply_run_scope_filter(q, run)
    assets_scope = q.all()
    asset_ids = [a.id for a in assets_scope]

    existing_statuses = RunAssetStatus.query.filter(
        RunAssetStatus.run_id == run.id,
        RunAssetStatus.asset_id.in_(asset_ids)
    ).all() if asset_ids else []
    existing_map = {s.asset_id: s for s in existing_statuses}

    created_not_found = 0
    for asset in assets_scope:
        if asset.id in existing_map:
            continue
        db.session.add(RunAssetStatus(
            run_id=run.id,
            asset_id=asset.id,
            status='No encontrado',
            scanned_at=now_iso_value,
            scanned_by=user,
        ))
        asset.estado_inventario = 'No encontrado'
        asset.fecha_verificacion = now_iso_value
        asset.usuario_verificador = user
        created_not_found += 1

    run.status = 'closed'
    run.closed_at = now_iso_value
    db.session.commit()

    found = RunAssetStatus.query.filter_by(run_id=run.id, status='Encontrado').count()
    not_found = RunAssetStatus.query.filter_by(run_id=run.id, status='No encontrado').count()
    return jsonify({
        'run': run.to_dict(),
        'summary': {
            'total': len(assets_scope),
            'found': found,
            'not_found': not_found,
            'pending': 0,
        },
        'auto_marked_not_found': created_not_found,
    })


@app.route('/runs/<int:run_id>/cancel', methods=['POST'])
def cancel_run(run_id):
    ensure_db()
    run, err = get_run_or_404(run_id)
    if err:
        return err
    if run.status == 'cancelled':
        return jsonify({'error': 'La jornada ya esta anulada'}), 400

    data = request.get_json() or {}
    reason = (data.get('reason') or '').strip()
    user = (data.get('user') or '').strip() or 'usuario_movil'
    if not reason:
        return jsonify({'error': 'Debes indicar el motivo de anulacion de la jornada'}), 400

    has_scan_trace = RunAssetStatus.query.filter_by(run_id=run.id).count() > 0
    if has_scan_trace:
        return jsonify({'error': 'No puedes anular la jornada porque ya tiene trazabilidad de escaneo'}), 400

    run.status = 'cancelled'
    run.closed_at = now_iso()
    run.cancelled_at = now_iso()
    run.cancelled_by = user
    run.cancel_reason = reason
    db.session.commit()
    return jsonify({'run': run.to_dict()})


@app.route('/export', methods=['GET', 'POST'])
def export():
    ensure_db()
    payload = request.get_json(silent=True) if request.method == 'POST' else None
    source = payload or request.args
    service = (source.get('service') or '').strip()
    run_id = source.get('run_id')
    try:
        run_id = int(run_id) if run_id not in (None, '') else None
    except Exception:
        run_id = None
    period_id = source.get('period_id')
    try:
        period_id = int(period_id) if period_id not in (None, '') else None
    except Exception:
        period_id = None
    receiver = (source.get('receiver') or '').strip()
    observation = (source.get('observation') or '').strip()
    report_date = (source.get('report_date') or '').strip()
    warehouse_lead = (source.get('warehouse_lead') or '').strip()
    assets_manager = (source.get('assets_manager') or '').strip()
    per_asset_observations = source.get('per_asset_observations') or {}
    if not isinstance(per_asset_observations, dict):
        per_asset_observations = {}
    if not period_id:
        return jsonify({'error': 'Debes seleccionar el periodo para generar A22'}), 400
    if not run_id:
        return jsonify({'error': 'Debes seleccionar la jornada del periodo para generar A22'}), 400
    if not warehouse_lead:
        return jsonify({'error': 'Lider de almacen es obligatorio'}), 400
    if not assets_manager:
        return jsonify({'error': 'Responsable de activos fijos es obligatorio'}), 400
    if report_date:
        try:
            datetime.strptime(report_date, '%Y-%m-%d')
        except Exception:
            return jsonify({'error': 'Fecha invalida. Usa formato YYYY-MM-DD'}), 400
    selected_date = report_date or now_local_dt().strftime('%Y-%m-%d')

    run, assets_scope, err = get_a22_scope(service=service or None, run_id=run_id, period_id=period_id)
    if err:
        return err
    if not assets_scope:
        return jsonify({'error': 'No hay activos encontrados para generar A22 con ese filtro'}), 400
    if not run or not run.service:
        return jsonify({'error': 'La jornada seleccionada no tiene centro de costo asociado'}), 400

    if not os.path.exists(TEMPLATE_A22_PATH):
        return jsonify({'error': 'No existe la plantilla formato a22.xlsx'}), 400

    wb = load_workbook(TEMPLATE_A22_PATH)
    ws = wb[wb.sheetnames[0]]

    assets_scope = sort_assets_for_a22(assets_scope)
    selected_service = run.service
    selected_receiver = receiver or assets_scope[0].nom_resp or ''
    work_area = classify_area(selected_service).upper()

    data_rows = [r for r in range(13, ws.max_row + 1) if isinstance(ws.cell(r, 1).value, (int, float))]
    capacity = len(data_rows)
    if capacity == 0:
        return jsonify({'error': 'La plantilla no tiene filas de detalle configuradas'}), 400

    data_start = min(data_rows)
    data_end = max(data_rows)

    logo_path = next((p for p in A22_LOGO_CANDIDATES if os.path.exists(p)), None)
    logo_bytes = None
    if logo_path is None:
        try:
            with zipfile.ZipFile(TEMPLATE_A22_PATH, 'r') as zf:
                media_files = [n for n in zf.namelist() if n.startswith('xl/media/')]
                if media_files:
                    logo_bytes = zf.read(media_files[0])
        except Exception:
            logo_bytes = None

    def apply_header_and_signature(sheet):
        sheet.cell(6, 6).value = selected_date
        sheet.cell(7, 6).value = selected_service
        sheet.cell(8, 6).value = selected_receiver
        sheet.cell(9, 6).value = f'LIDER {selected_service}'
        sheet.cell(10, 6).value = work_area

        # Reemplaza nombre fijo en firmas por el responsable seleccionado.
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                if isinstance(cell.value, str) and 'LYDA MARTINEZ' in cell.value.upper():
                    cell.value = selected_receiver

        # Texto legal dinámico (evita depender de fórmulas rotas por estructura).
        legal_text = (
            f'Una vez culminado el proceso de inventario general en el área de {selected_service} '
            f'se entrega a {selected_receiver} responsable de dicha área, un documento detallado '
            f'que incluye todos los activos fijos asignados, clasificados y verificados. '
            f'A partir de la entrega formal de este documento, el responsable del área asume la '
            f'obligación de velar por la custodia y el buen estado de cada activo listado. '
            f'En caso de pérdida, daño o cualquier irregularidad que afecte los activos bajo su '
            f'supervisión, el responsable deberá presentar una justificación oportuna y detallada, '
            f'y asumir las consecuencias correspondientes. Esta medida busca asegurar la transparencia '
            f'y la adecuada gestión de los recursos de la institución.'
        )
        sheet.cell(288, 1).value = legal_text
        legal_align = copy(sheet.cell(288, 1).alignment) if sheet.cell(288, 1).alignment else Alignment()
        legal_align.wrap_text = True
        legal_align.vertical = 'top'
        sheet.cell(288, 1).alignment = legal_align

        # Firma del responsable del área dinámica.
        sheet.cell(295, 2).value = warehouse_lead
        sheet.cell(295, 6).value = selected_receiver
        sheet.cell(295, 8).value = assets_manager

        # Ajuste visual de líneas de firma para que no se vean pegadas.
        # Primero limpia cualquier merge que se cruce con H:L en filas de firma.
        target_rows = (294, 295, 296)
        target_col_start = 8   # H
        target_col_end = 12    # L
        for m in list(sheet.merged_cells.ranges):
            min_col, min_row, max_col, max_row = m.bounds
            intersects_rows = not (max_row < min(target_rows) or min_row > max(target_rows))
            intersects_cols = not (max_col < target_col_start or min_col > target_col_end)
            if intersects_rows and intersects_cols:
                sheet.unmerge_cells(str(m))

        # Amplía horizontalmente el bloque de "Responsable de activos fijos" a H:L.
        for r in target_rows:
            sheet.merge_cells(start_row=r, start_column=8, end_row=r, end_column=12)

        # Líneas: las dos primeras estándar y la tercera más larga por el bloque más ancho.
        sheet.cell(294, 2).value = '______________________________'
        sheet.cell(294, 6).value = '______________________________'
        sheet.cell(294, 8).value = '_______________________________________________'

        for r, c in [(294, 2), (294, 6), (294, 8)]:
            align = copy(sheet.cell(r, c).alignment) if sheet.cell(r, c).alignment else Alignment()
            align.horizontal = 'center'
            align.vertical = 'center'
            sheet.cell(r, c).alignment = align

        # Asegura texto exacto bajo la tercera firma.
        sheet.cell(296, 2).value = 'LIDER DE ALMACEN'
        sheet.cell(296, 6).value = 'RESPONSABLE DE AREA'
        sheet.cell(296, 8).value = 'RESPONSABLE DE ACTIVOS FIJOS'

        # Centrado de nombres/cargos de firma para mejor presentación.
        for r, c in [(295, 2), (295, 6), (295, 8), (296, 2), (296, 6), (296, 8)]:
            align = copy(sheet.cell(r, c).alignment) if sheet.cell(r, c).alignment else Alignment()
            align.horizontal = 'center'
            align.vertical = 'center'
            align.wrap_text = True
            sheet.cell(r, c).alignment = align

        if logo_path is not None:
            try:
                img = XLImage(logo_path)
                img.anchor = 'A2'
                fit_logo_to_a22_box(sheet, img, from_col=1, to_col=2, from_row=2, to_row=5)
                sheet.add_image(img)
            except Exception:
                pass
        elif logo_bytes is not None:
            try:
                image_stream = BytesIO(logo_bytes)
                pil_img = PILImage.open(image_stream)
                img = XLImage(pil_img)
                img.anchor = 'A2'
                fit_logo_to_a22_box(sheet, img, from_col=1, to_col=2, from_row=2, to_row=5)
                sheet.add_image(img)
            except Exception:
                pass

    chunks = [assets_scope[i:i + capacity] for i in range(0, len(assets_scope), capacity)]
    template_ws = ws
    template_ws.title = 'A22 1'

    for chunk_index, chunk_assets in enumerate(chunks, start=1):
        if chunk_index == 1:
            ws_chunk = template_ws
        else:
            ws_chunk = wb.copy_worksheet(template_ws)
            ws_chunk.title = f'A22 {chunk_index}'

        apply_header_and_signature(ws_chunk)

        for row_idx in data_rows:
            for col in [2, 3, 6, 7, 8, 9, 10]:
                ws_chunk.cell(row_idx, col).value = None

        for idx, asset in enumerate(chunk_assets):
            row_idx = data_rows[idx]
            ws_chunk.cell(row_idx, 2).value = asset.c_act
            ws_chunk.cell(row_idx, 3).value = asset.nom or ''
            ws_chunk.cell(row_idx, 6).value = asset.des_ubi or ''
            asset_obs = (per_asset_observations.get(str(asset.c_act)) or '').strip()
            ws_chunk.cell(row_idx, 7).value = asset_obs or observation or (asset.observacion_inventario or '')
            ws_chunk.cell(row_idx, 8).value = classify_asset_group(asset)
            ws_chunk.cell(row_idx, 9).value = reference_serial(asset)
            ws_chunk.cell(row_idx, 10).value = asset.modelo or ''

            # Ajuste visual para textos largos.
            text_candidates = [
                ws_chunk.cell(row_idx, 3).value or '',
                ws_chunk.cell(row_idx, 6).value or '',
                ws_chunk.cell(row_idx, 7).value or '',
                ws_chunk.cell(row_idx, 9).value or '',
                ws_chunk.cell(row_idx, 10).value or '',
            ]
            max_len = max(len(str(x)) for x in text_candidates)
            lines = max(1, min(8, (max_len // 20) + 1))
            ws_chunk.row_dimensions[row_idx].height = max(24, 14 * lines)
            for col in [3, 6, 7, 8, 9, 10]:
                c = ws_chunk.cell(row_idx, col)
                base_align = copy(c.alignment) if c.alignment else Alignment()
                base_align.wrap_text = True
                base_align.vertical = 'top'
                c.alignment = base_align

        # Mantiene formato y celdas combinadas: solo oculta filas sobrantes del bloque de activos.
        used = len(chunk_assets)
        for i, row_idx in enumerate(data_rows):
            ws_chunk.row_dimensions[row_idx].hidden = i >= used

    out = BytesIO()
    wb.save(out)
    base_name = run.name if run else selected_service
    safe_name = clean_filename(base_name)
    filename = f'a22_inventario_{safe_name}.xlsx'
    content = out.getvalue()
    date_parts = selected_date.split('-') if selected_date else []
    year_num = int(date_parts[0]) if len(date_parts) == 3 and date_parts[0].isdigit() else now_local_dt().year
    month_num = int(date_parts[1]) if len(date_parts) == 3 and date_parts[1].isdigit() else now_local_dt().month
    period_label = f'{selected_service} - {selected_date}'
    persist_generated_report_file(
        content=content,
        report_type='a22_excel',
        title='Acta A22 - Excel',
        period_label=period_label,
        period_id=period_id,
        file_name=filename,
        folder_group='a22',
        year=year_num,
        month=month_num,
    )
    return send_file(BytesIO(content), download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/export_a22_pdf', methods=['GET', 'POST'])
def export_a22_pdf():
    ensure_db()
    payload = request.get_json(silent=True) if request.method == 'POST' else None
    source = payload or request.args
    service = (source.get('service') or '').strip()
    run_id = source.get('run_id')
    try:
        run_id = int(run_id) if run_id not in (None, '') else None
    except Exception:
        run_id = None
    period_id = source.get('period_id')
    try:
        period_id = int(period_id) if period_id not in (None, '') else None
    except Exception:
        period_id = None
    receiver = (source.get('receiver') or '').strip()
    observation = (source.get('observation') or '').strip()
    report_date = (source.get('report_date') or '').strip()
    warehouse_lead = (source.get('warehouse_lead') or '').strip()
    assets_manager = (source.get('assets_manager') or '').strip()
    per_asset_observations = source.get('per_asset_observations') or {}
    if not isinstance(per_asset_observations, dict):
        per_asset_observations = {}
    if not period_id:
        return jsonify({'error': 'Debes seleccionar el periodo para generar A22'}), 400
    if not run_id:
        return jsonify({'error': 'Debes seleccionar la jornada del periodo para generar A22'}), 400
    if not warehouse_lead:
        return jsonify({'error': 'Lider de almacen es obligatorio'}), 400
    if not assets_manager:
        return jsonify({'error': 'Responsable de activos fijos es obligatorio'}), 400
    if report_date:
        try:
            datetime.strptime(report_date, '%Y-%m-%d')
        except Exception:
            return jsonify({'error': 'Fecha invalida. Usa formato YYYY-MM-DD'}), 400
    selected_date = report_date or now_local_dt().strftime('%Y-%m-%d')

    run, assets_scope, err = get_a22_scope(service=service or None, run_id=run_id, period_id=period_id)
    if err:
        return err
    if not assets_scope:
        return jsonify({'error': 'No hay activos encontrados para generar A22 con ese filtro'}), 400
    if not run or not run.service:
        return jsonify({'error': 'La jornada seleccionada no tiene centro de costo asociado'}), 400

    assets_scope = sort_assets_for_a22(assets_scope)
    selected_service = run.service
    selected_receiver = receiver or assets_scope[0].nom_resp or ''
    work_area = classify_area(selected_service).upper()
    logo_path = get_hospital_logo_path()
    cod_path = get_codificacion_path()

    out = BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=letter,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        topMargin=22 * mm,
        bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle('a22h', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=13, textColor=colors.HexColor('#0B4F6C'))
    normal = ParagraphStyle('a22n', parent=styles['Normal'], fontSize=8, leading=10)
    centered = ParagraphStyle('a22c', parent=normal, alignment=1)
    legal_style = ParagraphStyle('a22legal', parent=styles['Normal'], fontSize=8, leading=11, textColor=colors.HexColor('#1F2937'))

    story = []
    story.append(Paragraph('FORMATO A22 - INVENTARIO GENERAL DE ACTIVOS FIJOS', header_style))
    story.append(Spacer(1, 4))
    meta_data = [
        [Paragraph('<b>Fecha</b>', normal), Paragraph(selected_date, normal)],
        [Paragraph('<b>Centro de costo</b>', normal), Paragraph(selected_service, normal)],
        [Paragraph('<b>Responsable centro de costo</b>', normal), Paragraph(selected_receiver, normal)],
        [Paragraph('<b>Cargo</b>', normal), Paragraph(f'LIDER {selected_service}', normal)],
        [Paragraph('<b>Area de trabajo</b>', normal), Paragraph(work_area, normal)],
        [Paragraph('<b>Cantidad activos entregados</b>', normal), Paragraph(str(len(assets_scope)), normal)],
    ]
    meta_table = Table(meta_data, colWidths=[48 * mm, 130 * mm])
    meta_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#EAF4FA')),
        ('BOX', (0, 0), (-1, -1), 0.6, colors.HexColor('#BFD5E3')),
        ('INNERGRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#D7E5EE')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 6))

    data = [[
        Paragraph('<b>N°</b>', normal),
        Paragraph('<b>CODIGO ACTIVO FIJO</b>', normal),
        Paragraph('<b>DESCRIPCION ACTIVO FIJO</b>', normal),
        Paragraph('<b>UBICACION</b>', normal),
        Paragraph('<b>OBSERVACION</b>', normal),
        Paragraph('<b>TIPO ACTIVO</b>', normal),
        Paragraph('<b>REFERENCIA/SERIAL</b>', normal),
        Paragraph('<b>MODELO</b>', normal),
    ]]
    for i, asset in enumerate(assets_scope, start=1):
        asset_obs = (per_asset_observations.get(str(asset.c_act)) or '').strip()
        data.append([
            Paragraph(str(i), normal),
            Paragraph(str(asset.c_act or ''), normal),
            Paragraph(str(asset.nom or ''), normal),
            Paragraph(str(asset.des_ubi or ''), normal),
            Paragraph(asset_obs or observation or str(asset.observacion_inventario or ''), normal),
            Paragraph(classify_asset_group(asset), normal),
            Paragraph(reference_serial(asset), normal),
            Paragraph(str(asset.modelo or ''), normal),
        ])
    table = Table(data, colWidths=[10 * mm, 24 * mm, 48 * mm, 32 * mm, 28 * mm, 22 * mm, 30 * mm, 20 * mm], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0B4F6C')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BOX', (0, 0), (-1, -1), 0.6, colors.HexColor('#BFD5E3')),
        ('INNERGRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#D7E5EE')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7FBFD')]),
    ]))
    story.append(table)

    legal_text = (
        f'Una vez culminado el proceso de inventario general en el area de {selected_service} '
        f'se entrega a {selected_receiver} responsable de dicha area, un documento detallado '
        f'que incluye todos los activos fijos asignados, clasificados y verificados. '
        f'A partir de la entrega formal de este documento, el responsable del area asume la '
        f'obligacion de velar por la custodia y el buen estado de cada activo listado. '
        f'En caso de perdida, dano o cualquier irregularidad que afecte los activos bajo su '
        f'supervision, el responsable debera presentar una justificacion oportuna y detallada, '
        f'y asumir las consecuencias correspondientes. Esta medida busca asegurar la transparencia '
        f'y la adecuada gestion de los recursos de la institucion.'
    )
    story.append(Spacer(1, 8))
    story.append(Paragraph(legal_text, legal_style))
    story.append(Spacer(1, 28))

    sign_table = Table([
        [
            Paragraph('______________________________', centered),
            Paragraph('______________________________', centered),
            Paragraph('______________________________', centered),
        ],
        [
            Paragraph(warehouse_lead, centered),
            Paragraph(selected_receiver, centered),
            Paragraph(assets_manager, centered),
        ],
        [
            Paragraph('LIDER DE ALMACEN', centered),
            Paragraph('RESPONSABLE DE AREA', centered),
            Paragraph('RESPONSABLE DE ACTIVOS FIJOS', centered),
        ],
    ], colWidths=[60 * mm, 60 * mm, 60 * mm])
    sign_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    story.append(sign_table)
    page_header = make_pdf_page_header(logo_path, right_image_path=cod_path)
    doc.build(story, onFirstPage=page_header, onLaterPages=page_header)
    content = out.getvalue()
    safe_name = str(selected_service).replace(' ', '_').replace('/', '_')
    filename = clean_filename(f'a22_inventario_{safe_name}.pdf')
    date_parts = selected_date.split('-') if selected_date else []
    year_num = int(date_parts[0]) if len(date_parts) == 3 and date_parts[0].isdigit() else now_local_dt().year
    month_num = int(date_parts[1]) if len(date_parts) == 3 and date_parts[1].isdigit() else now_local_dt().month
    period_label = f'{selected_service} - {selected_date}'
    persist_generated_report_file(
        content=content,
        report_type='a22_pdf',
        title='Acta A22 - PDF',
        period_label=period_label,
        period_id=period_id,
        file_name=filename,
        folder_group='a22',
        year=year_num,
        month=month_num,
    )
    return send_file(BytesIO(content), download_name=filename, as_attachment=True, mimetype='application/pdf')


@app.route('/reconciliation/export_found', methods=['GET'])
def reconciliation_export_found():
    ensure_db()
    service = (request.args.get('service') or '').strip()
    run_id = request.args.get('run_id', type=int)
    period_id = request.args.get('period_id', type=int)
    if not run_id and not period_id:
        return jsonify({'error': 'Debes seleccionar periodo o jornada para exportar'}), 400
    rows, err = build_reconciliation_rows(service=service, run_id=run_id, period_id=period_id)
    if err:
        return err
    found_rows = [r for r in rows if r['ESTADO_INVENTARIO'] == 'Encontrado']
    if not found_rows:
        return jsonify({'error': 'No hay activos encontrados para exportar'}), 400

    grouped = {}
    for row in found_rows:
        svc = str(row.get('SERVICIO') or '').strip() or 'SIN SERVICIO'
        grouped.setdefault(svc, []).append(row)

    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()
    for svc in sorted(grouped.keys(), key=lambda x: x.casefold()):
        sheet_name = excel_safe_sheet_name(svc, used_names)
        ws = wb.create_sheet(title=sheet_name)
        svc_rows = sorted(grouped[svc], key=lambda r: str(r.get('C_ACT') or ''))
        title = f'Base depurada - Encontrados ({svc})'
        write_reconciliation_sheet(ws, title, svc_rows)
        add_logo_to_excel_sheet(ws, logo_path=get_hospital_logo_path())

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    name = clean_filename(f"base_depurada_encontrados_{(service or 'todos').replace(' ', '_')}.xlsx")
    return send_file(out, as_attachment=True, download_name=name, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/reconciliation/export_not_found', methods=['GET'])
def reconciliation_export_not_found():
    ensure_db()
    service = (request.args.get('service') or '').strip()
    run_id = request.args.get('run_id', type=int)
    period_id = request.args.get('period_id', type=int)
    if not run_id and not period_id:
        return jsonify({'error': 'Debes seleccionar periodo o jornada para exportar'}), 400
    rows, err = build_reconciliation_rows(service=service, run_id=run_id, period_id=period_id)
    if err:
        return err
    not_found_rows = [r for r in rows if r['ESTADO_INVENTARIO'] == 'No encontrado']

    grouped = {}
    for row in not_found_rows:
        svc = str(row.get('SERVICIO') or '').strip() or 'SIN SERVICIO'
        grouped.setdefault(svc, []).append(row)

    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()
    if grouped:
        for svc in sorted(grouped.keys(), key=lambda x: x.casefold()):
            sheet_name = excel_safe_sheet_name(svc, used_names)
            ws = wb.create_sheet(title=sheet_name)
            svc_rows = sorted(grouped[svc], key=lambda r: str(r.get('C_ACT') or ''))
            title = f'Listado no encontrados ({svc})'
            write_reconciliation_sheet(ws, title, svc_rows)
            add_logo_to_excel_sheet(ws, logo_path=get_hospital_logo_path())
    else:
        ws = wb.create_sheet(title=excel_safe_sheet_name('SIN_NO_ENCONTRADOS', used_names))
        write_reconciliation_sheet(ws, 'Listado no encontrados (sin registros)', [])
        add_logo_to_excel_sheet(ws, logo_path=get_hospital_logo_path())

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    name = clean_filename(f"listado_no_encontrados_{(service or 'todos').replace(' ', '_')}.xlsx")
    return send_file(out, as_attachment=True, download_name=name, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/reconciliation/export_consolidated', methods=['GET'])
def reconciliation_export_consolidated():
    ensure_db()
    service = (request.args.get('service') or '').strip()
    run_id = request.args.get('run_id', type=int)
    period_id = request.args.get('period_id', type=int)
    if not run_id and not period_id:
        return jsonify({'error': 'Debes seleccionar periodo o jornada para exportar'}), 400
    rows, err = build_reconciliation_rows(service=service, run_id=run_id, period_id=period_id)
    if err:
        return err
    if not rows:
        return jsonify({'error': 'No hay activos para exportar'}), 400

    found_rows = [r for r in rows if r['ESTADO_INVENTARIO'] == 'Encontrado']
    not_found_rows = [r for r in rows if r['ESTADO_INVENTARIO'] == 'No encontrado']
    pending_rows = [r for r in rows if r['ESTADO_INVENTARIO'] == 'Pendiente']

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = 'Resumen'
    ws_summary.append(['CONSOLIDADO FINAL DE INVENTARIO'])
    ws_summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws_summary['A1'].font = Font(bold=True, size=14, color='0B4F6C')
    ws_summary.append([f"Servicio: {service or 'TODOS'}", f"Fecha: {now_local_dt().strftime('%Y-%m-%d %H:%M')}"])
    ws_summary.append(['Estado', 'Cantidad', 'Total costo', 'Total saldo'])
    ws_summary.append(['Encontrado', len(found_rows), sum(x['COSTO'] for x in found_rows), sum(x['SALDO'] for x in found_rows)])
    ws_summary.append(['No encontrado', len(not_found_rows), sum(x['COSTO'] for x in not_found_rows), sum(x['SALDO'] for x in not_found_rows)])
    ws_summary.append(['Pendiente', len(pending_rows), sum(x['COSTO'] for x in pending_rows), sum(x['SALDO'] for x in pending_rows)])
    ws_summary.append(['TOTAL', len(rows), sum(x['COSTO'] for x in rows), sum(x['SALDO'] for x in rows)])
    for col in ['A', 'B', 'C', 'D']:
        ws_summary.column_dimensions[col].width = [22, 12, 18, 18][ord(col) - ord('A')]
    for r in range(4, ws_summary.max_row + 1):
        ws_summary.cell(r, 3).number_format = '"$"#,##0'
        ws_summary.cell(r, 4).number_format = '"$"#,##0'
    add_logo_to_excel_sheet(ws_summary, logo_path=get_hospital_logo_path())

    ws_found = wb.create_sheet('Encontrados')
    write_reconciliation_sheet(ws_found, 'Activos encontrados', found_rows)
    add_logo_to_excel_sheet(ws_found, logo_path=get_hospital_logo_path())

    ws_not = wb.create_sheet('No encontrados')
    write_reconciliation_sheet(ws_not, 'Activos no encontrados', not_found_rows)
    add_logo_to_excel_sheet(ws_not, logo_path=get_hospital_logo_path())

    ws_pending = wb.create_sheet('Pendientes')
    write_reconciliation_sheet(ws_pending, 'Activos pendientes', pending_rows)
    add_logo_to_excel_sheet(ws_pending, logo_path=get_hospital_logo_path())

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    name = clean_filename(f"consolidado_final_inventario_{(service or 'todos').replace(' ', '_')}.xlsx")
    return send_file(out, as_attachment=True, download_name=name, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def normalize_family_code(value):
    text_value = str(value or '').strip()
    if text_value.endswith('.0'):
        text_value = text_value[:-2]
    return text_value


@lru_cache(maxsize=1)
def load_family_catalog_names():
    names = {}
    if not os.path.exists(FAMILY_CATALOG_PATH):
        return names
    try:
        df = pd.read_excel(FAMILY_CATALOG_PATH, dtype=str)
        cols = normalize_columns(df.columns)
        code_col = cols.get('C_FAM') or cols.get('CODIGO') or cols.get('COD_FAM') or cols.get('FAMILIA')
        name_col = cols.get('NOM_FAM') or cols.get('NOMBRE') or cols.get('DESCRIPCION') or cols.get('NOM')
        if not code_col or not name_col:
            return names
        for _, row in df.iterrows():
            code = normalize_family_code(row.get(code_col))
            name = str(row.get(name_col) or '').strip()
            if code and name and code.lower() != 'nan':
                names[code] = name
    except Exception:
        return names
    return names


def get_accounting_template_path():
    for path in ACCOUNTING_TEMPLATE_CANDIDATES:
        if os.path.exists(path):
            return path
    return ACCOUNTING_TEMPLATE_CANDIDATES[0]


def normalize_month_year(month_raw, year_raw):
    now = now_local_dt()
    try:
        month = int(str(month_raw or '').strip())
    except Exception:
        month = now.month
    try:
        year = int(str(year_raw or '').strip())
    except Exception:
        year = now.year
    month = min(12, max(1, month))
    year = min(2100, max(2000, year))
    return month, year


def sanitize_filename(text):
    raw = unicodedata.normalize('NFD', str(text or ''))
    raw = ''.join(ch for ch in raw if unicodedata.category(ch) != 'Mn')
    allowed = 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-_ .'
    clean = ''.join(ch if ch in allowed else '_' for ch in raw).strip()
    return clean or 'reporte'


def persist_accounting_report_file(content, file_name, period_label, month, year, report_title, period_id=None):
    reports_folder = os.path.join(REPORTS_DIR, 'accounting_monthly', str(year), f'{month:02d}')
    os.makedirs(reports_folder, exist_ok=True)

    stamped_name = f"{os.path.splitext(file_name)[0]}_{now_local_dt().strftime('%Y%m%d%H%M%S')}.xlsx"
    file_path = os.path.join(reports_folder, sanitize_filename(stamped_name))
    with open(file_path, 'wb') as f:
        f.write(content)

    report_row = GeneratedReport(
        report_type='accounting_monthly',
        title=report_title or 'Informe de conciliacion activos fijos - contabilidad',
        period_id=period_id,
        period_label=period_label,
        file_name=os.path.basename(file_path),
        file_path=file_path,
        generated_at=now_iso(),
    )
    db.session.add(report_row)
    db.session.commit()


def persist_generated_report_file(content, report_type, title, period_label, file_name, folder_group, year=None, month=None, period_id=None):
    yy = str(year or now_local_dt().year)
    mm = f"{int(month):02d}" if month else '00'
    reports_folder = os.path.join(REPORTS_DIR, folder_group, yy, mm)
    os.makedirs(reports_folder, exist_ok=True)

    stamped_name = f"{os.path.splitext(file_name)[0]}_{now_local_dt().strftime('%Y%m%d%H%M%S')}{os.path.splitext(file_name)[1] or ''}"
    stamped_name = sanitize_filename(stamped_name)
    file_path = os.path.join(reports_folder, stamped_name)
    with open(file_path, 'wb') as f:
        f.write(content)

    row = GeneratedReport(
        report_type=report_type,
        title=title,
        period_id=period_id,
        period_label=period_label,
        file_name=stamped_name,
        file_path=file_path,
        generated_at=now_iso(),
    )
    db.session.add(row)
    db.session.commit()


def asset_raw_payload(asset):
    if asset.raw_row_json:
        try:
            payload = json.loads(asset.raw_row_json)
            if isinstance(payload, dict):
                return payload
        except Exception:
            pass
    return {
        'C_ACT': asset.c_act,
        'NOM': asset.nom or '',
        'C_FAM': asset.c_fam or '',
        'NOM_FAM': asset.nom_fam or '',
        'MODELO': asset.modelo or '',
        'REF': asset.ref or '',
        'SERIE': asset.serie or '',
        'NOM_MARCA': asset.nom_marca or '',
        'C_TIAC': asset.c_tiac or '',
        'DESC_TIAC': asset.desc_tiac or '',
        'DES_SUBTIAC': asset.desc_subtiac or '',
        'DEPRECIA': asset.deprecia or '',
        'VIDA_UTIL': asset.vida_util or '',
        'DES_UBI': asset.des_ubi or '',
        'NOM_CCOS': asset.nom_ccos or '',
        'NOM_RESP': asset.nom_resp or '',
        'EST': asset.est or '',
        'COSTO': to_number(asset.costo),
        'SALDO': to_number(asset.saldo),
        'FECHA_COMPRA': asset.fecha_compra or '',
    }


def write_headers_row(ws, row_idx, columns):
    for col_idx, col_name in enumerate(columns, start=1):
        c = ws.cell(row_idx, col_idx, col_name)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill(fill_type='solid', fgColor='0B4F6C')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = Border(
            left=Side(style='thin', color='BFD5E3'),
            right=Side(style='thin', color='BFD5E3'),
            top=Side(style='thin', color='BFD5E3'),
            bottom=Side(style='thin', color='BFD5E3'),
        )


@app.route('/reports/accounting_monthly_excel', methods=['GET'])
def report_accounting_monthly_excel():
    ensure_db()
    force_refresh = str(request.args.get('refresh', '')).strip().lower() in {'1', 'true', 'yes', 'si'}
    report_title = str(request.args.get('report_title') or '').strip()
    if not report_title:
        return jsonify({'error': 'Debes indicar el titulo del informe contable'}), 400
    generated_by = str(request.args.get('generated_by') or '').strip()
    if not generated_by:
        return jsonify({'error': 'Debes indicar el usuario que genera el informe'}), 400
    month, year = normalize_month_year(request.args.get('month'), request.args.get('year'))
    month_label = MONTH_LABELS_ES.get(month, str(month))
    period_label = f'{month_label} {year}'
    template_path = get_accounting_template_path()
    if not os.path.exists(template_path):
        return jsonify({'error': 'No se encontro la plantilla "INFORME CONTABILIDAD REF.xlsx"'}), 400
    template_mtime = int(os.path.getmtime(template_path))
    current_cache_key = f"{get_assets_revision()}:{ACCOUNTING_REPORT_ALGO_VERSION}:{template_mtime}:{month}:{year}:{report_title}:{generated_by}"

    with ACCOUNTING_CACHE_LOCK:
        cached_version = ACCOUNTING_REPORT_CACHE.get('version')
        cached_bytes = ACCOUNTING_REPORT_CACHE.get('bytes')
        cached_filename = ACCOUNTING_REPORT_CACHE.get('filename')

    if (not force_refresh) and cached_version == current_cache_key and cached_bytes:
        safe_period = sanitize_filename(period_label.replace(' ', '_'))
        base_filename = f'informe_conciliacion_activos_fijos_contabilidad_{safe_period}.xlsx'
        persist_accounting_report_file(cached_bytes, base_filename, period_label, month, year, report_title, period_id=None)
        return send_file(
            BytesIO(cached_bytes),
            as_attachment=True,
            download_name=base_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    asset_rows = db.session.query(
        Asset.raw_row_json,
        Asset.c_act,
        Asset.c_fam,
        Asset.nom_fam,
        Asset.costo,
        Asset.saldo,
    ).all()
    if not asset_rows:
        return jsonify({'error': 'No hay activos cargados para generar el informe'}), 400

    def row_values(ws, row_idx):
        return [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]

    def non_empty_headers(ws, row_idx):
        values = row_values(ws, row_idx)
        return [str(v).strip() for v in values if str(v or '').strip()]

    def clear_sheet_data(ws, from_row):
        if ws.max_row >= from_row:
            ws.delete_rows(from_row, ws.max_row - from_row + 1)

    family_name_by_code = {}
    rows_by_code = {}
    base_rows = []
    all_codes_seen = set()

    wb = load_workbook(template_path)
    if 'BASE COMPLETA' not in wb.sheetnames or 'DESGLOSE' not in wb.sheetnames or 'INFORME' not in wb.sheetnames:
        return jsonify({'error': 'La plantilla no tiene las hojas requeridas: BASE COMPLETA, DESGLOSE, INFORME'}), 400

    ws_base = wb['BASE COMPLETA']
    ws_des = wb['DESGLOSE']
    ws_inf = wb['INFORME']
    ws_inf.column_dimensions['C'].width = max(18, ws_inf.column_dimensions['C'].width or 0)
    ws_inf.row_dimensions[3].height = max(55, ws_inf.row_dimensions[3].height or 0)
    logo_path = get_hospital_logo_path()
    if logo_path:
        try:
            logo_img = XLImage(logo_path)
            # Logo anclado y contenido en C3.
            fit_logo_to_a22_box(ws_inf, logo_img, from_col=3, to_col=3, from_row=3, to_row=3, padding_px=2, shrink=0.98)
            ws_inf.add_image(logo_img)
        except Exception:
            pass
    ws_inf.cell(3, 3, report_title)
    ws_inf.cell(49, 4, generated_by)
    d48_cell = ws_inf.cell(48, 4)
    d48_align = copy(d48_cell.alignment) if d48_cell.alignment else Alignment()
    d48_align.horizontal = 'center'
    d48_align.vertical = 'center'
    d48_align.wrap_text = True
    d48_cell.alignment = d48_align
    by_cell = ws_inf.cell(49, 4)
    by_align = copy(by_cell.alignment) if by_cell.alignment else Alignment()
    by_align.wrap_text = True
    by_align.shrink_to_fit = True
    by_align.horizontal = 'center'
    by_align.vertical = 'center'
    by_cell.alignment = by_align
    if len(generated_by) > 34:
        ws_inf.row_dimensions[49].height = max(28, ws_inf.row_dimensions[49].height or 0)

    template_headers = non_empty_headers(ws_base, 1)
    columns_order = template_headers[:] if template_headers else []
    seen_cols = set(columns_order)

    for raw_row_json, c_act, c_fam, nom_fam, costo, saldo in asset_rows:
        payload = {}
        if raw_row_json:
            try:
                maybe_payload = json.loads(raw_row_json)
                if isinstance(maybe_payload, dict):
                    payload = maybe_payload
            except Exception:
                payload = {}

        fam_code = normalize_family_code(payload.get('C_FAM') or c_fam)
        payload['C_FAM'] = fam_code
        payload['NOM_FAM'] = payload.get('NOM_FAM') or nom_fam or ''
        payload['COSTO'] = to_number(payload.get('COSTO') if payload.get('COSTO') is not None else costo)
        payload['SALDO'] = to_number(payload.get('SALDO') if payload.get('SALDO') is not None else saldo)
        payload['C_ACT'] = payload.get('C_ACT') or c_act

        for col in payload.keys():
            if col not in seen_cols:
                seen_cols.add(col)
                columns_order.append(col)

        base_rows.append(payload)
        all_codes_seen.add(fam_code)
        if fam_code and payload.get('NOM_FAM'):
            family_name_by_code[fam_code] = str(payload.get('NOM_FAM') or '')
        rows_by_code.setdefault(fam_code, []).append(payload)

    if 'COSTO' not in columns_order:
        columns_order.append('COSTO')

    report_rows_by_code = {}
    for fam_code, fam_rows in rows_by_code.items():
        code = normalize_family_code(fam_code)
        if not code or code in ACCOUNTING_EXCLUDED_FAMILIES:
            continue
        report_rows_by_code[code] = list(fam_rows)
    report_scope_assets_count = sum(len(rows) for rows in report_rows_by_code.values())
    excluded_in_scope = [code for code in report_rows_by_code.keys() if code in ACCOUNTING_EXCLUDED_FAMILIES]
    if STRICT_ACCOUNTING_VALIDATION and excluded_in_scope:
        return jsonify({'error': f'Validacion interna fallo: familias excluidas en alcance reportable ({", ".join(excluded_in_scope)})'}), 500

    catalog_names = load_family_catalog_names()
    configured_parent_codes = [code for code in ACCOUNTING_FAMILY_ORDER if len(code) == 4]
    parent_codes = []
    parent_codes_seen = set()

    for parent in configured_parent_codes:
        has_rows = any(code.startswith(parent) for code in report_rows_by_code.keys())
        if has_rows and parent not in parent_codes_seen:
            parent_codes.append(parent)
            parent_codes_seen.add(parent)

    dynamic_parent_codes = sorted({
        code[:4] for code in report_rows_by_code.keys()
        if len(code) >= 4 and code[:4].isdigit() and code[:4] not in parent_codes_seen
    })
    for parent in dynamic_parent_codes:
        parent_codes.append(parent)
        parent_codes_seen.add(parent)

    clear_sheet_data(ws_base, 2)
    for col_idx, header in enumerate(columns_order, start=1):
        ws_base.cell(1, col_idx, header)

    base_rows_sorted = sorted(
        base_rows,
        key=lambda r: (normalize_family_code(r.get('C_FAM')), str(r.get('C_ACT') or ''))
    )
    row_idx = 2
    for row in base_rows_sorted:
        for col_idx, col_name in enumerate(columns_order, start=1):
            value = row.get(col_name)
            cell = ws_base.cell(row_idx, col_idx, value if value is not None else '')
            if col_name == 'COSTO':
                cell.number_format = '"$"#,##0.00'
        row_idx += 1

    clear_sheet_data(ws_des, 2)
    for col_idx, header in enumerate(columns_order, start=1):
        ws_des.cell(1, col_idx, header)

    des_row = 3
    cost_col = columns_order.index('COSTO') + 1
    cost_col_letter = get_column_letter(cost_col)
    des_total_report_scope = 0.0
    des_detail_rows_written = 0
    assigned_codes = set()
    nc_total_row_des = None
    des_formula_cells = []
    family_total_refs = {}
    configured_children_by_parent = {
        parent: [code for code in ACCOUNTING_FAMILY_ORDER if len(code) > 4 and code.startswith(parent)]
        for parent in configured_parent_codes
    }

    for parent in parent_codes:
        family_codes_order = []
        if parent in report_rows_by_code:
            family_codes_order.append(parent)

        configured_children = configured_children_by_parent.get(parent, [])
        for child in configured_children:
            if child in report_rows_by_code and child not in family_codes_order:
                family_codes_order.append(child)

        dynamic_children = sorted(
            code for code in report_rows_by_code.keys()
            if len(code) > 4 and code.startswith(parent) and code not in family_codes_order
        )
        family_codes_order.extend(dynamic_children)
        assigned_codes.update(family_codes_order)

        rows = []
        for fam_code in family_codes_order:
            rows.extend(report_rows_by_code.get(fam_code, []))
        rows = sorted(rows, key=lambda r: (str(r.get('C_FAM') or ''), str(r.get('C_ACT') or '')))
        parent_name = (
            family_name_by_code.get(parent)
            or catalog_names.get(parent)
            or f'FAMILIA {parent}'
        )

        ws_des.cell(des_row, 2, parent)
        ws_des.cell(des_row, 3, parent_name)
        des_row += 2

        for col_idx, header in enumerate(columns_order, start=1):
            ws_des.cell(des_row, col_idx, header)
        des_row += 1

        subtotal = 0.0
        parent_total_refs = []
        for fam_code in family_codes_order:
            fam_rows = sorted(
                report_rows_by_code.get(fam_code, []),
                key=lambda r: str(r.get('C_ACT') or '')
            )
            if not fam_rows:
                continue

            fam_subtotal = 0.0
            detail_start_row = des_row
            for row in fam_rows:
                for col_idx, col_name in enumerate(columns_order, start=1):
                    value = row.get(col_name)
                    cell = ws_des.cell(des_row, col_idx, value if value is not None else '')
                    if col_name == 'COSTO':
                        cell.number_format = '"$"#,##0.00'
                fam_subtotal += to_number(row.get('COSTO'))
                des_detail_rows_written += 1
                des_row += 1

            detail_end_row = des_row - 1
            fam_name = (
                family_name_by_code.get(fam_code)
                or catalog_names.get(fam_code)
                or f'FAMILIA {fam_code}'
            )
            ws_des.cell(des_row, 3, f'TOTAL {fam_code} - {fam_name}').font = Font(bold=True)
            if detail_end_row >= detail_start_row:
                family_total_cell = ws_des.cell(
                    des_row,
                    cost_col,
                    f'=ROUND(SUM({cost_col_letter}{detail_start_row}:{cost_col_letter}{detail_end_row}),2)'
                )
                des_formula_cells.append(family_total_cell)
            else:
                family_total_cell = ws_des.cell(des_row, cost_col, 0)
            family_total_cell.font = Font(bold=True)
            family_total_cell.number_format = '"$"#,##0.00'
            family_total_refs[fam_code] = f'DESGLOSE!${cost_col_letter}${des_row}'
            parent_total_refs.append(f'${cost_col_letter}${des_row}')
            des_row += 1
            subtotal += fam_subtotal

        ws_des.cell(des_row, 3, f'TOTAL {parent} - {parent_name}').font = Font(bold=True)
        if parent_total_refs:
            total_cell = ws_des.cell(
                des_row,
                cost_col,
                f'=ROUND({"+".join(parent_total_refs)},2)'
            )
            des_formula_cells.append(total_cell)
        else:
            total_cell = ws_des.cell(des_row, cost_col, 0)
        total_cell.font = Font(bold=True)
        total_cell.number_format = '"$"#,##0.00'
        des_total_report_scope += subtotal
        des_row += 2

    unassigned_codes = sorted(code for code in report_rows_by_code.keys() if code not in assigned_codes)
    if unassigned_codes:
        ws_des.cell(des_row, 2, 'NC')
        ws_des.cell(des_row, 3, 'NO CLASIFICADAS / FUERA DE ESTRUCTURA')
        des_row += 2
        for col_idx, header in enumerate(columns_order, start=1):
            ws_des.cell(des_row, col_idx, header)
        des_row += 1

        nc_subtotal = 0.0
        nc_total_refs = []
        for fam_code in unassigned_codes:
            fam_rows = sorted(report_rows_by_code.get(fam_code, []), key=lambda r: str(r.get('C_ACT') or ''))
            if not fam_rows:
                continue

            fam_subtotal = 0.0
            detail_start_row = des_row
            for row in fam_rows:
                for col_idx, col_name in enumerate(columns_order, start=1):
                    value = row.get(col_name)
                    cell = ws_des.cell(des_row, col_idx, value if value is not None else '')
                    if col_name == 'COSTO':
                        cell.number_format = '"$"#,##0.00'
                fam_subtotal += to_number(row.get('COSTO'))
                des_detail_rows_written += 1
                des_row += 1

            detail_end_row = des_row - 1
            fam_name = (
                family_name_by_code.get(fam_code)
                or catalog_names.get(fam_code)
                or f'FAMILIA {fam_code}'
            )
            ws_des.cell(des_row, 3, f'TOTAL {fam_code} - {fam_name}').font = Font(bold=True)
            if detail_end_row >= detail_start_row:
                family_total_cell = ws_des.cell(
                    des_row,
                    cost_col,
                    f'=ROUND(SUM({cost_col_letter}{detail_start_row}:{cost_col_letter}{detail_end_row}),2)'
                )
                des_formula_cells.append(family_total_cell)
            else:
                family_total_cell = ws_des.cell(des_row, cost_col, 0)
            family_total_cell.font = Font(bold=True)
            family_total_cell.number_format = '"$"#,##0.00'
            family_total_refs[fam_code] = f'DESGLOSE!${cost_col_letter}${des_row}'
            nc_total_refs.append(f'${cost_col_letter}${des_row}')
            des_row += 1
            nc_subtotal += fam_subtotal

        ws_des.cell(des_row, 3, 'TOTAL NC - NO CLASIFICADAS / FUERA DE ESTRUCTURA').font = Font(bold=True)
        if nc_total_refs:
            total_cell = ws_des.cell(
                des_row,
                cost_col,
                f'=ROUND({"+".join(nc_total_refs)},2)'
            )
            des_formula_cells.append(total_cell)
        else:
            total_cell = ws_des.cell(des_row, cost_col, 0)
        total_cell.font = Font(bold=True)
        total_cell.number_format = '"$"#,##0.00'
        des_total_report_scope += nc_subtotal
        nc_total_row_des = des_row
        des_row += 2

    expected_scope_codes = set(report_rows_by_code.keys())
    covered_scope_codes = assigned_codes.union(set(unassigned_codes))
    if STRICT_ACCOUNTING_VALIDATION and covered_scope_codes != expected_scope_codes:
        missing_codes = sorted(expected_scope_codes - covered_scope_codes)
        return jsonify({'error': f'Validacion interna fallo: familias sin cobertura en desglose ({", ".join(missing_codes)})'}), 500
    if STRICT_ACCOUNTING_VALIDATION and des_detail_rows_written != report_scope_assets_count:
        return jsonify({
            'error': (
                'Validacion interna fallo: filas de detalle en DESGLOSE '
                f'({des_detail_rows_written}) no coinciden con activos reportables ({report_scope_assets_count})'
            )
        }), 500
    expected_scope_total = round(sum(
        to_number(row.get('COSTO'))
        for rows in report_rows_by_code.values()
        for row in rows
    ), 2)
    if STRICT_ACCOUNTING_VALIDATION and abs(round(des_total_report_scope, 2) - expected_scope_total) > 0.01:
        return jsonify({
            'error': (
                'Validacion interna fallo: total DESGLOSE no coincide con total reportable '
                f'({round(des_total_report_scope, 2)} vs {expected_scope_total})'
            )
        }), 500

    # INFORME: escribir exactamente sobre el bloque de la plantilla (filas 6..31)
    for r in range(6, 32):
        for c in range(3, 8):
            ws_inf.cell(r, c, None)
    # Quitar columnas de conciliacion contable del formato generado por sistema
    ws_inf.cell(5, 6, None)
    ws_inf.cell(5, 7, None)

    info_row = 6
    parent_rows_written = []
    inf_formula_cells = []

    def refs_for_prefix(prefix):
        return [
            ref
            for fam_code, ref in family_total_refs.items()
            if normalize_family_code(fam_code).startswith(prefix)
        ]

    for group in ACCOUNTING_REPORT_STRUCTURE:
        child_specs = []
        for child in group['children']:
            prefixes = child.get('source_prefixes') or [child.get('source_prefix')]
            formula_refs = []
            for p in prefixes:
                if not p:
                    continue
                formula_refs.extend(refs_for_prefix(p))
            child_specs.append((child, formula_refs))

        parent_row = info_row
        first_child_row = parent_row + 1
        last_child_row = parent_row + len(child_specs)
        ws_inf.cell(info_row, 3, group['parent_code'])
        ws_inf.cell(info_row, 4, group['parent_name'])
        parent_value_cell = ws_inf.cell(info_row, 5, f'=SUM(E{first_child_row}:E{last_child_row})')
        parent_value_cell.number_format = '"$"#,##0.00'
        ws_inf.cell(info_row, 3).font = Font(bold=True)
        ws_inf.cell(info_row, 4).font = Font(bold=True)
        parent_value_cell.font = Font(bold=True)
        inf_formula_cells.append(parent_value_cell)
        parent_rows_written.append(parent_row)
        info_row += 1

        for child, formula_refs in child_specs:
            fallback_prefix = (child.get('source_prefixes') or [child.get('source_prefix')] or [''])[0]
            child_name = str(child.get('name') or '').strip() or catalog_names.get(fallback_prefix) or f"SUBFAMILIA {fallback_prefix}"
            ws_inf.cell(info_row, 3, child['report_code'])
            ws_inf.cell(info_row, 4, child_name)
            child_formula = '+'.join(formula_refs) if formula_refs else '0'
            child_value_cell = ws_inf.cell(info_row, 5, f'=ROUND({child_formula},2)')
            child_value_cell.number_format = '"$"#,##0.00'
            inf_formula_cells.append(child_value_cell)
            info_row += 1

    # Primer subtotal (solo familias inventariadas)
    for c in range(3, 8):
        ws_inf.cell(31, c, None)
    ws_inf.cell(31, 4, 'SUBTOTAL').font = Font(bold=True)
    subtotal_refs = [f'E{r}' for r in parent_rows_written]
    if subtotal_refs:
        subtotal_formula = '+'.join(subtotal_refs)
        subtotal_cell = ws_inf.cell(31, 5, f'={subtotal_formula}')
        inf_formula_cells.append(subtotal_cell)
    else:
        subtotal_cell = ws_inf.cell(31, 5, 0)
    subtotal_cell.font = Font(bold=True)
    subtotal_cell.number_format = '"$"#,##0.00'
    ws_inf.cell(31, 6, None)
    ws_inf.cell(31, 7, None)

    # Mantener estructura visual pero sin valores numericos en seccion contable fija
    for r in range(34, 44):
        for c in range(5, 8):
            ws_inf.cell(r, c, None)
    for r in [34, 35, 36, 38, 41]:
        ws_inf.cell(r, 5, '')

    # Limpieza defensiva: borrar valores/formulas de filas no requeridas,
    # aunque la plantilla cambie de posicion en futuras versiones.
    blocked_labels = {
        'TERRENO',
        'MUEBLES EN BODEGA',
        'EDIFICACIONES',
        'DEPRECIACION ACUMULADA',
        'TOTAL PROPIEDAD PLANTA Y EQUIPO',
    }
    blocked_codes = {'1605', '1635', '1640', '1685'}
    subtotal_rows = []
    for r in range(1, ws_inf.max_row + 1):
        code_txt = str(ws_inf.cell(r, 3).value or '').strip()
        label_txt = str(ws_inf.cell(r, 4).value or '').strip().upper()
        if label_txt == 'SUBTOTAL':
            subtotal_rows.append(r)
        if label_txt in blocked_labels or code_txt in blocked_codes:
            for c in range(5, 8):
                ws_inf.cell(r, c, None)

    # Mantener solo el primer SUBTOTAL (familias). Limpiar cualquier subtotal adicional.
    if subtotal_rows:
        keep_row = min(subtotal_rows)
        for r in subtotal_rows:
            if r != keep_row:
                for c in range(5, 8):
                    ws_inf.cell(r, c, None)

    # Forzar celdas sin formula en seccion contable fija.
    for r in [34, 35, 36, 38, 41]:
        ws_inf.cell(r, 5, '')

    # Ocultar formulas en barra de formula y permitir solo edicion de valores de detalle en DESGLOSE.
    # Excel solo oculta formulas cuando la hoja esta protegida y la celda esta "hidden".
    max_des_edit_col = len(columns_order)
    for r in range(1, ws_des.max_row + 1):
        for c in range(1, max_des_edit_col + 1):
            cell = ws_des.cell(r, c)
            cell.protection = Protection(locked=False, hidden=False)
    for cell in des_formula_cells:
        cell.protection = Protection(locked=True, hidden=True)
    ws_des.protection.sheet = True

    for cell in inf_formula_cells:
        cell.protection = Protection(locked=True, hidden=True)
    ws_inf.protection.sheet = True
    ws_base.protection.sheet = True

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.calcMode = 'auto'

    out = BytesIO()
    wb.save(out)
    content = out.getvalue()
    safe_period = sanitize_filename(period_label.replace(' ', '_'))
    filename = f'informe_conciliacion_activos_fijos_contabilidad_{safe_period}.xlsx'

    persist_accounting_report_file(content, filename, period_label, month, year, report_title, period_id=None)

    with ACCOUNTING_CACHE_LOCK:
        ACCOUNTING_REPORT_CACHE['version'] = current_cache_key
        ACCOUNTING_REPORT_CACHE['algo_version'] = ACCOUNTING_REPORT_ALGO_VERSION
        ACCOUNTING_REPORT_CACHE['bytes'] = content
        ACCOUNTING_REPORT_CACHE['filename'] = filename

    return send_file(
        BytesIO(content),
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/reports/accounting_monthly_history', methods=['GET'])
def accounting_monthly_history():
    ensure_db()
    rows = GeneratedReport.query.filter_by(report_type='accounting_monthly').order_by(GeneratedReport.id.desc()).limit(200).all()
    return jsonify({'items': [r.to_dict() for r in rows]})


@app.route('/reports/accounting_monthly_history/<int:report_id>/download', methods=['GET'])
def accounting_monthly_history_download(report_id):
    ensure_db()
    row = GeneratedReport.query.filter_by(id=report_id, report_type='accounting_monthly').first()
    if not row:
        return jsonify({'error': 'Informe no encontrado'}), 404
    if not row.file_path or not os.path.exists(row.file_path):
        return jsonify({'error': 'El archivo no existe en almacenamiento'}), 404
    return send_file(
        row.file_path,
        as_attachment=True,
        download_name=row.file_name or os.path.basename(row.file_path),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/reports/a22_history', methods=['GET'])
def a22_history():
    ensure_db()
    period_id = request.args.get('period_id', type=int)
    q = GeneratedReport.query.filter(
        GeneratedReport.report_type.in_(['a22_excel', 'a22_pdf'])
    )
    if period_id:
        q = q.filter(GeneratedReport.period_id == period_id)
    rows = q.order_by(GeneratedReport.id.desc()).limit(300).all()
    return jsonify({'items': [r.to_dict() for r in rows]})


@app.route('/reports/a22_history/<int:report_id>/download', methods=['GET'])
def a22_history_download(report_id):
    ensure_db()
    row = GeneratedReport.query.filter(
        GeneratedReport.id == report_id,
        GeneratedReport.report_type.in_(['a22_excel', 'a22_pdf'])
    ).first()
    if not row:
        return jsonify({'error': 'Informe A22 no encontrado'}), 404
    if not row.file_path or not os.path.exists(row.file_path):
        return jsonify({'error': 'El archivo no existe en almacenamiento'}), 404
    ext = os.path.splitext(row.file_name or row.file_path)[1].lower()
    mime = 'application/pdf' if ext == '.pdf' else 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return send_file(
        row.file_path,
        as_attachment=True,
        download_name=row.file_name or os.path.basename(row.file_path),
        mimetype=mime
    )


if __name__ == '__main__':
    ensure_db()
    app.run(debug=True)

