# Helper para limpiar nombres de archivo
def clean_filename(text):
    import unicodedata
    text = str(text or '').lower().replace(' ', '_')
    text = unicodedata.normalize('NFD', text)
    text = ''.join(ch for ch in text if unicodedata.category(ch) != 'Mn')
    text = ''.join(c for c in text if c.isalnum() or c in ('_', '-', '.'))
    return text
import os
import shutil
import subprocess
import tempfile
import zipfile
import unicodedata
import json
import re
from difflib import SequenceMatcher
from threading import Lock
from functools import lru_cache
from copy import copy
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime, timezone, timedelta
from io import BytesIO
from zoneinfo import ZoneInfo

# Intentamos resoluciones más robustas de zona horaria en entornos Windows
def _resolve_zoneinfo(name):
    try:
        return ZoneInfo(name)
    except Exception:
        try:
            # python-dateutil normalmente está disponible (dependencia de pandas)
            from dateutil import tz as dateutil_tz

            tzinfo = dateutil_tz.gettz(name)
            if tzinfo:
                return tzinfo
        except Exception:
            pass
        return None


import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, Protection
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from PIL import Image as PILImage
from flask import Flask, render_template, request, jsonify, send_file, has_app_context, session, redirect, url_for, flash, g
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image as RLImage
from reportlab.lib.utils import ImageReader
from reportlab.graphics.shapes import Drawing, String
from reportlab.graphics.charts.barcharts import HorizontalBarChart
from reportlab.graphics.charts.piecharts import Pie
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from sqlalchemy import UniqueConstraint, text
from werkzeug.security import generate_password_hash, check_password_hash
from xml.sax.saxutils import escape

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
APP_TIMEZONE = os.getenv('APP_TIMEZONE', 'America/Bogota')
LOCAL_TZ = _resolve_zoneinfo(APP_TIMEZONE) or timezone.utc
DB_PATH = os.path.join(BASE_DIR, 'assets.db')
REPORTS_DIR = os.path.join(BASE_DIR, 'generated_reports')
DOCUMENTS_DIR = os.path.join(BASE_DIR, 'generated_documents')
TEMPLATE_A22_PATH = os.path.join(BASE_DIR, 'formato a22.xlsx')
ACCOUNTING_TEMPLATE_CANDIDATES = [
    os.path.join(BASE_DIR, 'INFORME CONTABILIDAD REF.xlsx'),
    os.path.join(BASE_DIR, 'INFORME CONTABILIDAD REFERENCIA.xlsx'),
]
FAMILY_CATALOG_PATH = os.path.join(BASE_DIR, 'FAMILIA DE ACTIVOS FIJOS.xlsx')
A22_LOGO_CANDIDATES = [
    os.path.join(BASE_DIR, 'logo_a22.png'),
    os.path.join(BASE_DIR, 'logo_a22.jpg'),
    os.path.join(BASE_DIR, 'logo_a22.jpeg'),
    os.path.join(BASE_DIR, 'logo.png'),
    os.path.join(BASE_DIR, 'logo.jpg'),
]
CODIFICACION_CANDIDATES = [
    os.path.join(BASE_DIR, 'codificacion.png'),
    os.path.join(BASE_DIR, 'codificacion.jpg'),
    os.path.join(BASE_DIR, 'codificacion.jpeg'),
]
TEMPLATES_DIR = os.path.join(BASE_DIR, 'templates')
STATIC_DIR = os.path.join(BASE_DIR, 'static')
DISPOSAL_TYPE_KEYS = ['BIOMEDICO', 'MUEBLE Y ENSER', 'INDUSTRIAL', 'TECNOLOGICO', 'CONTROL']
DISPOSAL_MANUAL_TYPE_OPTIONS = [
    'BIOMEDICO',
    'MUEBLE Y ENSER',
    'INDUSTRIAL',
    'TECNOLOGICO',
    'CONTROL - BIOMEDICO',
    'CONTROL - MUEBLE Y ENSER',
    'CONTROL - INDUSTRIAL',
    'CONTROL - TECNOLOGICO',
]
DOCUMENT_TYPE_OPTIONS = [
    'Salida de almacen',
    'RA recepcion',
    'Contrato',
    'Nota interna',
    'Oficio',
    'Certificacion',
    'Mantenimiento',
    'Baja',
    'Novedad',
    'Otro',
]
ACCOUNTING_FAMILY_ORDER = """
5504
5504001
5504002
5504003
5506
5506001
5511
5511001
5523
5523001
6002
6002001
6002002
6002003
6002004
6002005
6003
6003001
6003002
6005
6005001
6006
6006001
6006002
6006003
6006004
6006005
6006006
6006007
6006008
6006009
6006010
6006011
6007
6007001
6007002
6007003
6007004
6007005
6007006
6007007
6007008
6007009
6007010
6007011
6007012
6007013
6007015
6008
6008001
6008002
6008003
6008004
6008005
6008006
6008007
6501
6501001
6501002
6501003
6501004
6501005
6501006
6501007
6501008
6501009
6501010
6501011
6501012
6501013
6501014
6501015
6501016
6501017
6501018
6501019
6501020
6501021
6501022
6501023
6501024
6501025
6501026
6501027
6501028
6501029
6501030
6501031
6501032
6501033
6501034
6501035
6501036
6501037
6501038
6502
6502002
6502003
6502004
7001
7001001
7001002
7001003
7001004
7001005
7002
7002001
7002002
7002003
7002004
7002005
7002006
7003
7003001
7003002
7003003
7003004
7003005
7003006
7003007
7003008
7004
7004001
7502
7502001
7506
7506001
8002
8002001
8002002
8002003
8002004
8004
8004001
8004002
""".split()

ACCOUNTING_CACHE_LOCK = Lock()
BASE_DATA_VERSION = 0
ACCOUNTING_REPORT_CACHE = {
    'version': None,
    'algo_version': None,
    'bytes': None,
    'filename': None,
}


def now_local_dt():
    return datetime.now(LOCAL_TZ)


def now_iso():
    return now_local_dt().isoformat()


def parse_dt(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        dt = value
    else:
        raw = str(value).strip()
        if not raw:
            return None
        try:
            dt = datetime.fromisoformat(raw.replace('Z', '+00:00'))
        except Exception:
            return None
    # Si el datetime no tiene tzinfo, asumir que fue ingresado en la zona local
    # (APP_TIMEZONE). Esto evita interpretar horas locales como UTC y desplazar
    # la hora al convertir.
    if dt.tzinfo is None:
        try:
            dt = dt.replace(tzinfo=LOCAL_TZ)
        except Exception:
            dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(LOCAL_TZ)


def format_dt_local(value, fmt='%Y-%m-%d %H:%M'):
    dt = parse_dt(value)
    if not dt:
        return ''
    return dt.strftime(fmt)
ACCOUNTING_REPORT_ALGO_VERSION = 19
ACCOUNTING_EXCLUDED_FAMILIES = {'1114001', '4619001'}
STRICT_ACCOUNTING_VALIDATION = True
# Costos fijos definidos por contabilidad (reemplazo manual por C_ACT).
# Este diccionario aplica igual para todos los meses.
ACCOUNTING_COST_OVERRIDES = {
    '4978':  Decimal('3326296.00'),
    '6449': Decimal('11424000.33'),
    '6450': Decimal('11424000.33'),
    '6451': Decimal('11424000.33'),
    '6478':  Decimal('1062077.61'),
    '6575':  Decimal('1945569.12'),
    '6690':  Decimal('1903999.20'),
    '6925':  Decimal('297192.00'),
    '5140':  Decimal('1043997.72'),
    '5570':  Decimal('1411934.73'),
    '3681':  Decimal('3363478.00'),
    '2638':  Decimal('803333.41'),
}
_DEC_ZERO = Decimal('0')
_DEC_TWO = Decimal('0.01')
_YELLOW_FILL_DES = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
MONTH_LABELS_ES = {
    1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril', 5: 'mayo', 6: 'junio',
    7: 'julio', 8: 'agosto', 9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
}
EXCLUDED_SERVICE_NAMES = {
    'HOSPITAL FRANCISCO DE PAULA',
}
ISSUE_STATUSES = ['Nuevo', 'En analisis', 'Escalado', 'Cerrado']
ISSUE_SEVERITIES = ['Alta', 'Media', 'Baja']
TRANSFER_STATUSES = ['Pendiente aprobacion', 'Aprobado', 'Rechazado', 'Ejecutado']
ISSUE_TYPE_LABELS = {
    'NOT_FOUND_CRITICAL': 'No encontrado critico',
    'NOT_FOUND_HIGH_VALUE': 'No encontrado de alto valor',
    'SCANNED_OTHER_SERVICE': 'Escaneado en servicio distinto',
    'RESPONSIBLE_REVIEW': 'Revision de responsable',
    'LOCATION_REVIEW': 'Revision de ubicacion',
    'DUPLICATE_CODE': 'Duplicidad probable de codigo',
    'MISSING_SERIAL_REF': 'Sin serial/referencia',
    'MISSING_MODEL_BRAND': 'Falta marca/modelo',
    'MISSING_CUSTODY_DATA': 'Falta responsable/ubicacion',
    'PENDING_UNSCANNED': 'Activo pendiente sin escaneo',
    'INVALID_FINANCIAL_VALUES': 'Valores financieros inconsistentes',
    'DEPRECIATION_INCONSISTENT': 'Depreciacion/vida util inconsistente',
    'CANDIDATE_DISPOSAL': 'Riesgo por baja pendiente',
}
ASSET_ASSIST_MAX_IMAGE_MB = 12
ASSET_ASSIST_ALLOWED_MIME = {
    'image/jpeg', 'image/jpg', 'image/png', 'image/webp', 'image/bmp'
}
ASSET_ASSIST_OCR_MIN_TOKEN_SIZE = 3
ASSET_ASSIST_OCR_FIELDS = [
    'NOM', 'NOM_FAM', 'DESC_TIAC', 'DES_SUBTIAC', 'MODELO', 'SERIE', 'REF', 'NOM_MARCA'
]
ASSET_ASSIST_CATEGORY_RULES = [
    {
        'key': 'biomedico',
        'label': 'Biomedico',
        'keywords': [
            'biomed', 'medic', 'hospital', 'monitor de signos', 'signos vitales', 'multiparametro',
            'desfibrilador', 'bomba de infusion', 'ventilador', 'oximetro', 'pulsoximetro', 'electro', 'ecografo',
        ],
        'model_labels': [],
        'exclude_keywords': ['escritorio', 'mesa oficina', 'cpu', 'pc', 'impresora'],
    },
    {
        'key': 'mueble_enser',
        'label': 'Mueble y Enser',
        'keywords': [
            'mueble', 'enser', 'mesa', 'silla', 'atril', 'escritorio', 'archivador', 'gabinete',
            'estante', 'locker', 'camilla', 'vitrina', 'modulo',
        ],
        'model_labels': [],
        'exclude_keywords': ['bomba infusion', 'desfibrilador', 'ventilador'],
    },
    {
        'key': 'industrial',
        'label': 'Industrial',
        'keywords': [
            'industrial', 'planta', 'compresor', 'tablero', 'caldera', 'motor', 'generador',
            'transformador', 'subestacion', 'chiller',
        ],
        'model_labels': [],
        'exclude_keywords': ['monitor signos', 'portatil', 'escritorio'],
    },
    {
        'key': 'tecnologico',
        'label': 'Tecnologico',
        'keywords': [
            'tecnolog', 'computador', 'cpu', 'pc', 'all in one', 'portatil', 'laptop', 'teclado',
            'mouse', 'impresora', 'scanner', 'router', 'switch', 'comunicacion',
        ],
        'model_labels': [],
        'exclude_keywords': ['signos vitales', 'desfibrilador', 'camilla'],
    },
]
ACCOUNTING_REPORT_STRUCTURE = [
    {
        'parent_code': '1655',
        'parent_name': 'MAQUINARIA Y EQUIPO',
        'children': [
            {'report_code': '165504', 'name': 'MAQUINARIA INDUSTRIAL', 'source_prefix': '5504'},
            {'report_code': '165506', 'name': 'EQ DE RECREACION Y DEPORTES', 'source_prefix': '5506'},
            {'report_code': '165511', 'name': 'HERRAMIENTAS Y ACCESORIOS', 'source_prefix': '5511'},
            {'report_code': '165523', 'name': 'EQUIPO DE ASEO', 'source_prefix': '5523'},
        ],
    },
    {
        'parent_code': '1660',
        'parent_name': 'EQUIPO MEDICO Y CIENTIFICO',
        'children': [
            {'report_code': '166002', 'name': 'EQUIPO DE LABORATORIO', 'source_prefix': '6002'},
            {'report_code': '166003', 'name': 'EQUIPO DE URGENCIAS', 'source_prefix': '6003'},
            {'report_code': '166005', 'name': 'EQUIPO DE HOSPITALIZACION', 'source_prefix': '6005'},
            {'report_code': '166006', 'name': 'EQUIPO DE CX Y SALA DE PARTOS', 'source_prefix': '6006'},
            {'report_code': '166007', 'name': 'EQUIPO DE APOYO DIAGNOSTICO', 'source_prefix': '6007'},
            {'report_code': '166008', 'name': 'EQUIPO DE APOYO TERAPEUTICO', 'source_prefix': '6008'},
        ],
    },
    {
        'parent_code': '1665',
        'parent_name': 'MUEBLES, ENSERES Y EQUIPOS DE OFICINA',
        'children': [
            {'report_code': '166501', 'name': 'MUEBLES Y ENSERES', 'source_prefix': '6501'},
            {'report_code': '166502', 'name': 'EQUIPOS Y MAQUINAS DE OFICINA', 'source_prefix': '6502'},
        ],
    },
    {
        'parent_code': '1670',
        'parent_name': 'EQUIPO DE COMUNICACION Y COMPUTACION',
        'children': [
            {'report_code': '167001', 'name': 'EQUIPO DE COMUNICACION', 'source_prefixes': ['7001', '7004']},
            {'report_code': '167002', 'name': 'EQUIPO DE COMPUTACION', 'source_prefixes': ['7002', '7003']},
        ],
    },
    {
        'parent_code': '1675',
        'parent_name': 'EQUIPO DE TRANSPORTE Y TRACCION',
        'children': [
            {'report_code': '1675002', 'name': 'EQUIPO TERRESTRE', 'source_prefix': '7502'},
            {'report_code': '167506', 'name': 'EQUIPO DE TRACCION', 'source_prefix': '7506'},
        ],
    },
    {
        'parent_code': '1680',
        'parent_name': 'EQUIPO COMEDOR, DESPENSA Y COCINA',
        'children': [
            {'report_code': '168002', 'name': 'EQUIPO DE RESTAURANTE Y CAFETERIA', 'source_prefix': '8002'},
            {'report_code': '168004', 'name': 'EQUIPO DE LAVANDERIA', 'source_prefix': '8004'},
        ],
    },
]

app = Flask(__name__, template_folder=TEMPLATES_DIR, static_folder=STATIC_DIR)
CORS(app)
app.config['SECRET_KEY'] = os.getenv('SEII_SECRET_KEY', 'seii-dev-secret-change-me')
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=12)
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_PATH}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)


class Asset(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    c_act = db.Column(db.String, unique=True, nullable=False)
    nom = db.Column(db.String)
    modelo = db.Column(db.String)
    ref = db.Column(db.String)
    serie = db.Column(db.String)
    nom_marca = db.Column(db.String)
    c_fam = db.Column(db.String)
    nom_fam = db.Column(db.String)
    c_tiac = db.Column(db.String)
    desc_tiac = db.Column(db.String)
    desc_subtiac = db.Column(db.String)
    deprecia = db.Column(db.String)
    vida_util = db.Column(db.String)
    tipo_activo_cache = db.Column(db.String)
    des_ubi = db.Column(db.String)
    nom_ccos = db.Column(db.String)
    nom_resp = db.Column(db.String)
    est = db.Column(db.String)
    costo = db.Column(db.Float)
    saldo = db.Column(db.Float)
    fecha_compra = db.Column(db.String)
    codigo_inteligente = db.Column(db.String)
    subtipo_codigo = db.Column(db.String)
    color = db.Column(db.String)
    nit_proveedor = db.Column(db.String)
    desc_proveedor = db.Column(db.String)
    forma_adquisicion = db.Column(db.String)
    en_garantia = db.Column(db.String)
    entidad_garantia = db.Column(db.String)
    garantia_desde = db.Column(db.String)
    garantia_hasta = db.Column(db.String)
    agencia = db.Column(db.String)
    centro_costo_code = db.Column(db.String)

    # campos de inventario
    estado_inventario = db.Column(db.String, default='No verificado')
    fecha_verificacion = db.Column(db.String)
    usuario_verificador = db.Column(db.String)
    observacion_inventario = db.Column(db.String)
    raw_row_json = db.Column(db.Text)

    def to_dict(self):
        return {
            'id': self.id,
            'C_ACT': self.c_act,
            'NOM': self.nom,
            'MODELO': self.modelo,
            'REF': self.ref,
            'SERIE': self.serie,
            'NOM_MARCA': self.nom_marca,
            'C_FAM': self.c_fam,
            'NOM_FAM': self.nom_fam,
            'C_TIAC': self.c_tiac,
            'DESC_TIAC': self.desc_tiac,
            'DES_SUBTIAC': self.desc_subtiac,
            'DEPRECIA': self.deprecia,
            'VIDA_UTIL': self.vida_util,
            'TIPO_ACTIVO': self.tipo_activo_cache,
            'DES_UBI': self.des_ubi,
            'NOM_CCOS': self.nom_ccos,
            'NOM_RESP': self.nom_resp,
            'EST': self.est,
            'COSTO': self.costo,
            'SALDO': self.saldo,
            'FECHA_COMPRA': self.fecha_compra,
            'CODIGO_INTELIGENTE': self.codigo_inteligente,
            'SUBTIPO_CODIGO': self.subtipo_codigo,
            'COLOR': self.color,
            'NIT_PROVEEDOR': self.nit_proveedor,
            'DESCRIPCION_PROVEEDOR': self.desc_proveedor,
            'FORMA_ADQUISICION': self.forma_adquisicion,
            'EN_GARANTIA': self.en_garantia,
            'ENTIDAD_GARANTIA': self.entidad_garantia,
            'GARANTIA_DESDE': self.garantia_desde,
            'GARANTIA_HASTA': self.garantia_hasta,
            'AGENCIA': self.agencia,
            'CENTRO_COSTO_CODIGO': self.centro_costo_code,
            'estado_inventario': self.estado_inventario,
            'fecha_verificacion': self.fecha_verificacion,
            'usuario_verificador': self.usuario_verificador,
            'observacion_inventario': self.observacion_inventario,
        }


class InventoryRun(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String, nullable=False)
    period_id = db.Column(db.Integer, db.ForeignKey('inventory_period.id'))
    service = db.Column(db.String)
    service_scope_json = db.Column(db.String)
    status = db.Column(db.String, default='active', nullable=False)
    started_at = db.Column(db.String, nullable=False)
    closed_at = db.Column(db.String)
    created_by = db.Column(db.String)
    cancelled_at = db.Column(db.String)
    cancelled_by = db.Column(db.String)
    cancel_reason = db.Column(db.String)

    def to_dict(self):
        scope = []
        if self.service_scope_json:
            try:
                parsed = json.loads(self.service_scope_json)
                if isinstance(parsed, list):
                    scope = [str(x).strip() for x in parsed if str(x or '').strip()]
            except Exception:
                scope = []
        if (not scope) and self.service:
            scope = [str(self.service).strip()]
        return {
            'id': self.id,
            'name': self.name,
            'period_id': self.period_id,
            'service': self.service,
            'service_scope': scope,
            'service_scope_count': len(scope),
            'service_scope_label': ', '.join(scope[:3]) + (' ...' if len(scope) > 3 else ''),
            'status': self.status,
            'started_at': self.started_at,
            'started_at_local': format_dt_local(self.started_at),
            'closed_at': self.closed_at,
            'closed_at_local': format_dt_local(self.closed_at),
            'created_by': self.created_by,
            'cancelled_at': self.cancelled_at,
            'cancelled_at_local': format_dt_local(self.cancelled_at),
            'cancelled_by': self.cancelled_by,
            'cancel_reason': self.cancel_reason,
        }


class InventoryPeriod(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String, unique=True, nullable=False)
    period_type = db.Column(db.String, nullable=False, default='semestral')
    start_date = db.Column(db.String)
    end_date = db.Column(db.String)
    status = db.Column(db.String, nullable=False, default='open')
    notes = db.Column(db.String)
    created_at = db.Column(db.String, nullable=False)
    cancelled_at = db.Column(db.String)
    cancelled_by = db.Column(db.String)
    cancel_reason = db.Column(db.String)

    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'period_type': self.period_type,
            'start_date': self.start_date,
            'end_date': self.end_date,
            'status': self.status,
            'notes': self.notes,
            'created_at': self.created_at,
            'created_at_local': format_dt_local(self.created_at),
            'cancelled_at': self.cancelled_at,
            'cancelled_at_local': format_dt_local(self.cancelled_at),
            'cancelled_by': self.cancelled_by,
            'cancel_reason': self.cancel_reason,
        }


class RunAssetStatus(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    run_id = db.Column(db.Integer, db.ForeignKey('inventory_run.id'), nullable=False)
    asset_id = db.Column(db.Integer, db.ForeignKey('asset.id'), nullable=False)
    status = db.Column(db.String, nullable=False)
    scanned_at = db.Column(db.String, nullable=False)
    scanned_by = db.Column(db.String)

    __table_args__ = (
        UniqueConstraint('run_id', 'asset_id', name='uq_run_asset'),
    )


class AssetDisposal(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    asset_id = db.Column(db.Integer, db.ForeignKey('asset.id'), nullable=False, unique=True)
    period_id = db.Column(db.Integer, db.ForeignKey('inventory_period.id'))
    status = db.Column(db.String, nullable=False, default='Pendiente baja')
    reason = db.Column(db.String)
    requested_by = db.Column(db.String)
    requested_at = db.Column(db.String, nullable=False)
    reviewed_by = db.Column(db.String)
    reviewed_at = db.Column(db.String)
    review_notes = db.Column(db.String)

    def to_dict(self, asset=None):
        base = {
            'id': self.id,
            'asset_id': self.asset_id,
            'period_id': self.period_id,
            'status': self.status,
            'reason': self.reason,
            'requested_by': self.requested_by,
            'requested_at': self.requested_at,
            'requested_at_local': format_dt_local(self.requested_at),
            'reviewed_by': self.reviewed_by,
            'reviewed_at': self.reviewed_at,
            'reviewed_at_local': format_dt_local(self.reviewed_at),
            'review_notes': self.review_notes,
        }
        if asset:
            base['asset'] = asset.to_dict()
        return base


class SystemMeta(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    meta_key = db.Column(db.String, unique=True, nullable=False)
    meta_value = db.Column(db.String)


class UserAccount(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String, unique=True, nullable=False)
    full_name = db.Column(db.String)
    email = db.Column(db.String, unique=True)
    password_hash = db.Column(db.String, nullable=False)
    is_admin = db.Column(db.Boolean, nullable=False, default=False)
    is_active = db.Column(db.Boolean, nullable=False, default=True)
    created_at = db.Column(db.String, nullable=False)
    last_login_at = db.Column(db.String)

    def set_password(self, raw_password):
        self.password_hash = generate_password_hash(str(raw_password or ''))

    def check_password(self, raw_password):
        return check_password_hash(self.password_hash or '', str(raw_password or ''))

    def to_dict(self):
        return {
            'id': self.id,
            'username': self.username,
            'full_name': self.full_name or '',
            'email': self.email or '',
            'is_admin': bool(getattr(self, 'is_admin', False)),
            'is_active': bool(self.is_active),
            'created_at': self.created_at,
            'created_at_local': format_dt_local(self.created_at),
            'last_login_at': self.last_login_at,
            'last_login_at_local': format_dt_local(self.last_login_at),
        }


class GeneratedReport(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    report_type = db.Column(db.String, nullable=False)
    title = db.Column(db.String, nullable=False)
    period_id = db.Column(db.Integer, db.ForeignKey('inventory_period.id'))
    period_label = db.Column(db.String)
    accounting_base_id = db.Column(db.Integer, db.ForeignKey('accounting_monthly_base.id'))
    overrides_summary_json = db.Column(db.Text)
    file_name = db.Column(db.String, nullable=False)
    file_path = db.Column(db.String, nullable=False)
    generated_at = db.Column(db.String, nullable=False)

    def to_dict(self):
        overrides_summary = []
        if self.overrides_summary_json:
            try:
                parsed = json.loads(self.overrides_summary_json)
                if isinstance(parsed, list):
                    overrides_summary = parsed
            except Exception:
                overrides_summary = []
        return {
            'id': self.id,
            'report_type': self.report_type,
            'title': self.title,
            'period_id': self.period_id,
            'period_label': self.period_label,
            'accounting_base_id': self.accounting_base_id,
            'overrides_summary': overrides_summary,
            'file_name': self.file_name,
            'generated_at': self.generated_at,
            'generated_at_local': format_dt_local(self.generated_at),
        }


class AccountingMonthlyBase(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    period_year = db.Column(db.Integer, nullable=False)
    period_month = db.Column(db.Integer, nullable=False)
    period_label = db.Column(db.String, nullable=False)
    source_file_name = db.Column(db.String, nullable=False)
    source_file_path = db.Column(db.String, nullable=False)
    uploaded_by = db.Column(db.String)
    uploaded_at = db.Column(db.String, nullable=False)
    asset_count = db.Column(db.Integer, nullable=False, default=0)
    status = db.Column(db.String, nullable=False, default='active')

    def to_dict(self):
        return {
            'id': self.id,
            'period_year': self.period_year,
            'period_month': self.period_month,
            'period_label': self.period_label,
            'source_file_name': self.source_file_name,
            'uploaded_by': self.uploaded_by or '',
            'uploaded_at': self.uploaded_at,
            'uploaded_at_local': format_dt_local(self.uploaded_at),
            'asset_count': int(self.asset_count or 0),
            'status': self.status,
        }


class AccountingMonthlyBaseAsset(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    base_id = db.Column(db.Integer, db.ForeignKey('accounting_monthly_base.id'), nullable=False)
    c_act = db.Column(db.String, nullable=False)
    c_fam = db.Column(db.String)
    nom_fam = db.Column(db.String)
    costo = db.Column(db.Float)
    saldo = db.Column(db.Float)
    raw_row_json = db.Column(db.Text)

    __table_args__ = (
        UniqueConstraint('base_id', 'c_act', name='uq_accounting_base_asset'),
    )


class AssetIssue(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    issue_type = db.Column(db.String, nullable=False)
    title = db.Column(db.String, nullable=False)
    severity = db.Column(db.String, nullable=False, default='Media')
    status = db.Column(db.String, nullable=False, default='Nuevo')
    source = db.Column(db.String, nullable=False, default='auto')
    period_id = db.Column(db.Integer, db.ForeignKey('inventory_period.id'))
    run_id = db.Column(db.Integer, db.ForeignKey('inventory_run.id'))
    asset_id = db.Column(db.Integer, db.ForeignKey('asset.id'))
    service = db.Column(db.String)
    detected_value = db.Column(db.Float, default=0.0)
    description = db.Column(db.String)
    assigned_to = db.Column(db.String)
    due_date = db.Column(db.String)
    resolution_notes = db.Column(db.String)
    created_at = db.Column(db.String, nullable=False)
    updated_at = db.Column(db.String, nullable=False)

    def to_dict(self):
        asset = Asset.query.get(self.asset_id) if self.asset_id else None
        return {
            'id': self.id,
            'issue_type': self.issue_type,
            'issue_type_label': ISSUE_TYPE_LABELS.get(self.issue_type, self.issue_type),
            'title': self.title,
            'severity': self.severity,
            'status': self.status,
            'source': self.source,
            'period_id': self.period_id,
            'run_id': self.run_id,
            'asset_id': self.asset_id,
            'asset_code': asset.c_act if asset else '',
            'asset_name': asset.nom if asset else '',
            'service': self.service or (asset.nom_ccos if asset else ''),
            'detected_value': to_number(self.detected_value),
            'description': self.description or '',
            'assigned_to': self.assigned_to or '',
            'due_date': self.due_date or '',
            'resolution_notes': self.resolution_notes or '',
            'created_at': self.created_at,
            'created_at_local': format_dt_local(self.created_at),
            'updated_at': self.updated_at,
            'updated_at_local': format_dt_local(self.updated_at),
        }


class AssetTransferCase(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    issue_id = db.Column(db.Integer, db.ForeignKey('asset_issue.id'))
    asset_id = db.Column(db.Integer, db.ForeignKey('asset.id'), nullable=False)
    period_id = db.Column(db.Integer, db.ForeignKey('inventory_period.id'))
    run_id = db.Column(db.Integer, db.ForeignKey('inventory_run.id'))
    status = db.Column(db.String, nullable=False, default='Pendiente aprobacion')
    origin_service = db.Column(db.String)
    target_service = db.Column(db.String)
    origin_responsible = db.Column(db.String)
    target_responsible = db.Column(db.String)
    justification = db.Column(db.String)
    requested_by = db.Column(db.String)
    requested_at = db.Column(db.String)
    approved_by = db.Column(db.String)
    approved_at = db.Column(db.String)
    approval_notes = db.Column(db.String)
    executed_by = db.Column(db.String)
    executed_at = db.Column(db.String)
    execution_notes = db.Column(db.String)
    acta_doc_id = db.Column(db.Integer, db.ForeignKey('document_record.id'))
    acta_file_path = db.Column(db.String)
    created_at = db.Column(db.String, nullable=False)
    updated_at = db.Column(db.String, nullable=False)

    def to_dict(self):
        asset = Asset.query.get(self.asset_id) if self.asset_id else None
        issue = AssetIssue.query.get(self.issue_id) if self.issue_id else None
        return {
            'id': self.id,
            'issue_id': self.issue_id,
            'issue_type': issue.issue_type if issue else '',
            'issue_title': issue.title if issue else '',
            'asset_id': self.asset_id,
            'asset_code': asset.c_act if asset else '',
            'asset_name': asset.nom if asset else '',
            'period_id': self.period_id,
            'run_id': self.run_id,
            'status': self.status,
            'origin_service': self.origin_service or '',
            'target_service': self.target_service or '',
            'origin_responsible': self.origin_responsible or '',
            'target_responsible': self.target_responsible or '',
            'justification': self.justification or '',
            'requested_by': self.requested_by or '',
            'requested_at': self.requested_at,
            'requested_at_local': format_dt_local(self.requested_at),
            'approved_by': self.approved_by or '',
            'approved_at': self.approved_at,
            'approved_at_local': format_dt_local(self.approved_at),
            'approval_notes': self.approval_notes or '',
            'executed_by': self.executed_by or '',
            'executed_at': self.executed_at,
            'executed_at_local': format_dt_local(self.executed_at),
            'execution_notes': self.execution_notes or '',
            'acta_doc_id': self.acta_doc_id,
            'acta_available': bool(self.acta_doc_id),
            'acta_download_url': f"/transfers/{self.id}/acta" if self.acta_doc_id else '',
            'created_at': self.created_at,
            'created_at_local': format_dt_local(self.created_at),
            'updated_at': self.updated_at,
            'updated_at_local': format_dt_local(self.updated_at),
        }


class DocumentRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    link_type = db.Column(db.String, nullable=False, default='general')  # general | asset
    asset_id = db.Column(db.Integer, db.ForeignKey('asset.id'))
    asset_code = db.Column(db.String)
    asset_name = db.Column(db.String)
    document_type = db.Column(db.String, nullable=False)
    title = db.Column(db.String, nullable=False)
    description = db.Column(db.String)
    doc_date = db.Column(db.String)
    area_service = db.Column(db.String)
    radicado = db.Column(db.String)
    file_name = db.Column(db.String, nullable=False)
    file_path = db.Column(db.String, nullable=False)
    file_ext = db.Column(db.String)
    file_size = db.Column(db.Integer, default=0)
    uploaded_by = db.Column(db.String)
    uploaded_at = db.Column(db.String, nullable=False)
    status = db.Column(db.String, nullable=False, default='active')

    def to_dict(self):
        return {
            'id': self.id,
            'link_type': self.link_type,
            'asset_id': self.asset_id,
            'asset_code': self.asset_code or '',
            'asset_name': self.asset_name or '',
            'document_type': self.document_type or '',
            'title': self.title or '',
            'description': self.description or '',
            'doc_date': self.doc_date or '',
            'area_service': self.area_service or '',
            'radicado': self.radicado or '',
            'file_name': self.file_name or '',
            'file_ext': self.file_ext or '',
            'file_size': int(self.file_size or 0),
            'uploaded_by': self.uploaded_by or '',
            'uploaded_at': self.uploaded_at,
            'uploaded_at_local': format_dt_local(self.uploaded_at),
            'status': self.status,
        }


def ensure_db():
    os.makedirs(REPORTS_DIR, exist_ok=True)
    os.makedirs(DOCUMENTS_DIR, exist_ok=True)
    if has_app_context():
        db.create_all()
        ensure_schema_updates()
        backfill_asset_life_sheet_fields()
        assign_legacy_period_to_old_runs()
        assign_legacy_period_to_old_disposals()
        ensure_default_user()
        ensure_admin_user_exists()
    else:
        with app.app_context():
            db.create_all()
            ensure_schema_updates()
            backfill_asset_life_sheet_fields()
            assign_legacy_period_to_old_runs()
            assign_legacy_period_to_old_disposals()
            ensure_default_user()
            ensure_admin_user_exists()


def ensure_default_user():
    if UserAccount.query.count() > 0:
        return
    username = str(os.getenv('SEII_DEFAULT_USER', 'admin')).strip() or 'admin'
    default_password = str(os.getenv('SEII_DEFAULT_PASSWORD', 'admin123'))
    full_name = str(os.getenv('SEII_DEFAULT_FULLNAME', 'Administrador SEII')).strip() or 'Administrador SEII'
    row = UserAccount(
        username=username,
        full_name=full_name,
        email=None,
        is_admin=True,
        is_active=True,
        created_at=now_iso(),
    )
    row.set_password(default_password)
    db.session.add(row)
    db.session.commit()
    app.logger.warning('[AUTH] Usuario inicial creado: %s', username)


def ensure_admin_user_exists():
    admin_exists = UserAccount.query.filter_by(is_admin=True, is_active=True).count() > 0
    if admin_exists:
        return
    first_user = UserAccount.query.order_by(UserAccount.id.asc()).first()
    if not first_user:
        return
    first_user.is_admin = True
    if not str(first_user.created_at or '').strip():
        first_user.created_at = now_iso()
    db.session.commit()
    app.logger.warning('[AUTH] Se promovio a administrador inicial: %s', first_user.username)


def get_current_user():
    cached = getattr(g, '_current_user', None)
    if cached is not None:
        return cached
    user_id = session.get('user_id')
    if not user_id:
        g._current_user = None
        return None
    row = UserAccount.query.get(int(user_id))
    if not row or not row.is_active:
        session.clear()
        g._current_user = None
        return None
    g._current_user = row
    return row


def get_actor_username(fallback='system'):
    user = get_current_user()
    if user and str(user.username or '').strip():
        return str(user.username).strip()
    fb = str(fallback or '').strip()
    return fb or 'system'


def is_current_user_admin():
    user = get_current_user()
    return bool(user and getattr(user, 'is_admin', False))


def require_admin_or_403():
    if is_current_user_admin():
        return None
    if request.path.startswith('/admin/users/api'):
        return jsonify({'error': 'Acceso restringido a administradores'}), 403
    accept = str(request.headers.get('Accept') or '').lower()
    if request.method == 'GET' and request.path.startswith('/admin/') and 'text/html' in accept:
        flash('No tienes permisos para acceder al panel de administracion.', 'error')
        return redirect(url_for('index'))
    return jsonify({'error': 'Acceso restringido a administradores'}), 403


def login_user_session(user):
    session.permanent = True
    session['user_id'] = int(user.id)
    session['username'] = str(user.username or '')
    session['login_at'] = now_iso()


def logout_user_session():
    session.clear()


AUTH_PAGE_PATHS = {
    '/',
    '/inventario',
    '/jornadas',
    '/formatos',
    '/bajas',
    '/dashboard',
    '/informes',
    '/cronograma',
    '/novedades',
    '/hoja_vida',
    '/documentos',
    '/admin/usuarios',
}


AUTH_EXEMPT_ENDPOINTS = {
    'static',
    'login_page',
    'login_submit',
    'logout',
    'logo_file',
}


@app.before_request
def enforce_login():
    endpoint = request.endpoint or ''
    if endpoint in AUTH_EXEMPT_ENDPOINTS or endpoint.startswith('static'):
        return None
    user = get_current_user()
    if user:
        g.current_user = user
        return None
    if request.method == 'GET' and request.path in AUTH_PAGE_PATHS:
        next_url = request.full_path if request.query_string else request.path
        if next_url.endswith('?'):
            next_url = next_url[:-1]
        return redirect(url_for('login_page', next=next_url))
    return jsonify({'error': 'Debes iniciar sesion', 'auth_required': True}), 401


@app.context_processor
def inject_auth_context():
    return {'current_user': get_current_user()}


def invalidate_accounting_report_cache():
    global BASE_DATA_VERSION
    with ACCOUNTING_CACHE_LOCK:
        BASE_DATA_VERSION += 1
        ACCOUNTING_REPORT_CACHE['version'] = None
        ACCOUNTING_REPORT_CACHE['algo_version'] = None
        ACCOUNTING_REPORT_CACHE['bytes'] = None
        ACCOUNTING_REPORT_CACHE['filename'] = None


def get_system_meta(meta_key, default=None):
    row = SystemMeta.query.filter_by(meta_key=meta_key).first()
    if not row:
        return default
    return row.meta_value if row.meta_value is not None else default


def set_system_meta(meta_key, meta_value):
    row = SystemMeta.query.filter_by(meta_key=meta_key).first()
    if not row:
        row = SystemMeta(meta_key=meta_key)
        db.session.add(row)
    row.meta_value = str(meta_value)


def bump_assets_revision():
    raw = get_system_meta('assets_revision', '0')
    try:
        current = int(str(raw).strip())
    except Exception:
        current = 0
    next_value = current + 1
    set_system_meta('assets_revision', str(next_value))
    db.session.commit()
    return next_value


def get_assets_revision():
    raw = get_system_meta('assets_revision', '0')
    try:
        return int(str(raw).strip())
    except Exception:
        return 0


def ensure_schema_updates():
    with db.engine.begin() as conn:
        columns = {row[1] for row in conn.execute(text('PRAGMA table_info(asset)')).fetchall()}
        if 'c_fam' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN c_fam VARCHAR'))
        if 'nom_fam' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN nom_fam VARCHAR'))
        if 'desc_subtiac' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN desc_subtiac VARCHAR'))
        if 'deprecia' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN deprecia VARCHAR'))
        if 'vida_util' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN vida_util VARCHAR'))
        if 'tipo_activo_cache' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN tipo_activo_cache VARCHAR'))
        if 'raw_row_json' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN raw_row_json TEXT'))
        if 'codigo_inteligente' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN codigo_inteligente VARCHAR'))
        if 'subtipo_codigo' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN subtipo_codigo VARCHAR'))
        if 'color' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN color VARCHAR'))
        if 'nit_proveedor' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN nit_proveedor VARCHAR'))
        if 'desc_proveedor' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN desc_proveedor VARCHAR'))
        if 'forma_adquisicion' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN forma_adquisicion VARCHAR'))
        if 'en_garantia' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN en_garantia VARCHAR'))
        if 'entidad_garantia' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN entidad_garantia VARCHAR'))
        if 'garantia_desde' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN garantia_desde VARCHAR'))
        if 'garantia_hasta' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN garantia_hasta VARCHAR'))
        if 'agencia' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN agencia VARCHAR'))
        if 'centro_costo_code' not in columns:
            conn.execute(text('ALTER TABLE asset ADD COLUMN centro_costo_code VARCHAR'))

        run_columns = {row[1] for row in conn.execute(text('PRAGMA table_info(inventory_run)')).fetchall()}
        if 'period_id' not in run_columns:
            conn.execute(text('ALTER TABLE inventory_run ADD COLUMN period_id INTEGER'))
        if 'service_scope_json' not in run_columns:
            conn.execute(text('ALTER TABLE inventory_run ADD COLUMN service_scope_json VARCHAR'))
        if 'cancelled_at' not in run_columns:
            conn.execute(text('ALTER TABLE inventory_run ADD COLUMN cancelled_at VARCHAR'))
        if 'cancelled_by' not in run_columns:
            conn.execute(text('ALTER TABLE inventory_run ADD COLUMN cancelled_by VARCHAR'))
        if 'cancel_reason' not in run_columns:
            conn.execute(text('ALTER TABLE inventory_run ADD COLUMN cancel_reason VARCHAR'))

        period_columns = {row[1] for row in conn.execute(text('PRAGMA table_info(inventory_period)')).fetchall()}
        if 'cancelled_at' not in period_columns:
            conn.execute(text('ALTER TABLE inventory_period ADD COLUMN cancelled_at VARCHAR'))
        if 'cancelled_by' not in period_columns:
            conn.execute(text('ALTER TABLE inventory_period ADD COLUMN cancelled_by VARCHAR'))
        if 'cancel_reason' not in period_columns:
            conn.execute(text('ALTER TABLE inventory_period ADD COLUMN cancel_reason VARCHAR'))

        report_columns = {row[1] for row in conn.execute(text('PRAGMA table_info(generated_report)')).fetchall()}
        if 'period_id' not in report_columns:
            conn.execute(text('ALTER TABLE generated_report ADD COLUMN period_id INTEGER'))
        if 'accounting_base_id' not in report_columns:
            conn.execute(text('ALTER TABLE generated_report ADD COLUMN accounting_base_id INTEGER'))
        if 'overrides_summary_json' not in report_columns:
            conn.execute(text('ALTER TABLE generated_report ADD COLUMN overrides_summary_json TEXT'))

        disposal_columns = {row[1] for row in conn.execute(text('PRAGMA table_info(asset_disposal)')).fetchall()}
        if 'period_id' not in disposal_columns:
            conn.execute(text('ALTER TABLE asset_disposal ADD COLUMN period_id INTEGER'))

        user_columns = {row[1] for row in conn.execute(text('PRAGMA table_info(user_account)')).fetchall()}
        if 'is_admin' not in user_columns:
            conn.execute(text('ALTER TABLE user_account ADD COLUMN is_admin BOOLEAN DEFAULT 0'))
        if 'is_active' not in user_columns:
            conn.execute(text('ALTER TABLE user_account ADD COLUMN is_active BOOLEAN DEFAULT 1'))
        if 'created_at' not in user_columns:
            conn.execute(text("ALTER TABLE user_account ADD COLUMN created_at VARCHAR DEFAULT ''"))
        if 'last_login_at' not in user_columns:
            conn.execute(text('ALTER TABLE user_account ADD COLUMN last_login_at VARCHAR'))


def assign_legacy_period_to_old_runs():
    orphan_runs = InventoryRun.query.filter(InventoryRun.period_id.is_(None)).count()
    if orphan_runs <= 0:
        return
    legacy = get_or_create_default_period()
    InventoryRun.query.filter(InventoryRun.period_id.is_(None)).update({'period_id': legacy.id})
    db.session.commit()


def assign_legacy_period_to_old_disposals():
    orphan_disposals = AssetDisposal.query.filter(AssetDisposal.period_id.is_(None)).count()
    if orphan_disposals <= 0:
        return
    legacy = get_or_create_default_period()
    AssetDisposal.query.filter(AssetDisposal.period_id.is_(None)).update({'period_id': legacy.id})
    db.session.commit()


def backfill_asset_life_sheet_fields():
    if str(get_system_meta('life_sheet_backfill_v1', '0')).strip() == '1':
        return
    rows = Asset.query.all()
    changed = 0
    for asset in rows:
        payload = asset_raw_payload(asset)

        def set_if_empty(attr_name, candidate):
            nonlocal changed
            current = getattr(asset, attr_name, None)
            if str(current or '').strip():
                return
            val = str(candidate or '').strip()
            if not val:
                return
            setattr(asset, attr_name, val)
            changed += 1

        set_if_empty('codigo_inteligente', _pick_first_value(payload, [
            'CODINTELIGENTE', 'CODIGO_INTELIGENTE', 'COD_INTELIGENTE', 'CODIGO INTELIGENTE',
        ]))
        set_if_empty('subtipo_codigo', _pick_first_value(payload, [
            'SUBTIPO', 'SUBTIPO_ACTIVO', 'COD_SUBTIPO', 'COD_SUBTIPO_ACTIVO',
        ]))
        set_if_empty('color', _pick_first_value(payload, ['COLOR', 'COLORES']))
        set_if_empty('nit_proveedor', _pick_first_value(payload, ['NIT_PROVEEDOR', 'NIT PROVEEDOR', 'NIT']))
        set_if_empty('desc_proveedor', _pick_first_value(payload, ['PROVEEDOR', 'DESCRIPCION_PROVEEDOR', 'DESCRIPCION DEL PROVEEDOR']))
        set_if_empty('forma_adquisicion', _pick_first_value(payload, ['FORMA_ADQUISICION', 'FORMA DE ADQUISICION', 'ADQUISICION']))
        set_if_empty('en_garantia', _pick_first_value(payload, ['EN_GARANTIA', 'GARANTIA']))
        set_if_empty('entidad_garantia', _pick_first_value(payload, ['ENTIDAD', 'ENTIDAD_GARANTIA']))
        set_if_empty('garantia_desde', _pick_first_value(payload, ['GARANTIA_DESDE', 'DESDE']))
        set_if_empty('garantia_hasta', _pick_first_value(payload, ['GARANTIA_HASTA', 'HASTA']))
        set_if_empty('agencia', _pick_first_value(payload, ['AGENCIA']))
        set_if_empty('centro_costo_code', _pick_first_value(payload, ['C_CCOS', 'CENTRO_COSTO', 'COD_CENTRO_COSTO']))

    set_system_meta('life_sheet_backfill_v1', '1')
    db.session.commit()


def normalize_columns(cols):
    out = {}
    for c in cols:
        raw = str(c or '').strip()
        if not raw:
            continue
        out[raw.upper()] = c
        out[normalize_lookup_key(raw)] = c
    return out


def normalize_lookup_key(value):
    txt = str(value or '').strip().upper()
    if not txt:
        return ''
    txt = unicodedata.normalize('NFD', txt)
    txt = ''.join(ch for ch in txt if unicodedata.category(ch) != 'Mn')
    normalized = []
    last_underscore = False
    for ch in txt:
        if ch.isalnum():
            normalized.append(ch)
            last_underscore = False
        else:
            if not last_underscore:
                normalized.append('_')
                last_underscore = True
    return ''.join(normalized).strip('_')


def is_excluded_service_name(value):
    txt = str(value or '').strip().upper()
    return txt in EXCLUDED_SERVICE_NAMES


def parse_bool(value, default=False):
    if isinstance(value, bool):
        return value
    if value is None:
        return default
    txt = str(value).strip().lower()
    if txt in {'1', 'true', 'si', 'sí', 'yes', 'on'}:
        return True
    if txt in {'0', 'false', 'no', 'off', ''}:
        return False
    return default


def parse_int(value, default=None):
    try:
        if value in (None, ''):
            return default
        return int(value)
    except Exception:
        return default


def normalize_service_name(value):
    txt = str(value or '').strip()
    if not txt:
        return ''
    return txt


def normalize_service_scope(values):
    if values is None:
        return []
    if not isinstance(values, list):
        values = [values]
    result = []
    seen = set()
    for item in values:
        svc = normalize_service_name(item)
        if not svc or is_excluded_service_name(svc):
            continue
        key = svc.casefold()
        if key in seen:
            continue
        seen.add(key)
        result.append(svc)
    return result


def run_scope_services(run):
    if not run:
        return []
    scope = []
    raw = getattr(run, 'service_scope_json', None)
    if raw:
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, list):
                scope = parsed
        except Exception:
            scope = []
    scope = normalize_service_scope(scope)
    if (not scope) and getattr(run, 'service', None):
        scope = normalize_service_scope([run.service])
    return scope


def apply_run_scope_filter(query, run):
    scope = run_scope_services(run)
    if scope:
        query = query.filter(Asset.nom_ccos.in_(scope))
    return query


def get_cell(row, cols_map, key):
    """Retorna el valor de la columna mapeada o None si no existe."""
    candidates = [
        str(key or '').strip(),
        str(key or '').strip().upper(),
        normalize_lookup_key(key),
    ]
    for cand in candidates:
        if not cand:
            continue
        if cand in cols_map:
            try:
                v = row[cols_map[cand]]
                if pd.isna(v):
                    return None
                return v
            except Exception:
                return None
    return None


def get_cell_first(row, cols_map, keys):
    for key in keys:
        value = get_cell(row, cols_map, key)
        if value is not None and str(value).strip().lower() != 'nan':
            return value
    return None


def is_non_depreciable(value):
    if value is None:
        return False

    raw = str(value).strip()
    if not raw:
        return False

    # Si viene numérico (ej: 0, 0.0), aplica regla directa.
    numeric = raw.replace(',', '.')
    try:
        return float(numeric) <= 0
    except Exception:
        pass

    txt = raw.upper()
    txt = unicodedata.normalize('NFD', txt)
    txt = ''.join(ch for ch in txt if unicodedata.category(ch) != 'Mn')
    true_tokens = {'SI', 'S', 'TRUE', '1', 'Y', 'YES'}
    false_tokens = {'NO', 'N', 'FALSE', '0', '0.0', '0.00'}
    if txt in false_tokens:
        return True
    if txt in true_tokens:
        return False
    if 'NO' in txt and 'DEPREC' in txt:
        return True
    if 'SIN DEPREC' in txt:
        return True
    return False


def is_zero_useful_life(value):
    if value is None:
        return False
    txt = str(value).strip()
    if not txt:
        return False
    txt = txt.replace(',', '.')
    try:
        return float(txt) <= 0
    except Exception:
        return txt.upper() in {'CERO', 'SIN VIDA UTIL', 'NO APLICA', 'N/A'}


